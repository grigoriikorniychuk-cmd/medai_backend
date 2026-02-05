import asyncio
import json
import os
import shutil
import logging
import traceback
from datetime import datetime, timedelta, timezone
from typing import Dict, Any, List, Optional, Tuple, Union
from uuid import uuid4
from urllib.parse import urlparse
import io
import csv
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from fastapi import HTTPException

# Matplotlib настройка
import matplotlib

matplotlib.use("Agg")  # Для работы без GUI
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.colors import LinearSegmentedColormap
import numpy as np
import pandas as pd
import seaborn as sns
from bson.objectid import ObjectId
import calendar
from collections import defaultdict
import tempfile
import matplotlib.ticker as mticker

# AmoCRM клиент
from mlab_amo_async.amocrm_client import AsyncAmoCRMClient

# ReportLab импорты для PDF
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
    PageBreak,
    Flowable,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Openpyxl imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Внутренние импорты
from ..settings.paths import DATA_DIR
from ..settings.auth import get_mongodb
from ..models.call_report import ReportType, CallReportData, CallReportRequest, AnalyticsReportRequest, VisualizationType, ReportFormat
from ..services.mongodb_service import mongodb_service

# Настройка логирования
logger = logging.getLogger(__name__)


class CallReportService:
    """Сервис for генерации отчетов по звонкам"""

    def __init__(self):
        """Инициализация сервиса отчетов"""
        self.mongo_client = get_mongodb()
        self.db = self.mongo_client["medai"]
        self.mongodb_service = mongodb_service

        # Создаем директории для отчетов и временных файлов
        self.reports_dir = os.path.join(DATA_DIR, "reports")
        self.temp_dir = os.path.join(self.reports_dir, "temp")
        os.makedirs(self.reports_dir, exist_ok=True)
        os.makedirs(self.temp_dir, exist_ok=True)

    def _register_fonts(self):
        """Регистрирует шрифты с поддержкой кириллицы для PDF-отчётов"""
        try:
            # Пути к шрифтам
            font_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "fonts"
            )

            # Проверяем и регистрируем шрифты
            dejavu_sans_path = os.path.join(font_dir, "DejaVuSans.ttf")
            dejavu_sans_bold_path = os.path.join(font_dir, "DejaVuSans-Bold.ttf")

            if not os.path.exists(dejavu_sans_path) or not os.path.exists(
                dejavu_sans_bold_path
            ):
                logger.warning(f"Шрифты not найдены in {font_dir}")
                return "Helvetica", "Helvetica-Bold"

            # Регистрируем шрифты
            pdfmetrics.registerFont(TTFont("DejaVuSans", dejavu_sans_path))
            pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", dejavu_sans_bold_path))
            pdfmetrics.registerFontFamily(
                "DejaVuSans", normal="DejaVuSans", bold="DejaVuSans-Bold"
            )

            logger.info("Шрифты DejaVu Sans успешно зарегистрированы")
            return "DejaVuSans", "DejaVuSans-Bold"

        except Exception as e:
            logger.error(f"Ошибка при регистрации шрифтов: {e}")
            return "Helvetica", "Helvetica-Bold"

    def _convert_date_format(self, date_str: str) -> str:
        """Преобразует дату для формата DD.MM.YYYY в YYYY-MM-DD"""
        if not date_str or "." not in date_str:
            return date_str

        day, month, year = date_str.split(".")
        return f"{year}-{month}-{day}"

    async def debug_lead_data(
        self, lead_id: str, client_id: str
    ) -> Tuple[str, Optional[Dict]]:
        """Получает and анализирует данные сделки для отладки"""
        if not lead_id or not client_id:
            return "Неизвестно", None

        try:
            client = AsyncAmoCRMClient(
                client_id=client_id,
                client_secret="",
                subdomain="",
                redirect_url="",
                mongo_uri="mongodb://localhost:27017/",
                db_name="medai",
            )

            lead = await client.get_lead(lead_id)
            await client.close()

            if not lead:
                return "Неизвестно", None

            # Парсинг источника для названия сделки
            source = "Неизвестно"
            if "name" in lead:
                lead_name = lead.get("name", "")
                source_patterns = [
                    "Заявка с лендинга",
                    "Заявка с сайта",
                    "Заявка с формы",
                    "Заявка",
                ]

                for pattern in source_patterns:
                    if pattern in lead_name:
                        source_info = lead_name[lead_name.find(pattern) :]
                        if "http" in source_info:
                            url_start = source_info.find("http")
                            url_end = (
                                source_info.find(" ", url_start)
                                if source_info.find(" ", url_start) > -1
                                else len(source_info)
                            )
                            source = source_info[url_start:url_end]
                        else:
                            source = source_info
                        break

            return source, lead
        except Exception as e:
            logger.error(f"Ошибка при получении данных сделки {lead_id}: {e}")
            return "Неизвестно", None

    async def update_traffic_sources_in_analytics(self):
        """Обновляет источники трафика в записях аналитики звонков"""
        try:
            # Находим записи с неизвестным источником и lead_id
            records = (
                await self.db.call_analytics.find(
                    {
                        "client.source": "Unknown",
                        "lead_id": {"$exists": True, "$ne": None},
                    }
                )
                .limit(50)
                .to_list(length=50)
            )

            updated_count = 0
            for record in records:
                lead_id = record.get("lead_id")
                client_id = record.get("client_id", "")

                # Получаем данные сделки
                source, _ = await self.debug_lead_data(lead_id, client_id)

                # Если источник определен, обновляем запись
                if source != "Неизвестно":
                    result = await self.db.call_analytics.update_one(
                        {"_id": record.get("_id")}, {"$set": {"client.source": source}}
                    )

                    if result.modified_count > 0:
                        updated_count += 1

            logger.info(f"Обновлено {updated_count} источников трафика")
        except Exception as e:
            logger.error(f"Ошибка при обновлении источников трафика: {e}")
            
    async def get_calls_data(
        self,
        start_date: str,
        end_date: str,
        clinic_id: Optional[str] = None,
        administrator_ids: Optional[List[str]] = None,
    ) -> List[Dict[str, Any]]:
        """
        Получает данные о звонках из коллекции calls в MongoDB
        
        :param start_date: Начальная дата в формате YYYY-MM-DD
        :param end_date: Конечная дата в формате YYYY-MM-DD
        :param clinic_id: ID клиники (опционально)
        :param administrator_ids: Список ID администраторов (опционально)
        :return: Список звонков с данными
        """
        try:
            # Формируем запрос для фильтрации звонков
            query = {}
            
            # Добавляем фильтр, чтобы получать только звонки с анализом
            query["analysis"] = {"$exists": True}
            
            # Фильтр по дате
            if start_date and end_date:
                # Преобразуем строковые даты в datetime объекты
                start_datetime = datetime.strptime(f"{start_date} 00:00:00", "%Y-%m-%d %H:%M:%S")
                end_datetime = datetime.strptime(f"{end_date} 23:59:59", "%Y-%m-%d %H:%M:%S")
                
                # Добавляем фильтр по recorded_at (для новых записей) или created_at (для старых)
                query["$or"] = [
                    {"recorded_at": {"$gte": start_datetime, "$lte": end_datetime}},
                    {"created_at": {"$gte": int(start_datetime.timestamp()), "$lte": int(end_datetime.timestamp())}}
                ]
            
            # Фильтр по клинике
            if clinic_id:
                query["client_id"] = clinic_id
            
            # Фильтр по администраторам
            if administrator_ids and len(administrator_ids) > 0:
                # Преобразуем строковые ID в числовые, если необходимо
                numeric_admin_ids = []
                for admin_id in administrator_ids:
                    try:
                        numeric_admin_ids.append(int(admin_id))
                    except (ValueError, TypeError):
                        pass
                
                # Добавляем фильтр по administrator_id или amocrm_user_id
                if numeric_admin_ids:
                    query["$or"] = query.get("$or", []) + [
                        {"amocrm_user_id": {"$in": numeric_admin_ids}},
                        {"responsible_user_id": {"$in": numeric_admin_ids}}
                    ]
                
                # Добавляем фильтр по имени администратора (если ID не числовые)
                str_admin_ids = [admin_id for admin_id in administrator_ids if isinstance(admin_id, str)]
                if str_admin_ids:
                    query["$or"] = query.get("$or", []) + [
                        {"administrator": {"$in": str_admin_ids}}
                    ]
            
            # Получаем данные звонков из MongoDB
            calls = await self.db.calls.find(query).sort("recorded_at", -1).to_list(length=None)
            
            # Формируем словарь администраторов для логирования
            admins_dict = {}
            
            # Формируем список звонков с необходимыми данными
            processed_calls = []
            for call in calls:
                # Сохраняем имя администратора в словарь
                admin_name = call.get("administrator", "Неизвестный администратор")
                admin_id = call.get("amocrm_user_id") or call.get("responsible_user_id", "")
                
                if admin_id and admin_name != "Неизвестный администратор":
                    admins_dict[str(admin_id)] = admin_name
                
                # Преобразуем данные в нужный формат
                call_data = {
                    "id": str(call.get("_id", "")),
                    "administrator_id": call.get("amocrm_user_id", "") or call.get("responsible_user_id", ""),
                    "administrator_name": admin_name,
                    "timestamp": call.get("recorded_at", datetime.now()).isoformat(),
                    "duration": call.get("duration", 0),
                    "call_type": {
                        "direction": "incoming" if call.get("call_direction") == "Входящий" else "outgoing",
                    },
                    "client": {
                        "source": call.get("source", "Неизвестно"),
                        "phone": call.get("phone", "")
                    },
                    "lead_id": call.get("lead_id", ""),
                    "client_id": call.get("client_id", ""),
                    "metrics": {
                        "overall_score": 0  # У нас нет этих данных в коллекции calls, ставим 0
                    },
                    "conversion": False  # У нас нет этих данных в коллекции calls, ставим False
                }
                
                processed_calls.append(call_data)
            
            # Логируем найденные уникальные имена администраторов
            if admins_dict:
                logger.info(f"Найдены имена администраторов из коллекции calls: {admins_dict}")
            
            return processed_calls
        except Exception as e:
            logger.error(f"Ошибка при получении данных о звонках: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return []

    async def generate_report(self, request: CallReportRequest) -> Dict[str, Any]:
        """Генерация отчета по звонкам"""
        if request.report_type == ReportType.FULL:
            return await self.generate_full_report(request)
        elif request.report_type == ReportType.SUMMARY:
            return await self.generate_summary_report(request)
        elif request.report_type == ReportType.INDIVIDUAL:
            return await self.generate_individual_report(request)
        elif request.report_type == ReportType.ANALYTICS:
            analytics_request = AnalyticsReportRequest(
                start_date=request.start_date,
                end_date=request.end_date,
                clinic_id=request.clinic_id,
                administrator_ids=request.administrator_ids,
                visualization_type="category_distribution",
                include_raw_data=False,
                format="pdf"
            )
            return await self.generate_analytics_report(analytics_request)
        else:
            raise HTTPException(status_code=400, detail=f"Неподдерживаемый тип отчета: {request.report_type}")

    async def generate_full_report(self, request: CallReportRequest) -> Dict[str, Any]:
        """
        Генерирует отчет по звонкам

        :param start_date: Начальная дата в формате DD.MM.YYYY
        :param end_date: Конечная дата в формате DD.MM.YYYY
        :param report_type: Тип отчета
        :param clinic_id: ID клиники
        :param administrator_ids: Список ID администраторов
        :return: Кортеж from путей к PDF and Excel файлам и данных отчета
        """
        # Отладка источников трафика
        debug_leads = False  # Включаем только для отладки
        if debug_leads:
            try:
                debug_records = (
                    await self.db.call_analytics.find(
                        {
                            "client.source": "Unknown",
                            "lead_id": {"$exists": True, "$ne": None},
                        }
                    )
                    .sort("timestamp", -1)
                    .limit(3)
                    .to_list(length=3)
                )

                for record in debug_records:
                    lead_id = record.get("lead_id")
                    client_id = record.get("client_id", "")
                    source, lead_data = await self.debug_lead_data(lead_id, client_id)

                    if source != "Неизвестно" and lead_data:
                        # Сохраняем данные for анализа
                        debug_file = os.path.join(
                            self.temp_dir, f"lead_{lead_id}_debug.json"
                        )
                        with open(debug_file, "w", encoding="utf-8") as f:
                            json.dump(lead_data, f, ensure_ascii=False, indent=2)

                        # Обновляем запись
                        await self.db.call_analytics.update_one(
                            {"_id": record.get("_id")},
                            {"$set": {"client.source": source}},
                        )
            except Exception as e:
                logger.error(f"Ошибка with отладке источников трафика: {e}")

        try:
            # Преобразуем даты для MongoDB and создаем временную директорию
            start_date_db = self._convert_date_format(request.start_date)
            end_date_db = self._convert_date_format(request.end_date)

            report_id = str(uuid4())
            report_temp_dir = os.path.join(self.temp_dir, report_id)
            os.makedirs(report_temp_dir, exist_ok=True)

            # Получаем название клиники
            clinic_name = "Клиника not определена"
            found_clinic = None

            if request.clinic_id:
                try:
                    # Создаем список возможных ID форматов
                    possible_ids = [request.clinic_id]
                    if isinstance(request.clinic_id, str):
                        try:
                            possible_ids.append(ObjectId(request.clinic_id))
                        except:
                            pass

                    # Поиск по всем возможным форматам
                    for possible_id in possible_ids:
                        clinic = await self.db.clinics.find_one({"_id": possible_id})
                        if clinic:
                            found_clinic = clinic
                            break

                        clinic = await self.db.clinics.find_one(
                            {"client_id": possible_id}
                        )
                        if clinic:
                            found_clinic = clinic
                            break

                    # Извлекаем название, если клиника найдена
                    if found_clinic and "name" in found_clinic:
                        clinic_name = found_clinic.get("name")
                    else:
                        # Дополнительная проверка через call_analytics
                        analytics = await self.db.call_analytics.find_one(
                            {"clinic_id": request.clinic_id}
                        )
                        if analytics and analytics.get("clinic_id"):
                            clinic = await self.db.clinics.find_one(
                                {"client_id": analytics.get("clinic_id")}
                            )
                            if clinic and "name" in clinic:
                                clinic_name = clinic.get("name")
                except Exception as e:
                    logger.error(f"Ошибка with получении информации o клинике: {e}")
            
            # Получаем данные о звонках из обоих источников
            analytics_data = await mongodb_service.get_call_analytics(
                start_date=start_date_db,
                end_date=end_date_db,
                clinic_id=request.clinic_id,
                administrator_ids=request.administrator_ids,
                analysis_exists=True
            )
            
            calls_data = await self.get_calls_data(
                start_date=start_date_db,
                end_date=end_date_db,
                clinic_id=request.clinic_id,
                administrator_ids=request.administrator_ids,
            )
            
            # Объединяем данные из двух источников, приоритизируя информацию из calls_data
            # Создаем словарь администраторов для быстрого поиска
            admin_map = {}
            for call in calls_data:
                admin_id = call.get("administrator_id")
                admin_name = call.get("administrator_name")
                if admin_id and admin_name and admin_name != "Неизвестный администратор":
                    admin_map[str(admin_id)] = admin_name
            
            # Создаем словарь источников трафика
            source_map = {}
            for call in calls_data:
                lead_id = call.get("lead_id")
                source = call.get("client", {}).get("source")
                if lead_id and source and source not in ["Неизвестно", "Unknown", None]:
                    source_map[str(lead_id)] = source
            
            # Дополнительно проверяем источники из базы UTM меток
            for call in analytics_data:
                lead_id = call.get("lead_id")
                if lead_id and str(lead_id) not in source_map:
                    try:
                        source, _ = await self.debug_lead_data(str(lead_id), call.get("client_id", ""))
                        if source and source not in ["Неизвестно", "Unknown", None]:
                            source_map[str(lead_id)] = source
                    except Exception as e:
                        logger.debug(f"Не удалось получить источник для lead_id={lead_id}: {e}")
            
            # Обогащаем данные из call_analytics информацией из calls и обновляем источник
            for call in analytics_data:
                # Обновляем имя администратора
                admin_id = str(call.get("administrator_id", ""))
                if admin_id and admin_id in admin_map:
                    call["administrator_name"] = admin_map[admin_id]
                
                # Обновляем источник трафика
                lead_id = call.get("lead_id")
                if lead_id and str(lead_id) in source_map:
                    if "client" not in call:
                        call["client"] = {}
                    call["client"]["source"] = source_map[str(lead_id)]
                elif "client" in call and call["client"].get("source") in ["Unknown", "Неизвестно", None]:
                    # Если источник неизвестен, попробуем найти его через lead_id
                    try:
                        source, _ = await self.debug_lead_data(str(lead_id), call.get("client_id", ""))
                        if source and source not in ["Неизвестно", "Unknown", None]:
                            call["client"]["source"] = source
                    except Exception as e:
                        logger.debug(f"Не удалось получить источник для обновления call_analytics: {e}")
            
            # Добавляем данные о длительности звонков из calls в analytics_data
            duration_map = {}
            for call in calls_data:
                call_id = call.get("id")
                duration = call.get("duration", 0)
                if call_id and duration > 0:
                    duration_map[call_id] = duration
            
            # Добавляем длительность звонков в данные аналитики
            for call in analytics_data:
                call_id = str(call.get("_id", ""))
                if call_id in duration_map:
                    call["duration"] = duration_map[call_id]
            
            # Обогащенные данные аналитики
            enriched_analytics = analytics_data + [call for call in calls_data if call.get("id") not in [str(a.get("_id", "")) for a in analytics_data]]
            
            # Получаем метрики звонков с уточнением для calls
            metrics = await mongodb_service.calculate_call_metrics(
                start_date=start_date_db,
                end_date=end_date_db,
                clinic_id=request.clinic_id,
                administrator_ids=request.administrator_ids,
            )
            
            # Дополняем метрики данными из коллекции calls
            # Рассчитываем общую и среднюю длительность звонков
            total_duration = 0
            valid_duration_count = 0
            
            for call in calls_data:
                duration = call.get("duration", 0)
                if duration > 0:
                    total_duration += duration
                    valid_duration_count += 1
            
            # Обновляем метрики
            avg_duration = total_duration / valid_duration_count if valid_duration_count > 0 else 0
            metrics["avg_duration"] = avg_duration
            metrics["total_duration"] = total_duration
            
            # Также собираем статистику по источникам трафика из calls
            sources_stats = {}
            for call in calls_data:
                source = call.get("client", {}).get("source", "Неизвестно")
                if source == "Unknown":
                    source = "Неизвестно"
                
                if source not in sources_stats:
                    sources_stats[source] = {
                        "count": 0, 
                        "name": source,
                        "incoming_calls": 0,
                        "outgoing_calls": 0
                    }
                
                sources_stats[source]["count"] += 1
                
                # Определяем направление звонка
                direction = call.get("call_type", {}).get("direction", "")
                if direction == "incoming":
                    sources_stats[source]["incoming_calls"] += 1
                elif direction == "outgoing":
                    sources_stats[source]["outgoing_calls"] += 1
            
            # Обновляем метрики источников трафика из данных calls
            for source, stats in sources_stats.items():
                if "traffic_sources" not in metrics:
                    metrics["traffic_sources"] = {}
                
                if source not in metrics["traffic_sources"]:
                    metrics["traffic_sources"][source] = stats
                else:
                    # Объединяем данные, если источник уже есть
                    metrics["traffic_sources"][source]["count"] += stats["count"]
                    metrics["traffic_sources"][source]["incoming_calls"] = stats["incoming_calls"]
                    metrics["traffic_sources"][source]["outgoing_calls"] = stats["outgoing_calls"]
                    metrics["traffic_sources"][source]["name"] = stats["name"]
            
            # Группируем звонки по администраторам из calls
            admin_stats = {}
            # Счетчик общего количества звонков из коллекции calls
            total_calls_from_calls = 0
            for call in calls_data:
                total_calls_from_calls += 1
                admin_name = call.get("administrator_name", "Неизвестный администратор")
                admin_id = call.get("administrator_id", "")
                
                if not admin_id:
                    continue
                
                admin_key = str(admin_id)
                if admin_key not in admin_stats:
                    admin_stats[admin_key] = {
                        "id": admin_id,
                        "name": admin_name,
                        "total_calls": 0,
                        "incoming_calls": 0,
                        "outgoing_calls": 0
                    }
                
                admin_stats[admin_key]["total_calls"] += 1
                direction = call.get("call_type", {}).get("direction", "")
                if direction == "incoming":
                    admin_stats[admin_key]["incoming_calls"] += 1
                elif direction == "outgoing":
                    admin_stats[admin_key]["outgoing_calls"] += 1

            # Кэш для хранения имен администраторов
            admin_names_cache = {}
            admin_ids_to_query = []

            # Собираем ID администраторов для запроса
            for admin in metrics.get("administrators", []):
                admin_id = admin.get("id", "")
                if admin_id:
                    admin_ids_to_query.append(admin_id)
                    if isinstance(admin_id, str):
                        try:
                            admin_ids_to_query.append(ObjectId(admin_id))
                        except:
                            pass

            # Заполняем кэш имен администраторов
            if admin_ids_to_query:
                try:
                    admin_docs = await self.db.administrators.find(
                        {
                            "$or": [
                                {"_id": {"$in": admin_ids_to_query}},
                                {"amocrm_user_id": {"$in": admin_ids_to_query}},
                            ]
                        }
                    ).to_list(length=None)

                    for admin_doc in admin_docs:
                        admin_id = str(admin_doc.get("_id", ""))
                        amocrm_user_id = admin_doc.get("amocrm_user_id", "")
                        admin_name = admin_doc.get("name", "Неизвестный администратор")

                        admin_names_cache[admin_id] = admin_name
                        if amocrm_user_id:
                            admin_names_cache[amocrm_user_id] = admin_name
                except Exception as e:
                    logger.error(f"Ошибка with получении имен администраторов: {e}")

            # Формируем список администраторов s данными
            administrators = []
            for admin in metrics.get("administrators", []):
                admin_id = admin.get("id", "")
                name = admin_names_cache.get(
                    admin_id, admin.get("name", "Неизвестный администратор")
                )
                
                # Улучшаем с данными из calls, если есть
                admin_key = str(admin_id)
                if admin_key in admin_stats:
                    admin["name"] = admin_stats[admin_key]["name"]
                    name = admin["name"]
                    
                    # Обновляем количество звонков, если в calls больше данных
                    if admin_stats[admin_key]["total_calls"] > admin.get("total_calls", 0):
                        admin["total_calls"] = admin_stats[admin_key]["total_calls"]
                        admin["incoming_calls"] = admin_stats[admin_key]["incoming_calls"]
                        admin["outgoing_calls"] = admin_stats[admin_key]["outgoing_calls"]

                administrators.append(
                    {
                        "id": admin_id,
                        "name": name,
                        "total_calls": admin.get("total_calls", 0),
                        "incoming_calls": admin.get("incoming_calls", 0),
                        "outgoing_calls": admin.get("outgoing_calls", 0),
                        "incoming_conversion": admin.get("incoming_conversion", 0),
                        "outgoing_conversion": admin.get("outgoing_conversion", 0),
                        "overall_conversion": admin.get("overall_conversion", 0),
                        "avg_score": admin.get("avg_score", 0),
                    }
                )
            
            # Добавляем администраторов из calls, которых нет в metrics
            for admin_key, admin_data in admin_stats.items():
                if not any(str(admin.get("id", "")) == admin_key for admin in administrators):
                    administrators.append({
                        "id": admin_data["id"],
                        "name": admin_data["name"],
                        "total_calls": admin_data["total_calls"],
                        "incoming_calls": admin_data["incoming_calls"],
                        "outgoing_calls": admin_data["outgoing_calls"],
                        "incoming_conversion": 0,
                        "outgoing_conversion": 0,
                        "overall_conversion": 0,
                        "avg_score": 0,
                    })
            
            # Проверяем и корректируем общее количество звонков
            total_calls = metrics.get("total_calls", 0)
            # Если данные из calls_data показывают больше звонков, используем их
            if total_calls_from_calls > total_calls:
                metrics["total_calls"] = total_calls_from_calls
                # Также корректируем входящие/исходящие звонки
                incoming_calls = sum(call.get("call_type", {}).get("direction") == "incoming" for call in calls_data)
                outgoing_calls = sum(call.get("call_type", {}).get("direction") == "outgoing" for call in calls_data)
                metrics["incoming_calls"] = incoming_calls
                metrics["outgoing_calls"] = outgoing_calls
            
            # Получаем детальные данные o звонках
            calls_data = []
            if request.report_type in [ReportType.FULL, ReportType.INDIVIDUAL]:
                # Запускаем фоновое обновление источников трафика
                asyncio.create_task(self.update_traffic_sources_in_analytics())

                # Получаем данные звонков
                analytics = await mongodb_service.get_call_analytics(
                    start_date=start_date_db,
                    end_date=end_date_db,
                    clinic_id=request.clinic_id,
                    administrator_ids=request.administrator_ids,
                    analysis_exists=True
                )

                # Кэш для информации o сделках
                leads_cache = {}

                # Преобразуем данные в формат for отчета
                for call in analytics:
                    # Формируем дату and время
                    timestamp = call.get("timestamp", "")
                    date_str = ""
                    time_str = ""
                    if timestamp and isinstance(timestamp, str):
                        try:
                            date_obj = datetime.fromisoformat(timestamp)
                            date_str = date_obj.strftime("%d.%m.%Y")
                            time_str = date_obj.strftime("%H:%M:%S")
                        except ValueError:
                            date_str = str(timestamp)
                    elif timestamp and isinstance(timestamp, datetime):
                            date_str = timestamp.strftime("%d.%m.%Y")
                            time_str = timestamp.strftime("%H:%M:%S")
                    else:
                            date_str = str(timestamp) if timestamp else ""
                            time_str = ""

                    # Формируем строку критериев
                    script_compliance = call.get("script_compliance", {})
                    criteria = "".join(
                        "✅" if value else "❌"
                        for key, value in script_compliance.items()
                    )

                    # Получаем FG% from overall_score
                    fg_percent = call.get("metrics", {}).get("overall_score", 0) * 10

                    # Имя администратора из кэша
                    admin_id = call.get("administrator_id", "")
                    admin_name = admin_names_cache.get(
                        admin_id,
                        call.get("administrator_name", "Неизвестный администратор"),
                    )

                    # Получаем источник трафика
                    source = call.get("client", {}).get("source", "Неизвестно")

                    # Если источник неизвестен, пробуем получить для сделки
                    lead_id = call.get("lead_id")
                    if (source == "Unknown" or source == "Неизвестно") and lead_id:
                        # Сначала проверяем кэш
                        if lead_id in leads_cache:
                            lead_data = leads_cache[lead_id]
                        else:
                            # Получаем данные сделки через API
                            try:
                                client = AsyncAmoCRMClient(
                                    client_id=call.get("client_id", ""),
                                    client_secret="",
                                    subdomain="",
                                    redirect_url="",
                                    mongo_uri="mongodb://localhost:27017/",
                                    db_name="medai",
                                )

                                lead = await client.get_lead(lead_id)
                                leads_cache[lead_id] = lead
                                lead_data = lead
                                await client.close()
                            except Exception as e:
                                logger.error(
                                    f"Не удалось получить данные сделки {lead_id}: {e}"
                                )
                                lead_data = None

                        # Извлекаем источник из названия сделки
                        if lead_data and "name" in lead_data:
                            lead_name = lead_data.get("name", "")
                            source_patterns = [
                                "Заявка с лендинга",
                                "Заявка с сайта",
                                "Заявка с формы",
                                "Заявка",
                            ]

                            for pattern in source_patterns:
                                if pattern in lead_name:
                                    source_info = lead_name[lead_name.find(pattern) :]
                                    if "http" in source_info:
                                        url_start = source_info.find("http")
                                        url_end = len(source_info)
                                        for delim in [" ", ")", ",", ";"]:
                                            pos = source_info.find(delim, url_start)
                                            if pos > -1 and pos > url_start:
                                                url_end = min(url_end, pos)
                                        source = source_info[url_start:url_end]
                                    else:
                                        source = source_info
                                    break

                            # Асинхронно обновляем запись for будущих отчетов
                            if source != "Неизвестно":
                                asyncio.create_task(
                                    self.db.call_analytics.update_one(
                                        {"_id": call.get("_id")},
                                        {"$set": {"client.source": source}},
                                    )
                                )

                    # Сокращаем длинные URL
                    if isinstance(source, str) and len(source) > 30:
                        source_parts = source.split("/")
                        if len(source_parts) > 2:
                            source = f"{source_parts[2]}/.../{source_parts[-1]}"

                    # Добавляем звонок in отчет
                    calls_data.append(
                        {
                            "id": str(call.get("_id", "")),
                            "date": date_str,
                            "time": time_str,
                            "administrator_name": admin_name,
                            "call_type": (
                                "Входящий"
                                if call.get("call_type", {}).get("direction")
                                == "incoming"
                                else "Исходящий"
                            ),
                            "call_category": call.get("call_type", {}).get(
                                "category_name", ""
                            ),
                            "source": source,
                            "fg_percent": fg_percent,
                            "criteria": criteria,
                            "conversion": call.get("conversion", False),
                            "comment": (
                                ", ".join(call.get("recommendations", []))[:100]
                                if call.get("recommendations")
                                else ""
                            ),
                            "crm_link": call.get("links", {}).get("amocrm", ""),
                            "transcription_link": call.get("links", {}).get(
                                "transcription", ""
                            ),
                        }
                    )

            # Статистика по типам звонков
            call_types = []
            for type_id, data in metrics.get("call_types", {}).items():
                total = metrics.get("total_calls", 0)
                percentage = (data.get("count", 0) / total * 100) if total > 0 else 0

                call_types.append(
                    {
                        "type_id": int(type_id),
                        "name": data.get("name", ""),
                        "count": data.get("count", 0),
                        "percentage": round(percentage, 1),
                        "conversion_rate": round(data.get("conversion_rate", 0), 1),
                    }
                )

            # Сортируем по количеству (сначала больше)
            call_types.sort(key=lambda x: x["count"], reverse=True)

            # Статистика по источникам трафика
            sources = []
            for source in metrics.get("traffic_sources", {}).items():
                source_id = source[0]
                source_data = source[1]
                total = metrics.get("total_calls", 0)
                count = source_data.get("count", 0)
                percentage = (count / total * 100) if total > 0 else 0

                sources.append(
                    {
                        "id": source_id,
                        "source": source_data.get("name", "Неизвестный источник"),
                        "count": count,
                        "percentage": round(percentage, 1),
                        "incoming_calls": source_data.get("incoming_calls", 0),
                        "outgoing_calls": source_data.get("outgoing_calls", 0),
                        "incoming_conversion": source_data.get("incoming_conversion", 0),
                        "outgoing_conversion": source_data.get("outgoing_conversion", 0),
                        "overall_conversion": source_data.get("overall_conversion", 0),
                        "avg_score": source_data.get("avg_score", 0),
                    }
                )

            # Сортируем по количеству звонков (сначала больше)
            sources.sort(key=lambda x: x["count"], reverse=True)
            
            # Получаем статистику по дням недели
            weekday_stats = metrics.get("weekday_stats", [])
            
            # Получаем статистику по дням
            daily_stats = metrics.get("daily_stats", [])

            # Создаём словарь маппинга source_id -> source_name для унификации
            source_id_map = {}
            source_name_map = {}  # Обратный словарь name -> id
            for src in sources:
                # Проверяем, что src["source"] не None перед вызовом lower()
                if src.get("id") is not None:
                    source_id_map[src["id"]] = src.get("source") or "Неизвестный источник"
                
                # Добавляем в source_name_map только если source не None
                if src.get("source") is not None:
                    source_name_map[src["source"].lower()] = src["id"]
                    
            # Нормализуем источники в данных звонков для отображения в таблице
            for call in calls_data:
                # Если у звонка есть source_id, обновляем его источник согласно source_id_map
                if call.get("source_id") and call["source_id"] in source_id_map:
                    call["source"] = source_id_map[call["source_id"]]
                # Если у звонка нет source_id, но есть source, пробуем найти соответствие в source_name_map
                elif call.get("source") and isinstance(call["source"], str) and call["source"].lower() in source_name_map:
                    # Используем правильное название из словаря маппинга
                    source_id = source_name_map[call["source"].lower()]
                    call["source"] = source_id_map[source_id]
                # Если ничего не подошло, но source похож на ID, пробуем его
                elif call.get("source") and call["source"] in source_id_map:
                    call["source"] = source_id_map[call["source"]]
                # Если всё ещё не нашли, ищем частичное совпадение
                elif call.get("source") and isinstance(call["source"], str):
                    source_text = call["source"].lower()
                    for source_name, source_id in source_name_map.items():
                        # Ищем совпадение хотя бы 3 символов для уверенности
                        if len(source_text) >= 3 and len(source_name) >= 3:
                            if source_text in source_name or source_name in source_text:
                                call["source"] = source_id_map[source_id]
                                break

            # Формируем данные отчета
            report_data = {
                "report_id": report_id,
                "report_type": request.report_type,
                "clinic_id": request.clinic_id,
                "clinic_name": clinic_name,
                "start_date": request.start_date,
                "end_date": request.end_date,
                "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                "total_calls": metrics.get("total_calls", 0),
                "incoming_calls": metrics.get("incoming_calls", 0),
                "outgoing_calls": metrics.get("outgoing_calls", 0),
                "conversion_rate": metrics.get("conversion_rate", 0),
                "avg_score": metrics.get("avg_score", 0) * 10,  
                "avg_duration": metrics.get("avg_duration", 0),
                "administrators": administrators,
                "calls": calls_data,
                "call_types": call_types,
                "sources": sources,
                "weekday_stats": weekday_stats,
                "daily_stats": daily_stats,
            }

            # Генерируем графики, PDF and Excel отчеты
            chart_paths = self._generate_charts(report_data, report_temp_dir)
            pdf_path = self._generate_pdf_report(
                report_data, chart_paths, report_temp_dir
            )
            excel_path = self._generate_excel_report(report_data, report_temp_dir)

            # Сохраняем отчет в MongoDB
            await self._save_report_to_db(report_data, pdf_path, excel_path)

            return pdf_path, excel_path, report_data

        except Exception as e:
            logger.error(f"Ошибка генерации отчета: {e}")
            import traceback

            logger.error(traceback.format_exc())
            raise

    async def generate_summary_report(self, request: CallReportRequest) -> Dict[str, Any]:
        """
        Генерирует сводный отчет по звонкам

        :param request: Объект запроса с параметрами для создания отчета
        :return: Словарь с данными отчета и ссылками на скачивание
        """
        try:
            # Преобразуем даты для MongoDB и создаем временную директорию
            start_date_db = self._convert_date_format(request.start_date)
            end_date_db = self._convert_date_format(request.end_date)

            report_id = str(uuid4())
            report_temp_dir = os.path.join(self.temp_dir, report_id)
            os.makedirs(report_temp_dir, exist_ok=True)

            # Получаем название клиники
            clinic_name = "Все клиники" if not request.clinic_id else "Клиника не определена"
            if request.clinic_id:
                try:
                    # Создаем список возможных ID форматов
                    possible_ids = [request.clinic_id]
                    if isinstance(request.clinic_id, str):
                        try:
                            possible_ids.append(ObjectId(request.clinic_id))
                        except:
                            pass

                    # Поиск по всем возможным форматам
                    for possible_id in possible_ids:
                        clinic = await self.db.clinics.find_one({"_id": possible_id})
                        if clinic:
                            clinic_name = clinic.get("name", "Клиника не определена")
                            break

                        clinic = await self.db.clinics.find_one({"client_id": possible_id})
                        if clinic:
                            clinic_name = clinic.get("name", "Клиника не определена")
                            break
                except Exception as e:
                    logger.error(f"Ошибка при получении информации о клинике: {e}")
            
            # Получаем данные о звонках
            analytics_data = await mongodb_service.get_call_analytics(
                start_date=start_date_db,
                end_date=end_date_db,
                clinic_id=request.clinic_id,
                administrator_ids=request.administrator_ids,
                analysis_exists=True
            )
            
            calls_data = await self.get_calls_data(
                start_date=start_date_db,
                end_date=end_date_db,
                clinic_id=request.clinic_id,
                administrator_ids=request.administrator_ids,
            )
            
            # Объединяем данные из обоих источников
            combined_data = analytics_data + [call for call in calls_data if call.get("id") not in [str(a.get("_id", "")) for a in analytics_data]]
            
            # Подготавливаем данные для отчета
            total_calls = len(combined_data)
            incoming_calls = sum(1 for call in combined_data if call.get("call_type", {}).get("direction") == "incoming")
            outgoing_calls = total_calls - incoming_calls
            
            # Рассчитываем среднюю и общую длительность звонков
            total_duration = 0
            valid_duration_count = 0
            
            for call in combined_data:
                duration = call.get("duration", 0)
                if duration > 0:
                    total_duration += duration
                    valid_duration_count += 1
            
            avg_duration = total_duration / valid_duration_count if valid_duration_count > 0 else 0
            
            # Собираем статистику по администраторам
            admin_stats = {}
            for call in combined_data:
                admin_id = str(call.get("administrator_id", ""))
                admin_name = call.get("administrator_name", "Неизвестный администратор")
                
                if admin_id not in admin_stats:
                    admin_stats[admin_id] = {
                        "id": admin_id,
                        "name": admin_name,
                        "total_calls": 0,
                        "incoming_calls": 0,
                        "outgoing_calls": 0,
                        "total_duration": 0,
                        "avg_duration": 0
                    }
                
                admin_stats[admin_id]["total_calls"] += 1
                
                if call.get("call_type", {}).get("direction") == "incoming":
                    admin_stats[admin_id]["incoming_calls"] += 1
                else:
                    admin_stats[admin_id]["outgoing_calls"] += 1
                
                duration = call.get("duration", 0)
                if duration > 0:
                    admin_stats[admin_id]["total_duration"] += duration
            
            # Вычисляем среднюю длительность для каждого администратора
            for admin_id, stats in admin_stats.items():
                stats["avg_duration"] = stats["total_duration"] / stats["total_calls"] if stats["total_calls"] > 0 else 0
            
            # Формируем список администраторов
            administrators = []
            for admin_id, stats in admin_stats.items():
                administrators.append({
                    "id": stats["id"],
                    "name": stats["name"],
                    "total_calls": stats["total_calls"],
                    "incoming_calls": stats["incoming_calls"],
                    "outgoing_calls": stats["outgoing_calls"],
                    "avg_duration": round(stats["avg_duration"] / 60, 2),  # в минутах
                    "total_duration": round(stats["total_duration"] / 60, 2)  # в минутах
                })
            
            # Сортируем по количеству звонков (сначала больше)
            administrators.sort(key=lambda x: x["total_calls"], reverse=True)
            
            # Собираем статистику по источникам
            source_stats = {}
            for call in combined_data:
                source = call.get("client", {}).get("source", "Неизвестно")
                
                if source not in source_stats:
                    source_stats[source] = {
                        "count": 0,
                        "incoming_calls": 0,
                        "outgoing_calls": 0
                    }
                
                source_stats[source]["count"] += 1
                
                if call.get("call_type", {}).get("direction") == "incoming":
                    source_stats[source]["incoming_calls"] += 1
                else:
                    source_stats[source]["outgoing_calls"] += 1
            
            # Формируем список источников
            sources = []
            for source_name, stats in source_stats.items():
                percentage = (stats["count"] / total_calls * 100) if total_calls > 0 else 0
                sources.append({
                    "source": source_name,
                    "count": stats["count"],
                    "percentage": round(percentage, 1),
                    "incoming_calls": stats["incoming_calls"],
                    "outgoing_calls": stats["outgoing_calls"]
                })
            
            # Сортируем по количеству звонков (сначала больше)
            sources.sort(key=lambda x: x["count"], reverse=True)
            
            # Формируем данные отчета
            report_data = {
                "report_id": report_id,
                "report_type": ReportType.SUMMARY,
                "clinic_id": request.clinic_id,
                "clinic_name": clinic_name,
                "start_date": request.start_date,
                "end_date": request.end_date,
                "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                "total_calls": total_calls,
                "incoming_calls": incoming_calls,
                "outgoing_calls": outgoing_calls,
                "avg_duration": round(avg_duration / 60, 2),  # в минутах
                "total_duration": round(total_duration / 60, 2),  # в минутах
                "administrators": administrators,
                "sources": sources
            }
            
            # Генерируем PDF и Excel отчеты
            pdf_path, excel_path = self._generate_summary_reports(report_data, report_temp_dir)
            
            # Формируем пути для скачивания
            pdf_filename = os.path.basename(pdf_path)
            excel_filename = os.path.basename(excel_path)
            
            # Сохраняем информацию об отчете в базу данных
            report_record = {
                "_id": report_id,
                "report_type": "SUMMARY",
                "clinic_id": request.clinic_id,
                "clinic_name": clinic_name,
                "start_date": request.start_date,
                "end_date": request.end_date,
                "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                "metrics": {
                    "total_calls": total_calls,
                    "incoming_calls": incoming_calls,
                    "outgoing_calls": outgoing_calls,
                    "avg_duration": round(avg_duration / 60, 2),
                    "total_duration": round(total_duration / 60, 2)
                },
                "administrators": administrators,
                "sources": sources,
                "links": {
                    "pdf": f"/api/call/reports/{pdf_filename}/download",
                    "excel": f"/api/call/reports/{excel_filename}/download"
                }
            }
            
            # Сохраняем в базу данных
            try:
                await self.db.call_reports.insert_one(report_record)
                logger.info(f"Отчет {report_id} сохранен в базу данных")
            except Exception as e:
                logger.error(f"Ошибка при сохранении отчета в базу данных: {e}")
            
            # Формируем ответ
            return {
                "success": True,
                "message": "Отчет успешно сгенерирован",
                "data": {
                    "report_id": report_id,
                    "report_type": "summary",
                    "clinic_name": clinic_name,
                    "start_date": request.start_date,
                    "end_date": request.end_date,
                    "total_calls": total_calls,
                    "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                    "pdf_link": f"/api/call/reports/{pdf_filename}/download",
                    "excel_link": f"/api/call/reports/{excel_filename}/download",
                }
            }
            
        except Exception as e:
            logger.error(f"Ошибка при генерации сводного отчета: {e}")
            import traceback
            logger.error(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"Ошибка при генерации отчета: {str(e)}")
    
    def _generate_summary_reports(self, report_data: Dict[str, Any], temp_dir: str) -> Tuple[str, str]:
        """
        Генерирует PDF и Excel отчеты на основе данных
        
        :param report_data: Данные для отчета
        :param temp_dir: Временная директория для сохранения отчетов
        :return: Кортеж из путей к PDF и Excel файлам
        """
        # Генерируем название файлов
        clinic_name = report_data.get("clinic_name", "").replace(" ", "_")
        start_date = report_data.get("start_date", "").replace(".", "_")
        end_date = report_data.get("end_date", "").replace(".", "_")
        
        base_filename = f"summary_report_{clinic_name}_{start_date}_{end_date}"
        pdf_path = os.path.join(temp_dir, f"{base_filename}.pdf")
        excel_path = os.path.join(temp_dir, f"{base_filename}.xlsx")
        
        # 1. Создаем PDF-отчет
        self._create_summary_pdf(report_data, pdf_path)
        
        # 2. Создаем Excel-отчет
        self._create_summary_excel(report_data, excel_path)
        
        return pdf_path, excel_path
    
    def _create_summary_pdf(self, report_data: Dict[str, Any], output_path: str) -> None:
        """
        Создает PDF-отчет с сводной статистикой
        
        :param report_data: Данные для отчета
        :param output_path: Путь для сохранения PDF-файла
        """
        # Создаем документ
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=inch/2,
            leftMargin=inch/2,
            topMargin=inch/2,
            bottomMargin=inch/2
        )
        
        # Создаем стили
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontName='DejaVuSans-Bold',
            fontSize=18,
            alignment=1,  # По центру
            spaceAfter=12
        )
        
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontName='DejaVuSans-Bold',
            fontSize=14,
            spaceAfter=6
        )
        
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontName='DejaVuSans',
            fontSize=10,
            spaceAfter=4
        )
        
        # Создаем элементы для отчета
        elements = []
        
        # Добавляем заголовок
        elements.append(Paragraph(f"Сводный отчет по звонкам клиники {report_data.get('clinic_name', '')}", title_style))
        elements.append(Paragraph(f"Период: {report_data.get('start_date', '')} - {report_data.get('end_date', '')}", normal_style))
        elements.append(Paragraph(f"Дата создания: {report_data.get('generated_at', '')}", normal_style))
        elements.append(Spacer(1, 12))
        
        # Добавляем основную статистику
        elements.append(Paragraph("Основная статистика", heading_style))
        
        # Создаем таблицу
        basic_stats = [
            ["Показатель", "Значение"],
            ["Всего звонков", str(report_data.get("total_calls", 0))],
            ["Входящие звонки", str(report_data.get("incoming_calls", 0))],
            ["Исходящие звонки", str(report_data.get("outgoing_calls", 0))],
            ["Средняя длительность звонка (мин)", str(report_data.get("avg_duration", 0))],
            ["Общая длительность всех звонков (мин)", str(report_data.get("total_duration", 0))]
        ]
        
        basic_table = Table(basic_stats, colWidths=[3*inch, 3*inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(basic_table)
        elements.append(Spacer(1, 12))
        
        # Добавляем статистику по администраторам
        elements.append(Paragraph("Статистика по администраторам", heading_style))
        
        # Создаем таблицу по администраторам
        admin_data = [
            ["Администратор", "Всего", "Входящие", "Исходящие", "Ср. длит. (мин)", "Общ. длит. (мин)"]
        ]
        
        for admin in report_data.get("administrators", []):
            admin_data.append([
                admin.get("name", ""),
                str(admin.get("total_calls", 0)),
                str(admin.get("incoming_calls", 0)),
                str(admin.get("outgoing_calls", 0)),
                str(admin.get("avg_duration", 0)),
                str(admin.get("total_duration", 0))
            ])
        
        admin_table = Table(admin_data, colWidths=[2.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.9*inch, 0.9*inch])
        admin_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(admin_table)
        elements.append(Spacer(1, 12))
        
        # Добавляем статистику по источникам трафика
        elements.append(Paragraph("Статистика по источникам трафика", heading_style))
        
        # Создаем таблицу по источникам
        source_data = [
            ["Источник", "Количество", "Процент", "Входящие", "Исходящие"]
        ]
        
        for source in report_data.get("sources", []):
            source_data.append([
                source.get("source", ""),
                str(source.get("count", 0)),
                f"{source.get('percentage', 0)}%",
                str(source.get("incoming_calls", 0)),
                str(source.get("outgoing_calls", 0))
            ])
        
        source_table = Table(source_data, colWidths=[3*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        source_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(source_table)
        
        # Создаем PDF
        doc.build(elements)
    
    def _create_summary_excel(self, report_data: Dict[str, Any], output_path: str) -> None:
        """
        Создает Excel-отчет с сводной статистикой
        
        :param report_data: Данные для отчета
        :param output_path: Путь для сохранения Excel-файла
        """
        # Создаем рабочую книгу и листы
        workbook = Workbook()
        
        # Основной лист со сводной информацией
        ws_summary = workbook.active
        ws_summary.title = "Сводная информация"
        
        # Лист с данными по администраторам
        ws_admins = workbook.create_sheet("Администраторы")
        
        # Лист с данными по источникам
        ws_sources = workbook.create_sheet("Источники")
        
        # 1. Заполняем лист со сводной информацией
        ws_summary.append(["Сводный отчет по звонкам"])
        ws_summary.append(["Клиника:", report_data.get("clinic_name", "")])
        ws_summary.append(["Период:", f"{report_data.get('start_date', '')} - {report_data.get('end_date', '')}" ])
        ws_summary.append(["Дата создания:", report_data.get("generated_at", "")])
        ws_summary.append([])
        
        ws_summary.append(["Основная статистика"])
        ws_summary.append(["Показатель", "Значение"])
        ws_summary.append(["Всего звонков", report_data.get("total_calls", 0)])
        ws_summary.append(["Входящие звонки", report_data.get("incoming_calls", 0)])
        ws_summary.append(["Исходящие звонки", report_data.get("outgoing_calls", 0)])
        ws_summary.append(["Средняя длительность звонка (мин)", report_data.get("avg_duration", 0)])
        ws_summary.append(["Общая длительность всех звонков (мин)", report_data.get("total_duration", 0)])
        
        # 2. Заполняем лист с данными по администраторам
        ws_admins.append(["Статистика по администраторам"])
        ws_admins.append(["Администратор", "Всего звонков", "Входящие", "Исходящие", "Средняя длительность (мин)", "Общая длительность (мин)"])
        
        for admin in report_data.get("administrators", []):
            ws_admins.append([
                admin.get("name", ""),
                admin.get("total_calls", 0),
                admin.get("incoming_calls", 0),
                admin.get("outgoing_calls", 0),
                admin.get("avg_duration", 0),
                admin.get("total_duration", 0)
            ])
        
        # 3. Заполняем лист с данными по источникам
        ws_sources.append(["Статистика по источникам трафика"])
        ws_sources.append(["Источник", "Количество звонков", "Процент от общего", "Входящие", "Исходящие"])
        
        for source in report_data.get("sources", []):
            ws_sources.append([
                source.get("source", ""),
                source.get("count", 0),
                source.get("percentage", 0)/100,  # Для процентного форматирования
                source.get("incoming_calls", 0),
                source.get("outgoing_calls", 0)
            ])
        
        # Форматирование
        # Заголовки
        for sheet in [ws_summary, ws_admins, ws_sources]:
            sheet.row_dimensions[1].height = 20
            sheet["A1"].font = Font(bold=True, size=14)
        
        # Форматирование процентов
        for i in range(3, len(report_data.get("sources", [])) + 3):
            ws_sources[f"C{i}"].number_format = '0.0%'
        
        # Автоматическая ширина колонок
        for sheet in [ws_summary, ws_admins, ws_sources]:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width
        
        # Сохраняем файл
        workbook.save(output_path)

    async def generate_analytics_report(self, request: AnalyticsReportRequest) -> Dict[str, Any]:
        """
        Создает аналитический отчет на основе данных из коллекции calls с использованием метрик анализа нейронки.
        
        :param request: Объект запроса с параметрами отчета
        :return: Словарь с результатами и ссылками на файлы отчета
        """
        try:
            # Регистрируем шрифты
            self._register_fonts()
            
            # Определяем диапазон дат
            start_date, end_date = self._parse_date_range(request.start_date, request.end_date)
            logger.info(f"Генерация аналитического отчета за период {start_date} - {end_date}")
            
            # Создаем временную директорию для отчета
            temp_dir = tempfile.mkdtemp(prefix="analytics_report_")
            
            # Формируем условия запроса
            query = {
                "created_at": {
                    "$gte": int(datetime.strptime(start_date, "%d.%m.%Y").timestamp()),
                    "$lte": int(datetime.strptime(end_date, "%d.%m.%Y").timestamp())
                }
            }
            
            # Добавляем фильтр по клинике, только если он указан
            if request.clinic_id:
                query["clinic_id"] = request.clinic_id
            
            # Добавляем фильтр по администраторам, только если он указан
            if request.administrator_ids and len(request.administrator_ids) > 0:
                query["administrator"] = {"$in": request.administrator_ids}
            
            # Получаем данные звонков из MongoDB
            calls_data = await self.db.calls.find(query).to_list(length=None)
            
            if not calls_data:
                return {
                    "success": False,
                    "message": "Нет данных за указанный период",
                    "data": None
                }
            
            # Преобразуем данные в DataFrame для анализа
            df = pd.DataFrame(calls_data)
            
            # Добавляем столбец с датой в формате datetime для группировки
            df['date'] = pd.to_datetime(df['created_at'], unit='s')
            df['week'] = df['date'].dt.isocalendar().week
            df['weekday'] = df['date'].dt.day_name()
            df['hour'] = df['date'].dt.hour
            
            # Фильтруем только звонки с анализом
            analyzed_calls = df[df['analysis'].notna()]
            
            # 1. Расчет основных метрик
            total_calls = len(df)
            total_analyzed_calls = len(analyzed_calls)
            
            # Расчет FG% (процент звонков с высокой оценкой приветствия)
            # Извлекаем метрику greeting из вложенного словаря metrics
            analyzed_calls_with_metrics = analyzed_calls[analyzed_calls['metrics'].notna()]
            greeting_scores = [call.get('metrics', {}).get('greeting', 0) for _, call in analyzed_calls_with_metrics.iterrows()]
            high_greeting_scores = [score for score in greeting_scores if score >= 7]  # Считаем высокими оценки от 7 и выше
            fg_percentage = round((len(high_greeting_scores) / total_analyzed_calls * 100) if total_analyzed_calls > 0 else 0, 1)
            
            # Расчет конверсии в запись (процент звонков с категорией "Запись на приём")
            appointment_calls = analyzed_calls[analyzed_calls['call_category'] == 'Запись на приём']
            conversion_rate = round((len(appointment_calls) / total_analyzed_calls * 100) if total_analyzed_calls > 0 else 0, 1)
            
            # Суммарное время диалогов в минутах
            total_duration_minutes = round(df['duration'].sum() / 60, 1)
            
            # Средняя скорость обработки в минутах
            avg_processing_speed = round(df['processing_speed'].mean(), 1)
            
            # 2. Подготовка данных для визуализаций
            
            # Данные по администраторам
            admin_stats = df.groupby('administrator').agg({
                'administrator': 'count',
                'duration': 'sum',
                'call_direction': lambda x: (x == 'Входящий').sum()
            }).reset_index()
            
            admin_stats.columns = ['administrator', 'calls', 'total_duration', 'incoming_calls']
            admin_stats['outgoing_calls'] = admin_stats['calls'] - admin_stats['incoming_calls']
            admin_stats['avg_duration'] = admin_stats['total_duration'] / admin_stats['calls']
            admin_stats['total_duration_minutes'] = round(admin_stats['total_duration'] / 60, 1)
            admin_stats = admin_stats.sort_values('calls', ascending=False)
            
            # Расчет метрик по каждому администратору
            admin_metrics = []
            for admin in admin_stats['administrator'].unique():
                admin_calls = analyzed_calls_with_metrics[analyzed_calls_with_metrics['administrator'] == admin]
                if len(admin_calls) > 0:
                    metrics = {}
                    for field in ['greeting', 'needs_identification', 'solution_proposal', 
                                 'objection_handling', 'call_closing', 'overall_score']:
                        metrics[field] = round(np.mean([
                            call.get('metrics', {}).get(field, 0) 
                            for _, call in admin_calls.iterrows()
                        ]), 1)
                    
                    admin_appointments = admin_calls[admin_calls['call_category'] == 'Запись на приём']
                    admin_conv_rate = round((len(admin_appointments) / len(admin_calls) * 100), 1) if len(admin_calls) > 0 else 0
                    
                    admin_metrics.append({
                        'administrator': admin,
                        'metrics': metrics,
                        'conversion_rate': admin_conv_rate,
                        'calls_analyzed': len(admin_calls)
                    })
            
            # Данные по источникам трафика
            source_stats = df.groupby('source').size().reset_index(name='counts')
            source_stats = source_stats.sort_values('counts', ascending=False)
            
            # Данные по дням недели
            weekday_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            weekday_stats = df.groupby('weekday').size().reset_index(name='counts')
            weekday_stats['weekday_id'] = weekday_stats['weekday'].apply(lambda x: weekday_order.index(x))
            weekday_stats = weekday_stats.sort_values('weekday_id')
            
            # Данные по часам
            hourly_stats = df.groupby('hour').size().reset_index(name='counts')
            hourly_stats = hourly_stats.sort_values('hour')
            
            # Данные по категориям звонков
            category_stats = analyzed_calls.groupby('call_category').size().reset_index(name='counts')
            category_stats = category_stats.sort_values('counts', ascending=False)
            
            # Данные по неделям
            weekly_stats = df.groupby(['week', 'call_category']).size().reset_index(name='counts')
            
            # 3. Создание визуализаций
            # Устанавливаем темную тему для seaborn
            sns.set_theme(style="darkgrid")
            plt.rcParams.update({
                "figure.facecolor": "#2e3440",
                "axes.facecolor": "#2e3440",
                "axes.edgecolor": "#d8dee9",
                "axes.labelcolor": "#d8dee9",
                "xtick.color": "#d8dee9",
                "ytick.color": "#d8dee9",
                "text.color": "#d8dee9",
                "grid.color": "#3b4252",
                "font.family": "Roboto"
            })
            
            charts = {}
            
            # Создаем график распределения по администраторам
            admin_chart_buffer = io.BytesIO()
            self._create_admin_distribution_plot(admin_stats, admin_chart_buffer, 'png')
            admin_chart_path = os.path.join(temp_dir, 'admin_distribution.png')
            with open(admin_chart_path, 'wb') as f:
                f.write(admin_chart_buffer.getvalue())
            charts['admin_chart'] = admin_chart_path
            
            # Создаем круговую диаграмму по источникам трафика
            source_chart_buffer = io.BytesIO()
            self._create_source_distribution_plot(source_stats, source_chart_buffer, 'png')
            source_chart_path = os.path.join(temp_dir, 'source_distribution.png')
            with open(source_chart_path, 'wb') as f:
                f.write(source_chart_buffer.getvalue())
            charts['source_chart'] = source_chart_path
            
            # Создаем график по категориям звонков
            if not category_stats.empty:
                category_chart_buffer = io.BytesIO()
                self._create_category_distribution_plot(category_stats, category_chart_buffer, 'png')
                category_chart_path = os.path.join(temp_dir, 'category_distribution.png')
                with open(category_chart_path, 'wb') as f:
                    f.write(category_chart_buffer.getvalue())
                charts['category_chart'] = category_chart_path
            
            # Создаем график по дням недели
            weekday_chart_buffer = io.BytesIO()
            self._create_weekday_distribution_plot(weekday_stats, weekday_chart_buffer, 'png')
            weekday_chart_path = os.path.join(temp_dir, 'weekday_distribution.png')
            with open(weekday_chart_path, 'wb') as f:
                f.write(weekday_chart_buffer.getvalue())
            charts['weekday_chart'] = weekday_chart_path
            
            # Создаем график по часам
            hourly_chart_buffer = io.BytesIO()
            self._create_hourly_distribution_plot(hourly_stats, hourly_chart_buffer, 'png')
            hourly_chart_path = os.path.join(temp_dir, 'hourly_distribution.png')
            with open(hourly_chart_path, 'wb') as f:
                f.write(hourly_chart_buffer.getvalue())
            charts['hourly_chart'] = hourly_chart_path
            
            # Создаем график по неделям
            if not weekly_stats.empty:
                weekly_chart_buffer = io.BytesIO()
                self._create_weekly_distribution_plot(weekly_stats, weekly_chart_buffer, 'png')
                weekly_chart_path = os.path.join(temp_dir, 'weekly_distribution.png')
                with open(weekly_chart_path, 'wb') as f:
                    f.write(weekly_chart_buffer.getvalue())
                charts['weekly_chart'] = weekly_chart_path
            
            # Создаем график метрик нейронки для каждого администратора
            for admin_data in admin_metrics:
                admin_name = admin_data['administrator']
                admin_metrics_chart_buffer = io.BytesIO()
                self._create_metrics_chart(admin_data['metrics'], admin_metrics_chart_buffer, 'png')
                admin_metrics_chart_path = os.path.join(temp_dir, f'metrics_chart_{admin_name.replace(" ", "_")}.png')
                with open(admin_metrics_chart_path, 'wb') as f:
                    f.write(admin_metrics_chart_buffer.getvalue())
                charts[f'metrics_chart_{admin_name}'] = admin_metrics_chart_path
            
            # Создаем тепловую карту звонков по дням недели и часам
            heatmap_buffer = io.BytesIO()
            self._create_call_heatmap(df, heatmap_buffer, 'png')
            heatmap_path = os.path.join(temp_dir, 'call_heatmap.png')
            with open(heatmap_path, 'wb') as f:
                f.write(heatmap_buffer.getvalue())
            charts['call_heatmap'] = heatmap_path
            
            # 4. Создание PDF и Excel отчетов
            report_id = str(uuid.uuid4())
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_filename = f"analytics_report_{timestamp}.pdf"
            excel_filename = f"analytics_report_{timestamp}.xlsx"
            
            pdf_path = os.path.join(temp_dir, pdf_filename)
            excel_path = os.path.join(temp_dir, excel_filename)
            
            # Формируем данные для отчета
            report_data = {
                "_id": report_id,
                "clinic_id": request.clinic_id,
                "report_type": "analytics",
                "start_date": start_date,
                "end_date": end_date,
                "generated_at": datetime.now().isoformat(),
                "metrics": {
                    "total_calls": total_calls,
                    "total_analyzed_calls": total_analyzed_calls,
                    "fg_percentage": fg_percentage,
                    "conversion_rate": conversion_rate,
                    "total_duration_minutes": total_duration_minutes,
                    "avg_processing_speed": avg_processing_speed,
                    "admin_stats": admin_stats.to_dict('records'),
                    "admin_metrics": admin_metrics,
                    "source_stats": source_stats.to_dict('records'),
                    "category_stats": category_stats.to_dict('records') if not category_stats.empty else [],
                    "weekday_stats": weekday_stats.to_dict('records'),
                    "hourly_stats": hourly_stats.to_dict('records')
                },
                "charts": charts
            }
            
            # Создаем PDF отчет
            self._create_analytics_pdf(report_data, pdf_path)
            
            # Создаем Excel отчет
            self._create_analytics_excel(report_data, excel_path)
            
            # 5. Копируем файлы в постоянное хранилище
            final_pdf_path = os.path.join(self.reports_dir, pdf_filename)
            final_excel_path = os.path.join(self.reports_dir, excel_filename)
            
            shutil.copy2(pdf_path, final_pdf_path)
            shutil.copy2(excel_path, final_excel_path)
            
            # 6. Сохраняем запись об отчете в базу данных
            report_record = {
                "_id": report_id,
                "clinic_id": request.clinic_id,
                "report_type": "analytics",
                "start_date": start_date,
                "end_date": end_date,
                "generated_at": datetime.now().isoformat(),
                "metrics": {
                    "total_calls": total_calls,
                    "fg_percentage": fg_percentage,
                    "conversion_rate": conversion_rate,
                    "total_duration_minutes": total_duration_minutes,
                    "avg_processing_speed": avg_processing_speed
                },
                "links": {
                    "pdf": f"/api/call/reports/{pdf_filename}/download",
                    "excel": f"/api/call/reports/{excel_filename}/download"
                }
            }
            
            await self.db.reports.insert_one(report_record)
            
            # Очищаем временную директорию
            shutil.rmtree(temp_dir)
            
            return {
                "success": True,
                "message": "Аналитический отчет успешно сгенерирован",
                "data": {
                    "report_id": report_id,
                    "total_calls": total_calls,
                    "total_analyzed_calls": total_analyzed_calls,
                    "admins_count": len(admin_stats),
                    "fg_percentage": fg_percentage,
                    "conversion_rate": conversion_rate,
                    "total_duration_minutes": total_duration_minutes,
                    "avg_processing_speed": avg_processing_speed,
                    "pdf_link": f"/api/call/reports/{pdf_filename}/download",
                    "excel_link": f"/api/call/reports/{excel_filename}/download"
                }
            }
        except Exception as e:
            logger.error(f"Ошибка при генерации аналитического отчета: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            
            return {
                "success": False,
                "message": f"Ошибка при генерации аналитического отчета: {str(e)}",
                "data": None
            }

    def _create_admin_distribution_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """
        Создает столбчатую диаграмму распределения звонков по администраторам
        """
        plt.figure(figsize=(10, 6))
        
        # Ограничиваем количество отображаемых администраторов для читаемости
        top_admins = df.head(10) if len(df) > 10 else df
        
        ax = sns.barplot(
            x='calls', 
            y='administrator', 
            data=top_admins,
            palette='viridis'
        )
        
        # Добавляем значения на столбцы
        for i, v in enumerate(top_admins['calls']):
            ax.text(v + 0.5, i, str(v), color='white', va='center')
        
        plt.title('Распределение звонков по администраторам', fontsize=16, color='white')
        plt.xlabel('Количество звонков', fontsize=12)
        plt.ylabel('Администратор', fontsize=12)
        plt.tight_layout()
        
        plt.savefig(buffer, format=format, dpi=100, bbox_inches='tight')
        plt.close()
        buffer.seek(0)

    def _create_source_distribution_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """
        Создает круговую диаграмму распределения звонков по источникам трафика
        """
        plt.figure(figsize=(10, 8))
        
        colors = sns.color_palette('viridis', len(df))
        
        # Если источников много, группируем малочисленные в "Другие"
        if len(df) > 6:
            top_sources = df.head(5)
            other_sources = pd.DataFrame({
                'source': ['Другие'],
                'counts': [df.iloc[5:]['counts'].sum()]
            })
            plot_df = pd.concat([top_sources, other_sources], ignore_index=True)
        else:
            plot_df = df
        
        plt.pie(
            plot_df['counts'], 
            labels=plot_df['source'], 
            autopct='%1.1f%%',
            colors=colors[:len(plot_df)],
            textprops={'color': 'white'}
        )
        
        plt.title('Распределение звонков по источникам трафика', fontsize=16, color='white')
        plt.tight_layout()
        
        plt.savefig(buffer, format=format, dpi=100, bbox_inches='tight')
        plt.close()
        buffer.seek(0)

    def _create_weekly_distribution_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """
        Создает график распределения звонков по неделям и категориям
        """
        plt.figure(figsize=(12, 6))
        
        # Получаем список уникальных недель и категорий
        weeks = sorted(df['week'].unique())
        categories = df['call_category'].unique()
        
        # Создаем сводную таблицу для удобства построения графика
        pivot_df = df.pivot_table(
            index='week', 
            columns='call_category', 
            values='counts', 
            aggfunc='sum',
            fill_value=0
        ).reindex(weeks)
        
        # Создаем столбчатую диаграмму
        ax = pivot_df.plot(
            kind='bar', 
            stacked=False,
            figsize=(12, 6),
            color=sns.color_palette('viridis', len(categories))
        )
        
        plt.title('Распределение звонков по неделям и категориям', fontsize=16, color='white')
        plt.xlabel('Неделя', fontsize=12)
        plt.ylabel('Количество звонков', fontsize=12)
        plt.legend(title='Категория звонка', bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.tight_layout()
        
        plt.savefig(buffer, format=format, dpi=100, bbox_inches='tight')
        plt.close()
        buffer.seek(0)

    def _create_metrics_chart(self, metrics: Dict[str, float], buffer: io.BytesIO, format: str):
        """
        Создает радарную диаграмму для отображения метрик нейронки
        """
        # Подготавливаем данные для радарной диаграммы
        categories = list(metrics.keys())
        values = list(metrics.values())
        
        # Дублируем первое значение в конец для замыкания многоугольника
        values_closed = values + [values[0]]
        categories_closed = categories + [categories[0]]
        
        # Вычисляем углы для категорий
        N = len(categories)
        angles = [n / float(N) * 2 * np.pi for n in range(N)]
        angles += angles[:1]
        
        # Создаем график
        plt.figure(figsize=(8, 8), facecolor='#2e3440')
        ax = plt.subplot(111, polar=True)
        
        # Настраиваем внешний вид
        ax.set_theta_offset(np.pi / 2)
        ax.set_theta_direction(-1)
        
        # Рисуем оси для каждой категории
        plt.xticks(angles[:-1], categories, color='white', size=12)
        
        # Добавляем метки для оценок (от 0 до 10)
        ax.set_rlabel_position(0)
        plt.yticks([2, 4, 6, 8, 10], ["2", "4", "6", "8", "10"], color="white", size=10)
        plt.ylim(0, 10)
        
        # Строим график
        ax.plot(angles, values_closed, 'o-', linewidth=2, color='#88c0d0')
        ax.fill(angles, values_closed, alpha=0.25, color='#88c0d0')
        
        # Добавляем значения на график
        for i, (angle, value) in enumerate(zip(angles[:-1], values)):
            ax.text(angle, value + 0.5, f"{value}", 
                    ha='center', va='center', color='white', fontsize=12,
                    bbox=dict(boxstyle="round,pad=0.3", facecolor='#4c566a', alpha=0.7))
        
        # Название графика
        plt.title('Средние оценки метрик нейронки', size=16, color='white', y=1.1)
        
        plt.tight_layout()
        plt.savefig(buffer, format=format, dpi=100, bbox_inches='tight')
        plt.close()
        buffer.seek(0)

    def _create_analytics_pdf(self, report_data: Dict[str, Any], output_path: str) -> None:
        """
        Создает PDF-файл аналитического отчета
        """
        # Создаем PDF документ
        c = canvas.Canvas(output_path, pagesize=letter)
        width, height = letter
        
        # Настраиваем шрифты
        pdfmetrics.registerFont(TTFont('Roboto', 'app/fonts/Roboto-Regular.ttf'))
        pdfmetrics.registerFont(TTFont('Roboto-Bold', 'app/fonts/Roboto-Bold.ttf'))
        
        # Добавляем заголовок
        c.setFont('Roboto-Bold', 18)
        c.drawString(50, height - 50, f"Аналитический отчет по звонкам")
        
        c.setFont('Roboto', 12)
        c.drawString(50, height - 70, f"Период: {report_data['start_date']} - {report_data['end_date']}")
        c.drawString(50, height - 90, f"Дата формирования: {report_data['generated_at'].split('T')[0]}")
        
        # Добавляем основные метрики
        c.setFont('Roboto-Bold', 14)
        c.drawString(50, height - 130, "Основные показатели:")
        
        metrics = report_data['metrics']
        y_pos = height - 160
        
        c.setFont('Roboto', 12)
        c.drawString(50, y_pos, f"Всего звонков: {metrics['total_calls']}")
        c.drawString(300, y_pos, f"Проанализировано: {metrics['total_analyzed_calls']}")
        
        y_pos -= 25
        c.drawString(50, y_pos, f"FG%: {metrics['fg_percentage']}%")
        c.drawString(300, y_pos, f"Конверсия в запись: {metrics['conversion_rate']}%")
        
        y_pos -= 25
        c.drawString(50, y_pos, f"Общее время диалогов: {metrics['total_duration_minutes']} мин.")
        c.drawString(300, y_pos, f"Средняя скорость обработки: {metrics['avg_processing_speed']} мин.")
        
        # Добавляем графики
        y_pos -= 50
        c.setFont('Roboto-Bold', 14)
        c.drawString(50, y_pos, "Распределение звонков по администраторам:")
        
        admin_chart = ImageReader(report_data['charts']['admin_chart'])
        c.drawImage(admin_chart, 50, y_pos - 250, width=450, height=230, preserveAspectRatio=True)
        
        # Переходим на следующую страницу - источники трафика
        c.showPage()
        
        c.setFont('Roboto-Bold', 18)
        c.drawString(50, height - 50, f"Аналитический отчет по звонкам (продолжение)")
        
        c.setFont('Roboto-Bold', 14)
        c.drawString(50, height - 90, "Распределение звонков по источникам трафика:")
        
        source_chart = ImageReader(report_data['charts']['source_chart'])
        c.drawImage(source_chart, 50, height - 380, width=450, height=270, preserveAspectRatio=True)
        
        # Переходим на следующую страницу - категории звонков
        c.showPage()
        
        c.setFont('Roboto-Bold', 18)
        c.drawString(50, height - 50, f"Аналитический отчет по звонкам (продолжение)")
        
        if 'category_chart' in report_data['charts']:
            c.setFont('Roboto-Bold', 14)
            c.drawString(50, height - 90, "Распределение звонков по категориям:")
            
            category_chart = ImageReader(report_data['charts']['category_chart'])
            c.drawImage(category_chart, 50, height - 380, width=450, height=270, preserveAspectRatio=True)
        else:
            c.setFont('Roboto', 12)
            c.drawString(50, height - 90, "Нет данных о категориях звонков")
        
        # Переходим на следующую страницу - дни недели
        c.showPage()
        
        c.setFont('Roboto-Bold', 18)
        c.drawString(50, height - 50, f"Аналитический отчет по звонкам (продолжение)")
        
        c.setFont('Roboto-Bold', 14)
        c.drawString(50, height - 90, "Распределение звонков по дням недели:")
        
        weekday_chart = ImageReader(report_data['charts']['weekday_chart'])
        c.drawImage(weekday_chart, 50, height - 380, width=450, height=270, preserveAspectRatio=True)
        
        # Переходим на следующую страницу - часы
        c.showPage()
        
        c.setFont('Roboto-Bold', 18)
        c.drawString(50, height - 50, f"Аналитический отчет по звонкам (продолжение)")
        
        c.setFont('Roboto-Bold', 14)
        c.drawString(50, height - 90, "Распределение звонков по часам дня:")
        
        hourly_chart = ImageReader(report_data['charts']['hourly_chart'])
        c.drawImage(hourly_chart, 50, height - 380, width=450, height=270, preserveAspectRatio=True)
        
        # Переходим на следующую страницу - тепловая карта
        c.showPage()
        
        c.setFont('Roboto-Bold', 18)
        c.drawString(50, height - 50, f"Аналитический отчет по звонкам (продолжение)")
        
        c.setFont('Roboto-Bold', 14)
        c.drawString(50, height - 90, "Тепловая карта распределения звонков по дням и часам:")
        
        heatmap_chart = ImageReader(report_data['charts']['call_heatmap'])
        c.drawImage(heatmap_chart, 50, height - 460, width=500, height=350, preserveAspectRatio=True)
        
        # Добавляем по странице для каждого администратора
        for admin_data in report_data['metrics']['admin_metrics']:
            admin_name = admin_data['administrator']
            chart_key = f'metrics_chart_{admin_name}'
            
            if chart_key in report_data['charts']:
                c.showPage()
                
                c.setFont('Roboto-Bold', 18)
                c.drawString(50, height - 50, f"Результаты администратора: {admin_name}")
                
                # Общая информация по администратору
                c.setFont('Roboto-Bold', 14)
                c.drawString(50, height - 90, "Общие показатели:")
                
                # Найдем данные администратора из admin_stats
                admin_stats = next((stats for stats in report_data['metrics']['admin_stats'] 
                                   if stats['administrator'] == admin_name), {})
                
                y_pos = height - 120
                c.setFont('Roboto', 12)
                
                if admin_stats:
                    c.drawString(50, y_pos, f"Всего звонков: {admin_stats.get('calls', 0)}")
                    c.drawString(300, y_pos, f"Проанализировано: {admin_data.get('calls_analyzed', 0)}")
                    
                    y_pos -= 25
                    c.drawString(50, y_pos, f"Входящих: {admin_stats.get('incoming_calls', 0)}")
                    c.drawString(300, y_pos, f"Исходящих: {admin_stats.get('outgoing_calls', 0)}")
                    
                    y_pos -= 25
                    c.drawString(50, y_pos, f"Конверсия: {admin_data.get('conversion_rate', 0)}%")
                    c.drawString(300, y_pos, f"Общее время: {admin_stats.get('total_duration_minutes', 0)} мин.")
                
                # Метрики нейронки для администратора
                y_pos -= 50
                c.setFont('Roboto-Bold', 14)
                c.drawString(50, y_pos, "Показатели качества звонков:")
                
                metrics_chart = ImageReader(report_data['charts'][chart_key])
                c.drawImage(metrics_chart, 100, y_pos - 350, width=400, height=300, preserveAspectRatio=True)
        
        # Завершаем создание PDF
        c.save()

    def _create_analytics_excel(self, report_data: Dict[str, Any], output_path: str) -> None:
        """
        Создает Excel-файл аналитического отчета
        """
        workbook = xlsxwriter.Workbook(output_path)
        
        # Форматы для заголовков и данных
        header_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#2e3440',
            'border': 1,
            'font_name': 'Arial',
            'font_size': 12,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        subheader_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#434c5e',
            'border': 1,
            'font_name': 'Arial',
            'font_size': 11,
            'align': 'left',
            'valign': 'vcenter'
        })
        
        data_format = workbook.add_format({
            'font_name': 'Arial',
            'font_size': 11,
            'align': 'left',
            'valign': 'vcenter',
            'border': 1
        })
        
        metrics = report_data['metrics']
        
        # 1. Лист с основными метриками
        ws_summary = workbook.add_worksheet('Основные показатели')
        
        # Настраиваем ширину столбцов
        ws_summary.set_column('A:A', 30)
        ws_summary.set_column('B:B', 20)
        
        # Заголовок
        ws_summary.merge_range('A1:B1', 'Аналитический отчет по звонкам', header_format)
        ws_summary.write('A2', 'Период:', subheader_format)
        ws_summary.write('B2', f"{report_data['start_date']} - {report_data['end_date']}", data_format)
        
        # Основные метрики
        row = 3
        ws_summary.write(row, 0, 'Показатель', subheader_format)
        ws_summary.write(row, 1, 'Значение', subheader_format)
        
        metrics_data = [
            ('Всего звонков', metrics['total_calls']),
            ('Проанализировано звонков', metrics['total_analyzed_calls']),
            ('FG%', f"{metrics['fg_percentage']}%"),
            ('Конверсия в запись', f"{metrics['conversion_rate']}%"),
            ('Общее время диалогов (мин)', metrics['total_duration_minutes']),
            ('Средняя скорость обработки (мин)', metrics['avg_processing_speed'])
        ]
        
        for i, (label, value) in enumerate(metrics_data):
            row = i + 4
            ws_summary.write(row, 0, label, data_format)
            ws_summary.write(row, 1, value, data_format)
        
        # 2. Лист с данными по администраторам
        ws_admins = workbook.add_worksheet('Администраторы')
        
        ws_admins.set_column('A:A', 30)
        ws_admins.set_column('B:B', 15)
        
        # Заголовок
        ws_admins.merge_range('A1:B1', 'Распределение звонков по администраторам', header_format)
        
        # Данные администраторов
        ws_admins.write('A2', 'Администратор', subheader_format)
        ws_admins.write('B2', 'Количество звонков', subheader_format)
        
        for i, admin_data in enumerate(metrics['admin_stats']):
            row = i + 2
            ws_admins.write(row, 0, admin_data['administrator'], data_format)
            ws_admins.write(row, 1, admin_data['counts'], data_format)
        
        # 3. Лист с данными по источникам трафика
        ws_sources = workbook.add_worksheet('Источники трафика')
        
        ws_sources.set_column('A:A', 30)
        ws_sources.set_column('B:B', 15)
        
        # Заголовок
        ws_sources.merge_range('A1:B1', 'Распределение звонков по источникам трафика', header_format)
        
        # Данные источников
        ws_sources.write('A2', 'Источник', subheader_format)
        ws_sources.write('B2', 'Количество звонков', subheader_format)
        
        for i, source_data in enumerate(metrics['source_stats']):
            row = i + 2
            ws_sources.write(row, 0, source_data['source'], data_format)
            ws_sources.write(row, 1, source_data['counts'], data_format)
        
        # 4. Лист с метриками нейронки
        ws_metrics = workbook.add_worksheet('Метрики нейронки')
        
        ws_metrics.set_column('A:A', 30)
        ws_metrics.set_column('B:B', 15)
        
        # Заголовок
        ws_metrics.merge_range('A1:B1', 'Средние оценки по метрикам нейронки', header_format)
        
        # Данные метрик
        ws_metrics.write('A2', 'Метрика', subheader_format)
        ws_metrics.write('B2', 'Средняя оценка', subheader_format)
        
        metrics_fields = {
            'greeting': 'Приветствие',
            'needs_identification': 'Выявление потребностей',
            'solution_proposal': 'Предложение решения',
            'objection_handling': 'Работа с возражениями',
            'call_closing': 'Закрытие звонка',
            'overall_score': 'Общая оценка'
        }
        
        row = 2
        for field, label in metrics_fields.items():
            ws_metrics.write(row, 0, label, data_format)
            ws_metrics.write(row, 1, metrics['metrics_means'].get(field, 0), data_format)
            row += 1
        
        # Закрываем файл
        workbook.close()
            
    def _parse_date_range(self, start_date_str: str, end_date_str: str) -> tuple:
        """Преобразует строковые даты в формат для запроса данных"""
        try:
            start_date = datetime.strptime(start_date_str, "%d.%m.%Y")
            end_date = datetime.strptime(end_date_str, "%d.%m.%Y")
            return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Некорректный формат даты. Используйте DD.MM.YYYY")
            
    async def _get_calls_data(self, start_date: str, end_date: str, clinic_id: Optional[str], administrator_ids: Optional[List[str]]) -> List[Dict[str, Any]]:
        """Получает данные звонков для аналитического отчета"""
        try:
            # Получаем данные из обоих источников
            analytics_data = await mongodb_service.get_call_analytics(
                start_date=start_date,
                end_date=end_date,
                clinic_id=clinic_id,
                administrator_ids=administrator_ids,
                analysis_exists=True
            )
            
            calls_data = await self.get_calls_data(
                start_date=start_date,
                end_date=end_date,
                clinic_id=clinic_id,
                administrator_ids=administrator_ids,
            )
            
            # Объединяем данные из обоих источников
            combined_data = analytics_data + [call for call in calls_data if call.get("id") not in [str(a.get("_id", "")) for a in analytics_data]]
            
            # Добавляем категории звонков для визуализации
            for call in combined_data:
                # Определяем категорию звонка (для визуализации)
                direction = call.get("call_type", {}).get("direction", "unknown")
                duration = call.get("duration", 0)
                
                if direction == "incoming":
                    call_category = "Входящий"
                    if duration < 60:  # меньше минуты
                        call_category += " (короткий)"
                    elif duration > 300:  # более 5 минут
                        call_category += " (длительный)"
                elif direction == "outgoing":
                    call_category = "Исходящий"
                    if duration < 60:
                        call_category += " (короткий)"
                    elif duration > 300:
                        call_category += " (длительный)"
                else:
                    call_category = "Неизвестный"
                
                call["call_category"] = call_category
                
                # Определяем успешность звонка (для визуализации)
                conversion = call.get("conversion", False)
                score = call.get("metrics", {}).get("overall_score", 0)
                
                call["is_successful"] = conversion or score > 0.7  # Считаем успешным, если есть конверсия или высокий скор
                
                # Добавляем время звонка (для визуализации)
                timestamp = call.get("timestamp", "")
                if timestamp:
                    try:
                        if isinstance(timestamp, str):
                            date_obj = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
                        else:
                            date_obj = timestamp
                        
                        call["call_time"] = date_obj.strftime("%H:%M:%S")
                    except:
                        call["call_time"] = "00:00:00"
                else:
                    call["call_time"] = "00:00:00"
            
            return combined_data
        except Exception as e:
            logger.error(f"Ошибка при получении данных для аналитического отчета: {e}")
            return []

    def _create_category_distribution_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """Создание визуализации распределения категорий звонков"""
        # Подсчитываем количество звонков по категориям
        category_counts = df['call_category'].value_counts().reset_index()
        category_counts.columns = ['Категория', 'Количество']
        
        # Создаем график
        plt.title('Распределение звонков по категориям', fontsize=16, pad=20)
        ax = sns.barplot(x='Категория', y='Количество', data=category_counts)
        
        # Поворачиваем подписи для лучшей читаемости
        plt.xticks(rotation=45, ha='right')
        
        # Добавляем подписи значений
        for p in ax.patches:
            ax.annotate(f'{int(p.get_height())}', 
                      (p.get_x() + p.get_width() / 2., p.get_height()), 
                      ha = 'center', va = 'bottom', 
                      fontsize=12)
        
        plt.tight_layout()
        output_format = 'png' if format == ReportFormat.PDF else format
        plt.savefig(buffer, format=output_format)
        plt.close()

    def _create_call_duration_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """Создание визуализации продолжительности звонков"""
        # Преобразуем продолжительность в числовой формат (секунды)
        if 'duration' in df.columns:
            # Создаем график
            plt.title('Распределение продолжительности звонков', fontsize=16, pad=20)
            ax = sns.histplot(df['duration'], kde=True, bins=20)
            
            plt.xlabel('Продолжительность (секунды)')
            plt.ylabel('Количество звонков')
            
            plt.tight_layout()
            output_format = 'png' if format == ReportFormat.PDF else format
            plt.savefig(buffer, format=output_format)
            plt.close()
        else:
            # Если нет данных о продолжительности, создаем информационный график
            plt.text(0.5, 0.5, 'Нет данных о продолжительности звонков', 
                    horizontalalignment='center', verticalalignment='center', fontsize=14)
            plt.axis('off')
            output_format = 'png' if format == ReportFormat.PDF else format
            plt.savefig(buffer, format=output_format)
            plt.close()

    def _create_time_distribution_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """Создание визуализации распределения звонков по времени"""
        if 'call_time' in df.columns:
            # Преобразуем время звонка в часы
            df['hour'] = pd.to_datetime(df['call_time']).dt.hour
            
            # Создаем график
            plt.title('Распределение звонков по времени суток', fontsize=16, pad=20)
            hour_counts = df['hour'].value_counts().sort_index().reset_index()
            hour_counts.columns = ['Час', 'Количество']
            
            ax = sns.barplot(x='Час', y='Количество', data=hour_counts)
            
            # Добавляем подписи значений
            for p in ax.patches:
                ax.annotate(f'{int(p.get_height())}', 
                          (p.get_x() + p.get_width() / 2., p.get_height()), 
                          ha = 'center', va = 'bottom', 
                          fontsize=10)
            
            plt.xlabel('Час дня (0-23)')
            plt.ylabel('Количество звонков')
            
            plt.tight_layout()
            output_format = 'png' if format == ReportFormat.PDF else format
            plt.savefig(buffer, format=output_format)
            plt.close()
        else:
            # Если нет данных о времени, создаем информационный график
            plt.text(0.5, 0.5, 'Нет данных о времени звонков', 
                    horizontalalignment='center', verticalalignment='center', fontsize=14)
            plt.axis('off')
            output_format = 'png' if format == ReportFormat.PDF else format
            plt.savefig(buffer, format=output_format)
            plt.close()

    def _create_success_rate_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """Создание визуализации успешности звонков"""
        if 'is_successful' in df.columns:
            # Создаем график
            plt.title('Соотношение успешных и неуспешных звонков', fontsize=16, pad=20)
            
            # Подсчитываем количество успешных/неуспешных звонков
            success_counts = df['is_successful'].value_counts().reset_index()
            success_counts.columns = ['Успешность', 'Количество']
            
            # Заменяем булевы значения на текстовые метки
            success_counts['Успешность'] = success_counts['Успешность'].map({True: 'Успешные', False: 'Неуспешные'})
            
            # Создаем круговую диаграмму
            plt.pie(success_counts['Количество'], labels=success_counts['Успешность'], 
                   autopct='%1.1f%%', startangle=90, colors=['#5cb85c', '#d9534f'])
            
            plt.axis('equal')  # Равные пропорции для круговой диаграммы
            plt.tight_layout()
            output_format = 'png' if format == ReportFormat.PDF else format
            plt.savefig(buffer, format=output_format)
            plt.close()
        else:
            # Если нет данных об успешности, создаем информационный график
            plt.text(0.5, 0.5, 'Нет данных об успешности звонков', 
                    horizontalalignment='center', verticalalignment='center', fontsize=14)
            plt.axis('off')
            output_format = 'png' if format == ReportFormat.PDF else format
            plt.savefig(buffer, format=output_format)
            plt.close()

    async def generate_individual_report(self, request: CallReportRequest) -> Dict[str, Any]:
        """
        Генерирует индивидуальный отчет по конкретному администратору
        
        :param request: Объект запроса с параметрами для создания отчета
        :return: Словарь с данными отчета и ссылками на скачивание
        """
        try:
            # Проверяем, что указан хотя бы один администратор
            if not request.administrator_ids or len(request.administrator_ids) == 0:
                raise HTTPException(status_code=400, detail="Для индивидуального отчета необходимо указать ID администратора")
            
            # Если передано несколько ID, используем только первый
            admin_id = request.administrator_ids[0]
            
            # Преобразуем даты для MongoDB и создаем временную директорию
            start_date_db = self._convert_date_format(request.start_date)
            end_date_db = self._convert_date_format(request.end_date)
            
            report_id = str(uuid4())
            report_temp_dir = os.path.join(self.temp_dir, report_id)
            os.makedirs(report_temp_dir, exist_ok=True)
            
            # Получаем название клиники
            clinic_name = "Все клиники" if not request.clinic_id else "Клиника не определена"
            if request.clinic_id:
                try:
                    clinic = await self.db.clinics.find_one({"_id": request.clinic_id})
                    if clinic:
                        clinic_name = clinic.get("name", "Клиника не определена")
                    else:
                        clinic = await self.db.clinics.find_one({"client_id": request.clinic_id})
                        if clinic:
                            clinic_name = clinic.get("name", "Клиника не определена")
                except Exception as e:
                    logger.error(f"Ошибка при получении информации о клинике: {e}")
            
            # Получаем данные звонков администратора
            calls_data = await self.get_calls_data(
                start_date=start_date_db,
                end_date=end_date_db,
                clinic_id=request.clinic_id,
                administrator_ids=[admin_id]
            )
            
            # Получаем информацию об администраторе
            admin_name = "Неизвестный администратор"
            for call in calls_data:
                if call.get("administrator_name") and call.get("administrator_name") != "Неизвестный администратор":
                    admin_name = call.get("administrator_name")
                    break
            
            # Считаем основные метрики
            total_calls = len(calls_data)
            incoming_calls = sum(1 for call in calls_data if call.get("call_type", {}).get("direction") == "incoming")
            outgoing_calls = total_calls - incoming_calls
            
            # Длительность звонков
            total_duration = 0
            valid_duration_count = 0
            
            for call in calls_data:
                duration = call.get("duration", 0)
                if duration > 0:
                    total_duration += duration
                    valid_duration_count += 1
            
            avg_duration = total_duration / valid_duration_count if valid_duration_count > 0 else 0
            
            # Считаем статистику по дням
            daily_stats = {}
            for call in calls_data:
                timestamp = call.get("timestamp", "")
                if timestamp:
                    try:
                        if isinstance(timestamp, str):
                            date_obj = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
                        else:
                            date_obj = timestamp
                        
                        day_key = date_obj.strftime("%d.%m.%Y")
                        
                        if day_key not in daily_stats:
                            daily_stats[day_key] = {
                                "date": day_key,
                                "total": 0,
                                "incoming": 0,
                                "outgoing": 0,
                                "duration": 0
                            }
                        
                        daily_stats[day_key]["total"] += 1
                        
                        if call.get("call_type", {}).get("direction") == "incoming":
                            daily_stats[day_key]["incoming"] += 1
                        else:
                            daily_stats[day_key]["outgoing"] += 1
                        
                        duration = call.get("duration", 0)
                        if duration > 0:
                            daily_stats[day_key]["duration"] += duration
                    except Exception as e:
                        logger.debug(f"Ошибка при обработке даты {timestamp}: {e}")
            
            # Преобразуем в список и сортируем по дате
            daily_calls = []
            for day_key, stats in daily_stats.items():
                avg_day_duration = stats["duration"] / stats["total"] if stats["total"] > 0 else 0
                daily_calls.append({
                    "date": day_key,
                    "total": stats["total"],
                    "incoming": stats["incoming"],
                    "outgoing": stats["outgoing"],
                    "avg_duration": round(avg_day_duration / 60, 2)  # в минутах
                })
            
            # Сортируем по дате (сначала старые)
            daily_calls.sort(key=lambda x: datetime.strptime(x["date"], "%d.%m.%Y"))
            
            # Формируем детальные данные по звонкам
            detailed_calls = []
            for call in calls_data:
                # Форматируем дату и время
                timestamp = call.get("timestamp", "")
                date_str = ""
                time_str = ""
                
                if timestamp:
                    try:
                        if isinstance(timestamp, str):
                            date_obj = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
                        else:
                            date_obj = timestamp
                        
                        date_str = date_obj.strftime("%d.%m.%Y")
                        time_str = date_obj.strftime("%H:%M:%S")
                    except Exception as e:
                        logger.debug(f"Ошибка при форматировании даты {timestamp}: {e}")
                        date_str = str(timestamp)
                
                # Добавляем данные о звонке
                detailed_calls.append({
                    "id": call.get("id", ""),
                    "date": date_str,
                    "time": time_str,
                    "phone": call.get("client", {}).get("phone", ""),
                    "duration": round(call.get("duration", 0) / 60, 2),  # в минутах
                    "direction": "Входящий" if call.get("call_type", {}).get("direction") == "incoming" else "Исходящий",
                    "source": call.get("client", {}).get("source", "Неизвестно")
                })
            
            # Сортируем по дате (сначала новые)
            detailed_calls.sort(key=lambda x: (x["date"], x["time"]), reverse=True)
            
            # Формируем данные отчета
            report_data = {
                "report_id": report_id,
                "report_type": ReportType.INDIVIDUAL,
                "clinic_id": request.clinic_id,
                "clinic_name": clinic_name,
                "administrator_id": admin_id,
                "administrator_name": admin_name,
                "start_date": request.start_date,
                "end_date": request.end_date,
                "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                "total_calls": total_calls,
                "incoming_calls": incoming_calls,
                "outgoing_calls": outgoing_calls,
                "avg_duration": round(avg_duration / 60, 2),  # в минутах
                "total_duration": round(total_duration / 60, 2),  # в минутах
                "daily_calls": daily_calls,
                "detailed_calls": detailed_calls
            }
            
            # Генерируем PDF и Excel отчеты
            pdf_path, excel_path = self._generate_individual_reports(report_data, report_temp_dir)
            
            # Формируем пути для скачивания
            pdf_filename = os.path.basename(pdf_path)
            excel_filename = os.path.basename(excel_path)
            
            # Сохраняем информацию об отчете в базу данных
            report_record = {
                "_id": report_id,
                "report_type": "INDIVIDUAL",
                "clinic_id": request.clinic_id,
                "clinic_name": clinic_name,
                "administrator_id": admin_id,
                "administrator_name": admin_name,
                "start_date": request.start_date,
                "end_date": request.end_date,
                "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                "metrics": {
                    "total_calls": total_calls,
                    "incoming_calls": incoming_calls,
                    "outgoing_calls": outgoing_calls,
                    "avg_duration": round(avg_duration / 60, 2),
                    "total_duration": round(total_duration / 60, 2)
                },
                "daily_stats": daily_calls,
                "links": {
                    "pdf": f"/api/call/reports/{pdf_filename}/download",
                    "excel": f"/api/call/reports/{excel_filename}/download"
                }
            }
            
            # Сохраняем в базу данных
            try:
                await self.db.call_reports.insert_one(report_record)
                logger.info(f"Индивидуальный отчет {report_id} сохранен в базу данных")
            except Exception as e:
                logger.error(f"Ошибка при сохранении индивидуального отчета в базу данных: {e}")
            
            # Формируем ответ
            return {
                "success": True,
                "message": "Индивидуальный отчет успешно сгенерирован",
                "data": {
                    "report_id": report_id,
                    "report_type": "individual",
                    "administrator_name": admin_name,
                    "clinic_name": clinic_name,
                    "start_date": request.start_date,
                    "end_date": request.end_date,
                    "total_calls": total_calls,
                    "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                    "pdf_link": f"/api/call/reports/{pdf_filename}/download",
                    "excel_link": f"/api/call/reports/{excel_filename}/download",
                }
            }
            
        except Exception as e:
            logger.error(f"Ошибка при генерации индивидуального отчета: {e}")
            import traceback
            logger.error(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"Ошибка при генерации отчета: {str(e)}")
    
    def _generate_individual_reports(self, report_data: Dict[str, Any], temp_dir: str) -> Tuple[str, str]:
        """
        Генерирует PDF и Excel файлы для индивидуального отчета
        
        :param report_data: Данные для отчета
        :param temp_dir: Временная директория для сохранения отчетов
        :return: Кортеж из путей к PDF и Excel файлам
        """
        # Генерируем имена файлов
        admin_name = report_data.get("administrator_name", "").replace(" ", "_")
        start_date = report_data.get("start_date", "").replace(".", "_")
        end_date = report_data.get("end_date", "").replace(".", "_")
        
        base_filename = f"individual_report_{admin_name}_{start_date}_{end_date}"
        pdf_path = os.path.join(temp_dir, f"{base_filename}.pdf")
        excel_path = os.path.join(temp_dir, f"{base_filename}.xlsx")
        
        # Создаем PDF отчет
        self._create_individual_pdf(report_data, pdf_path)
        
        # Создаем Excel отчет
        self._create_individual_excel(report_data, excel_path)
        
        return pdf_path, excel_path
    
    def _create_individual_pdf(self, report_data: Dict[str, Any], output_path: str) -> None:
        """
        Создает PDF-отчет для индивидуальной статистики администратора
        
        :param report_data: Данные для отчета
        :param output_path: Путь сохранения PDF-файла
        """
        # Создаем документ
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=inch/2,
            leftMargin=inch/2,
            topMargin=inch/2,
            bottomMargin=inch/2
        )
        
        # Создаем стили
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontName='DejaVuSans-Bold',
            fontSize=16,
            alignment=1,  # По центру
            spaceAfter=10
        )
        
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontName='DejaVuSans-Bold',
            fontSize=14,
            spaceAfter=6
        )
        
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontName='DejaVuSans',
            fontSize=10,
            spaceAfter=4
        )
        
        # Создаем элементы для отчета
        elements = []
        
        # Добавляем заголовок
        elements.append(Paragraph(f"Индивидуальный отчет администратора {report_data.get('administrator_name', '')}", title_style))
        elements.append(Paragraph(f"Клиника: {report_data.get('clinic_name', '')}", normal_style))
        elements.append(Paragraph(f"Период: {report_data.get('start_date', '')} - {report_data.get('end_date', '')}", normal_style))
        elements.append(Paragraph(f"Дата создания: {report_data.get('generated_at', '')}", normal_style))
        elements.append(Spacer(1, 12))
        
        # Добавляем основную статистику
        elements.append(Paragraph("Основная статистика", heading_style))
        
        # Создаем таблицу
        basic_stats = [
            ["Показатель", "Значение"],
            ["Всего звонков", str(report_data.get("total_calls", 0))],
            ["Входящие звонки", str(report_data.get("incoming_calls", 0))],
            ["Исходящие звонки", str(report_data.get("outgoing_calls", 0))],
            ["Средняя длительность звонка (мин)", str(report_data.get("avg_duration", 0))],
            ["Общая длительность всех звонков (мин)", str(report_data.get("total_duration", 0))]
        ]
        
        basic_table = Table(basic_stats, colWidths=[3*inch, 3*inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(basic_table)
        elements.append(Spacer(1, 12))
        
        # Добавляем статистику по дням
        elements.append(Paragraph("Статистика по дням", heading_style))
        
        # Создаем таблицу по дням
        daily_data = [
            ["Дата", "Всего", "Входящие", "Исходящие", "Ср. длит. (мин)"]
        ]
        
        for day in report_data.get("daily_calls", []):
            daily_data.append([
                day.get("date", ""),
                str(day.get("total", 0)),
                str(day.get("incoming", 0)),
                str(day.get("outgoing", 0)),
                str(day.get("avg_duration", 0))
            ])
        
        daily_table = Table(daily_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        daily_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(daily_table)
        elements.append(Spacer(1, 12))
        
        # Добавляем детальную статистику звонков
        elements.append(Paragraph("Детальная информация о звонках", heading_style))
        
        # Список звонков (ограничиваем первыми 50 для PDF)
        calls_data = [
            ["Дата", "Время", "Телефон", "Направление", "Длительность", "Источник"]
        ]
        
        for i, call in enumerate(report_data.get("detailed_calls", [])[:50]):  # Ограничиваем для PDF
            calls_data.append([
                call.get("date", ""),
                call.get("time", ""),
                call.get("phone", ""),
                call.get("direction", ""),
                f"{call.get('duration', 0)} мин",
                call.get("source", "")
            ])
        
        calls_table = Table(calls_data, colWidths=[1*inch, 0.8*inch, 1.3*inch, 1*inch, 1*inch, 2.3*inch])
        calls_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(calls_table)
        
        # Примечание
        if len(report_data.get("detailed_calls", [])) > 50:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"* Показаны первые 50 из {len(report_data.get('detailed_calls', []))} звонков. Полный список доступен в Excel-отчете.", normal_style))
        
        # Создаем PDF
        doc.build(elements)
    
    def _create_individual_excel(self, report_data: Dict[str, Any], output_path: str) -> None:
        """
        Создает Excel-отчет для индивидуальной статистики администратора
        
        :param report_data: Данные для отчета
        :param output_path: Путь сохранения Excel-файла
        """
        # Создаем рабочую книгу
        workbook = Workbook()
        
        # Основной лист со сводной информацией
        ws_summary = workbook.active
        ws_summary.title = "Общая информация"
        
        # Лист с данными по дням
        ws_daily = workbook.create_sheet("Статистика по дням")
        
        # Лист с детальными данными о звонках
        ws_calls = workbook.create_sheet("Список звонков")
        
        # 1. Заполняем лист со сводной информацией
        ws_summary.append(["Индивидуальный отчет администратора"])
        ws_summary.append(["Администратор:", report_data.get("administrator_name", "")])
        ws_summary.append(["Клиника:", report_data.get("clinic_name", "")])
        ws_summary.append(["Период:", f"{report_data.get('start_date', '')} - {report_data.get('end_date', '')}" ])
        ws_summary.append(["Дата создания:", report_data.get("generated_at", "")])
        ws_summary.append([])
        
        ws_summary.append(["Основная статистика"])
        ws_summary.append(["Показатель", "Значение"])
        ws_summary.append(["Всего звонков", report_data.get("total_calls", 0)])
        ws_summary.append(["Входящие звонки", report_data.get("incoming_calls", 0)])
        ws_summary.append(["Исходящие звонки", report_data.get("outgoing_calls", 0)])
        ws_summary.append(["Средняя длительность звонка (мин)", report_data.get("avg_duration", 0)])
        ws_summary.append(["Общая длительность всех звонков (мин)", report_data.get("total_duration", 0)])
        
        # 2. Заполняем лист с данными по дням
        ws_daily.append(["Статистика по дням"])
        ws_daily.append(["Дата", "Всего звонков", "Входящие", "Исходящие", "Средняя длительность (мин)"])
        
        for day in report_data.get("daily_calls", []):
            ws_daily.append([
                day.get("date", ""),
                day.get("total", 0),
                day.get("incoming", 0),
                day.get("outgoing", 0),
                day.get("avg_duration", 0)
            ])
        
        # 3. Заполняем лист с детальными данными о звонках
        ws_calls.append(["Список звонков"])
        ws_calls.append(["Дата", "Время", "Телефон", "Направление", "Длительность (мин)", "Источник"])
        
        for call in report_data.get("detailed_calls", []):
            ws_calls.append([
                call.get("date", ""),
                call.get("time", ""),
                call.get("phone", ""),
                call.get("direction", ""),
                call.get("duration", 0),
                call.get("source", "")
            ])
        
        # Форматирование
        # Заголовки
        for sheet in [ws_summary, ws_daily, ws_calls]:
            sheet.row_dimensions[1].height = 20
            sheet["A1"].font = Font(bold=True, size=14)
        
        # Автоматическая ширина колонок
        for sheet in [ws_summary, ws_daily, ws_calls]:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width
        
        # Сохраняем файл
        workbook.save(output_path)

    async def cleanup(self, max_age_hours: int = 24):
        """
        Очищает старые временные файлы отчетов
        
        :param max_age_hours: Максимальный возраст файлов в часах, после которого они удаляются
        """
        try:
            now = datetime.now()
            max_age = timedelta(hours=max_age_hours)
            
            # Очищаем временную директорию
            for root, dirs, files in os.walk(self.temp_dir):
                for dir_name in dirs:
                    dir_path = os.path.join(root, dir_name)
                    try:
                        # Проверяем возраст директории
                        dir_time = datetime.fromtimestamp(os.path.getmtime(dir_path))
                        if now - dir_time > max_age:
                            # Удаляем содержимое директории
                            for file in os.listdir(dir_path):
                                try:
                                    file_path = os.path.join(dir_path, file)
                                    if os.path.isfile(file_path):
                                        os.remove(file_path)
                                except Exception as e:
                                    logger.error(f"Ошибка при удалении файла {file}: {e}")
                            
                            # Удаляем директорию
                            os.rmdir(dir_path)
                            logger.info(f"Удалена старая директория отчета: {dir_path}")
                    except Exception as e:
                        logger.error(f"Ошибка при очистке директории {dir_path}: {e}")
            
            # Также очищаем старые файлы в корневой директории отчетов
            for file in os.listdir(self.reports_dir):
                file_path = os.path.join(self.reports_dir, file)
                try:
                    if os.path.isfile(file_path):
                        file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                        if now - file_time > max_age:
                            os.remove(file_path)
                            logger.info(f"Удален старый файл отчета: {file_path}")
                except Exception as e:
                    logger.error(f"Ошибка при удалении файла {file_path}: {e}")
                    
        except Exception as e:
            logger.error(f"Ошибка при очистке временных файлов: {e}")

    async def get_reports(
        self,
        clinic_id: Optional[str] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        report_type: Optional[str] = None,
        limit: int = 20,
        skip: int = 0,
    ) -> List[Dict[str, Any]]:
        """
        Получает список доступных отчетов из базы данных
        
        :param clinic_id: ID клиники для фильтрации
        :param start_date: Начальная дата для фильтрации в формате DD.MM.YYYY
        :param end_date: Конечная дата для фильтрации в формате DD.MM.YYYY
        :param report_type: Тип отчета для фильтрации
        :param limit: Максимальное количество отчетов для возврата
        :param skip: Количество отчетов для пропуска (для пагинации)
        :return: Список отчетов
        """
        try:
            # Формируем запрос
            query = {}
            
            # Фильтр по клинике
            if clinic_id:
                query["clinic_id"] = clinic_id
                
            # Фильтр по типу отчета
            if report_type:
                query["report_type"] = report_type
                
            # Фильтр по дате начала
            if start_date:
                try:
                    start_datetime = datetime.strptime(start_date, "%d.%m.%Y")
                    query["start_date"] = {"$gte": start_datetime.strftime("%d.%m.%Y")}
                except ValueError:
                    logger.warning(f"Некорректный формат даты начала: {start_date}")
                    
            # Фильтр по дате окончания
            if end_date:
                try:
                    end_datetime = datetime.strptime(end_date, "%d.%m.%Y")
                    query["end_date"] = {"$lte": end_datetime.strftime("%d.%m.%Y")}
                except ValueError:
                    logger.warning(f"Некорректный формат даты окончания: {end_date}")
                    
            # Получаем отчеты из базы данных
            reports = await self.db.call_reports.find(query).sort("generated_at", -1).skip(skip).limit(limit).to_list(length=limit)
            
            # Если отчетов нет в базе, попробуем собрать информацию из файловой системы
            if not reports:
                reports = await self._get_reports_from_filesystem(
                    clinic_id=clinic_id,
                    report_type=report_type,
                    limit=limit,
                    skip=skip
                )
                
            return reports
        except Exception as e:
            logger.error(f"Ошибка при получении списка отчетов: {e}")
            return []
            
    async def _get_reports_from_filesystem(
        self,
        clinic_id: Optional[str] = None,
        report_type: Optional[str] = None,
        limit: int = 20,
        skip: int = 0,
    ) -> List[Dict[str, Any]]:
        """
        Получает список отчетов из файловой системы, если они не найдены в базе данных
        
        :param clinic_id: ID клиники для фильтрации
        :param report_type: Тип отчета для фильтрации
        :param limit: Максимальное количество отчетов для возврата
        :param skip: Количество отчетов для пропуска (для пагинации)
        :return: Список отчетов
        """
        try:
            reports = []
            
            # Получаем список PDF-файлов в директории отчетов
            for file in os.listdir(self.reports_dir):
                if file.endswith(".pdf") and not os.path.isdir(os.path.join(self.reports_dir, file)):
                    # Извлекаем информацию из имени файла
                    try:
                        # Формат имени: {тип}_report_{клиника}_{дата_начала}_{дата_окончания}.pdf
                        parts = file.replace(".pdf", "").split("_")
                        
                        if len(parts) >= 5:
                            report_id = str(uuid4())
                            type_part = parts[0]
                            
                            # Определяем тип отчета
                            if type_part == "full":
                                file_report_type = "FULL"
                            elif type_part == "summary":
                                file_report_type = "SUMMARY"
                            elif type_part == "individual":
                                file_report_type = "INDIVIDUAL"
                            elif type_part == "analytics":
                                file_report_type = "ANALYTICS"
                            else:
                                file_report_type = "UNKNOWN"
                                
                            # Фильтрация по типу отчета, если указан
                            if report_type and file_report_type != report_type:
                                continue
                                
                            # Находим соответствующий Excel-файл
                            excel_file = file.replace(".pdf", ".xlsx")
                            pdf_link = f"/api/call/reports/{file}/download"
                            excel_link = f"/api/call/reports/{excel_file}/download" if os.path.exists(os.path.join(self.reports_dir, excel_file)) else ""
                            
                            # Определяем метаданные отчета
                            file_clinic_id = parts[2] if len(parts) > 2 else ""
                            
                            # Фильтрация по клинике, если указана
                            if clinic_id and file_clinic_id != clinic_id:
                                continue
                                
                            # Получаем дату создания файла как дату генерации отчета
                            file_path = os.path.join(self.reports_dir, file)
                            created_time = datetime.fromtimestamp(os.path.getctime(file_path))
                            
                            # Добавляем отчет в список
                            reports.append({
                                "_id": report_id,
                                "report_type": file_report_type,
                                "clinic_id": file_clinic_id,
                                "generated_at": created_time.strftime("%d.%m.%Y %H:%M:%S"),
                                "links": {
                                    "pdf": pdf_link,
                                    "excel": excel_link
                                }
                            })
                    except Exception as e:
                        logger.error(f"Ошибка при обработке файла отчета {file}: {e}")
                        
            # Сортируем по дате создания (сначала новые)
            reports.sort(key=lambda x: x.get("generated_at", ""), reverse=True)
            
            # Применяем пагинацию
            paginated_reports = reports[skip:skip+limit] if skip < len(reports) else []
            
            return paginated_reports
        except Exception as e:
            logger.error(f"Ошибка при получении списка отчетов из файловой системы: {e}")
            return []

    async def get_report(self, report_id: str) -> Optional[Dict[str, Any]]:
        """
        Получает информацию об отчете по его ID
        
        :param report_id: ID отчета
        :return: Данные отчета или None, если отчет не найден
        """
        try:
            # Проверяем, если ID является ObjectId
            try:
                object_id = ObjectId(report_id)
                report = await self.db.call_reports.find_one({"_id": object_id})
                if report:
                    return report
            except:
                pass
                
            # Проверяем как обычный строковый ID
            report = await self.db.call_reports.find_one({"_id": report_id})
            if report:
                return report
                
            # Если отчет не найден в базе данных, проверяем файловую систему
            # Обычно если ID был сгенерирован для отчета из файловой системы,
            # то он не будет найден повторно, но это нормальный механизм работы
            reports = await self._get_reports_from_filesystem()
            for report in reports:
                if str(report.get("_id", "")) == report_id:
                    return report
                    
            # Отчет не найден
            return None
        except Exception as e:
            logger.error(f"Ошибка при получении отчета по ID {report_id}: {e}")
            return None

    def _generate_charts(self, report_data: Dict[str, Any], temp_dir: str) -> Dict[str, str]:
        """
        Генерирует графики для отчета по звонкам
        
        :param report_data: Данные отчета
        :param temp_dir: Временная директория для сохранения графиков
        :return: Словарь с путями к графикам
        """
        chart_paths = {}
        
        try:
            # Настройка шрифтов и стилей для графиков
            plt.rcParams['font.size'] = 12
            plt.rcParams['axes.titlesize'] = 14
            plt.rcParams['axes.labelsize'] = 12
            plt.rcParams['legend.fontsize'] = 10
            plt.rcParams['figure.titlesize'] = 16
            
            # 1. График входящих и исходящих звонков
            fig, ax = plt.subplots(figsize=(10, 6))
            labels = ['Входящие', 'Исходящие']
            values = [report_data.get('incoming_calls', 0), report_data.get('outgoing_calls', 0)]
            colors = ['#4CAF50', '#2196F3']
            
            ax.bar(labels, values, color=colors)
            ax.set_title('Распределение входящих и исходящих звонков')
            ax.set_ylabel('Количество звонков')
            
            # Добавляем значения на столбцы
            for i, v in enumerate(values):
                ax.text(i, v + 0.5, str(v), ha='center')
            
            call_types_path = os.path.join(temp_dir, 'call_types.png')
            plt.tight_layout()
            plt.savefig(call_types_path, dpi=100)
            plt.close(fig)
            chart_paths['call_types'] = call_types_path
            
            # 2. Круговая диаграмма по источникам трафика
            if report_data.get('sources'):
                fig, ax = plt.subplots(figsize=(10, 8))
                
                # Получаем топ-5 источников по количеству звонков
                top_sources = sorted(report_data.get('sources', []), key=lambda x: x.get('count', 0), reverse=True)[:5]
                
                # Добавляем 'Другие' для всех остальных источников
                other_sources = report_data.get('sources', [])[5:] if len(report_data.get('sources', [])) > 5 else []
                other_count = sum(s.get('count', 0) for s in other_sources)
                
                labels = [s.get('source', 'Неизвестно')[:20] for s in top_sources]  # Ограничиваем длину названия
                if other_count > 0:
                    labels.append('Другие')
                
                values = [s.get('count', 0) for s in top_sources]
                if other_count > 0:
                    values.append(other_count)
                
                # Если все значения равны 0, показываем заглушку
                if sum(values) == 0:
                    ax.text(0.5, 0.5, 'Нет данных по источникам', ha='center', va='center', fontsize=14)
                    ax.axis('off')
                else:
                    # Генерируем цвета
                    colors = plt.cm.tab10(np.linspace(0, 1, len(labels)))
                    
                    # Рисуем круговую диаграмму
                    wedges, texts, autotexts = ax.pie(
                        values, 
                        labels=None,
                        autopct='%1.1f%%', 
                        startangle=90, 
                        colors=colors
                    )
                    
                    # Настраиваем отображение процентов
                    for autotext in autotexts:
                        autotext.set_fontsize(10)
                        autotext.set_color('white')
                    
                    # Добавляем легенду
                    ax.legend(wedges, labels, title="Источники", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
                    
                    ax.set_title('Распределение звонков по источникам трафика')
                
                sources_path = os.path.join(temp_dir, 'sources.png')
                plt.tight_layout()
                plt.savefig(sources_path, dpi=100)
                plt.close(fig)
                chart_paths['sources'] = sources_path
            
            # 3. График по дням за последние 4 недели
            if report_data.get('daily_stats'):
                fig, ax = plt.subplots(figsize=(12, 6))
                
                # Подготавливаем данные
                dates = []
                incoming = []
                outgoing = []
                
                # Сортируем по дате
                sorted_stats = sorted(
                    report_data.get('daily_stats', []), 
                    key=lambda x: datetime.strptime(x.get('date', '01.01.2000'), '%d.%m.%Y')
                )
                
                # Берем последние 4 недели (28 дней) или все, если меньше
                last_4_weeks = sorted_stats[-28:] if len(sorted_stats) > 28 else sorted_stats
                
                for day in last_4_weeks:
                    date_str = day.get('date', '')
                    dates.append(date_str)
                    incoming.append(day.get('incoming_calls', 0))
                    outgoing.append(day.get('outgoing_calls', 0))
                
                # Создаем группированную столбчатую диаграмму
                x = np.arange(len(dates))
                width = 0.35
                
                ax.bar(x - width/2, incoming, width, label='Входящие', color='#4CAF50')
                ax.bar(x + width/2, outgoing, width, label='Исходящие', color='#2196F3')
                
                # Добавляем подписи и легенду
                ax.set_title('Динамика звонков за последние 4 недели')
                ax.set_ylabel('Количество звонков')
                
                # Форматируем даты на оси X
                if len(dates) > 14:
                    # Если дат много, показываем только каждую 3-ю
                    ax.set_xticks(x[::3])
                    ax.set_xticklabels(dates[::3], rotation=45, ha='right')
                else:
                    ax.set_xticks(x)
                    ax.set_xticklabels(dates, rotation=45, ha='right')
                
                ax.legend()
                
                plt.tight_layout()
                weekly_path = os.path.join(temp_dir, 'weekly.png')
                plt.savefig(weekly_path, dpi=100)
                plt.close(fig)
                chart_paths['weekly'] = weekly_path
            
            # 4. График скорости обработки звонков (в минутах)
            if report_data.get('administrators'):
                fig, ax = plt.subplots(figsize=(10, 6))
                
                # Получаем данные по администраторам
                admin_names = [a.get('name', 'Неизвестно')[:15] for a in report_data.get('administrators', [])]
                avg_durations = [a.get('avg_duration', 0) / 60 for a in report_data.get('administrators', [])]  # переводим в минуты
                
                # Сортируем по средней продолжительности
                sorted_indices = np.argsort(avg_durations)[::-1]  # по убыванию
                admin_names = [admin_names[i] for i in sorted_indices]
                avg_durations = [avg_durations[i] for i in sorted_indices]
                
                # Ограничиваем количество администраторов для отображения
                max_admins = 10
                if len(admin_names) > max_admins:
                    admin_names = admin_names[:max_admins]
                    avg_durations = avg_durations[:max_admins]
                
                # Создаем горизонтальную столбчатую диаграмму
                y_pos = np.arange(len(admin_names))
                ax.barh(y_pos, avg_durations, align='center', color='#FF9800')
                ax.set_yticks(y_pos)
                ax.set_yticklabels(admin_names)
                
                # Добавляем подписи
                ax.set_xlabel('Средняя длительность звонка (мин)')
                ax.set_title('Скорость обработки звонков по администраторам')
                
                # Добавляем значения на столбцы
                for i, v in enumerate(avg_durations):
                    ax.text(v + 0.1, i, f"{v:.2f}", va='center')
                
                plt.tight_layout()
                speed_path = os.path.join(temp_dir, 'speed.png')
                plt.savefig(speed_path, dpi=100)
                plt.close(fig)
                chart_paths['speed'] = speed_path
            
            # 5. Конверсия по администраторам
            if report_data.get('administrators'):
                fig, ax = plt.subplots(figsize=(10, 6))
                
                # Получаем данные
                admin_names = [a.get('name', 'Неизвестно')[:15] for a in report_data.get('administrators', [])]
                conversions = [a.get('overall_conversion', 0) * 100 for a in report_data.get('administrators', [])]  # в проценты
                
                # Сортируем по конверсии
                sorted_indices = np.argsort(conversions)[::-1]  # по убыванию
                admin_names = [admin_names[i] for i in sorted_indices]
                conversions = [conversions[i] for i in sorted_indices]
                
                # Ограничиваем количество администраторов
                max_admins = 10
                if len(admin_names) > max_admins:
                    admin_names = admin_names[:max_admins]
                    conversions = conversions[:max_admins]
                
                # Создаем горизонтальную столбчатую диаграмму
                y_pos = np.arange(len(admin_names))
                ax.barh(y_pos, conversions, align='center', color='#9C27B0')
                ax.set_yticks(y_pos)
                ax.set_yticklabels(admin_names)
                
                # Добавляем подписи
                ax.set_xlabel('Конверсия (%)')
                ax.set_title('Конверсия по администраторам')
                
                # Добавляем значения на столбцы
                for i, v in enumerate(conversions):
                    ax.text(v + 0.5, i, f"{v:.1f}%", va='center')
                
                plt.tight_layout()
                conversion_path = os.path.join(temp_dir, 'conversion.png')
                plt.savefig(conversion_path, dpi=100)
                plt.close(fig)
                chart_paths['conversion'] = conversion_path
            
            # 6. FG% (First Goal) по администраторам
            if report_data.get('administrators'):
                fig, ax = plt.subplots(figsize=(10, 6))
                
                # Получаем данные
                admin_names = [a.get('name', 'Неизвестно')[:15] for a in report_data.get('administrators', [])]
                scores = [a.get('avg_score', 0) * 10 for a in report_data.get('administrators', [])]  # FG%
                
                # Сортируем по показателю
                sorted_indices = np.argsort(scores)[::-1]  # по убыванию
                admin_names = [admin_names[i] for i in sorted_indices]
                scores = [scores[i] for i in sorted_indices]
                
                # Ограничиваем количество администраторов
                max_admins = 10
                if len(admin_names) > max_admins:
                    admin_names = admin_names[:max_admins]
                    scores = scores[:max_admins]
                
                # Создаем горизонтальную столбчатую диаграмму
                y_pos = np.arange(len(admin_names))
                bars = ax.barh(y_pos, scores, align='center')
                
                # Раскрашиваем в зависимости от значения
                for i, bar in enumerate(bars):
                    if scores[i] >= 7:
                        bar.set_color('#4CAF50')  # хороший показатель - зеленый
                    elif scores[i] >= 5:
                        bar.set_color('#FFC107')  # средний показатель - желтый
                    else:
                        bar.set_color('#F44336')  # низкий показатель - красный
                
                ax.set_yticks(y_pos)
                ax.set_yticklabels(admin_names)
                
                # Добавляем подписи
                ax.set_xlabel('FG%')
                ax.set_title('First-touch Goal % (FG%) по администраторам')
                
                # Добавляем значения на столбцы
                for i, v in enumerate(scores):
                    ax.text(v + 0.2, i, f"{v:.1f}", va='center')
                
                plt.tight_layout()
                fg_path = os.path.join(temp_dir, 'fg.png')
                plt.savefig(fg_path, dpi=100)
                plt.close(fig)
                chart_paths['fg'] = fg_path
            
            # 7. График по дням недели
            if report_data.get('weekday_stats'):
                fig, ax = plt.subplots(figsize=(10, 6))
                
                # Получаем данные
                weekdays = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
                weekday_data = {day: {'total': 0, 'incoming': 0, 'outgoing': 0} for day in range(7)}
                
                # Заполняем данные по дням недели
                for day_data in report_data.get('weekday_stats', []):
                    day_num = day_data.get('weekday', 0)
                    if 0 <= day_num < 7:
                        weekday_data[day_num]['total'] = day_data.get('total_calls', 0)
                        weekday_data[day_num]['incoming'] = day_data.get('incoming_calls', 0)
                        weekday_data[day_num]['outgoing'] = day_data.get('outgoing_calls', 0)
                
                # Формируем данные для графика
                total_calls = [weekday_data[i]['total'] for i in range(7)]
                incoming = [weekday_data[i]['incoming'] for i in range(7)]
                outgoing = [weekday_data[i]['outgoing'] for i in range(7)]
                
                # Создаем группированную столбчатую диаграмму
                x = np.arange(len(weekdays))
                width = 0.3
                
                ax.bar(x - width, incoming, width, label='Входящие', color='#4CAF50')
                ax.bar(x, total_calls, width, label='Всего', color='#9C27B0')
                ax.bar(x + width, outgoing, width, label='Исходящие', color='#2196F3')
                
                # Добавляем подписи и легенду
                ax.set_title('Распределение звонков по дням недели')
                ax.set_ylabel('Количество звонков')
                ax.set_xticks(x)
                ax.set_xticklabels(weekdays, rotation=45, ha='right')
                
                ax.legend()
                
                plt.tight_layout()
                weekday_path = os.path.join(temp_dir, 'weekday.png')
                plt.savefig(weekday_path, dpi=100)
                plt.close(fig)
                chart_paths['weekday'] = weekday_path
            
            return chart_paths
        
        except Exception as e:
            logger.error(f"Ошибка при создании графиков: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return chart_paths

    def _create_category_distribution_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """
        Создает столбчатую диаграмму распределения звонков по категориям
        """
        plt.figure(figsize=(10, 6))
        
        # Ограничиваем количество категорий для читаемости
        top_categories = df.head(10) if len(df) > 10 else df
        
        ax = sns.barplot(
            x='counts', 
            y='call_category', 
            data=top_categories,
            palette='viridis'
        )
        
        # Добавляем значения на столбцы
        for i, v in enumerate(top_categories['counts']):
            ax.text(v + 0.5, i, str(v), color='white', va='center')
        
        plt.title('Распределение звонков по категориям', fontsize=16, color='white')
        plt.xlabel('Количество звонков', fontsize=12)
        plt.ylabel('Категория звонка', fontsize=12)
        plt.tight_layout()
        
        plt.savefig(buffer, format=format, dpi=100, bbox_inches='tight')
        plt.close()
        buffer.seek(0)

    def _create_weekday_distribution_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """
        Создает столбчатую диаграмму распределения звонков по дням недели
        """
        plt.figure(figsize=(12, 6))
        
        # Переводим названия дней недели на русский
        weekday_mapping = {
            'Monday': 'Понедельник',
            'Tuesday': 'Вторник',
            'Wednesday': 'Среда',
            'Thursday': 'Четверг',
            'Friday': 'Пятница',
            'Saturday': 'Суббота',
            'Sunday': 'Воскресенье'
        }
        
        # Создаем копию данных для локализации
        plot_df = df.copy()
        plot_df['weekday_ru'] = plot_df['weekday'].map(weekday_mapping)
        
        ax = sns.barplot(
            x='weekday_ru', 
            y='counts', 
            data=plot_df,
            palette='viridis'
        )
        
        # Добавляем значения над столбцами
        for i, v in enumerate(plot_df['counts']):
            ax.text(i, v + 0.5, str(v), ha='center', color='white')
        
        plt.title('Распределение звонков по дням недели', fontsize=16, color='white')
        plt.xlabel('День недели', fontsize=12)
        plt.ylabel('Количество звонков', fontsize=12)
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        plt.savefig(buffer, format=format, dpi=100, bbox_inches='tight')
        plt.close()
        buffer.seek(0)

    def _create_hourly_distribution_plot(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """
        Создает линейную диаграмму распределения звонков по часам дня
        """
        plt.figure(figsize=(14, 6))
        
        # Создаем полный диапазон часов (0-23)
        all_hours = pd.DataFrame({'hour': range(24)})
        
        # Объединяем с данными, заполняя пропущенные часы нулями
        plot_df = all_hours.merge(df, on='hour', how='left').fillna(0)
        
        # Сортируем по часам
        plot_df = plot_df.sort_values('hour')
        
        # Создаем график
        ax = sns.lineplot(
            x='hour', 
            y='counts', 
            data=plot_df,
            marker='o',
            color='#88c0d0',
            linewidth=3
        )
        
        # Добавляем значения к точкам
        for i, v in enumerate(plot_df['counts']):
            if v > 0:  # Показываем только ненулевые значения
                ax.text(plot_df['hour'].iloc[i], v + 0.5, str(int(v)), ha='center', color='white',
                      bbox=dict(boxstyle="round,pad=0.3", facecolor='#4c566a', alpha=0.7))
        
        plt.title('Распределение звонков по часам дня', fontsize=16, color='white')
        plt.xlabel('Час (0-23)', fontsize=12)
        plt.ylabel('Количество звонков', fontsize=12)
        plt.xticks(range(0, 24, 2))
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.tight_layout()
        
        plt.savefig(buffer, format=format, dpi=100, bbox_inches='tight')
        plt.close()
        buffer.seek(0)

    def _create_call_heatmap(self, df: pd.DataFrame, buffer: io.BytesIO, format: str):
        """
        Создает тепловую карту распределения звонков по дням недели и часам
        """
        plt.figure(figsize=(14, 8))
        
        # Создаем сводную таблицу: дни недели по вертикали, часы по горизонтали
        weekday_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        weekday_mapping = {
            'Monday': 'Понедельник',
            'Tuesday': 'Вторник',
            'Wednesday': 'Среда',
            'Thursday': 'Четверг',
            'Friday': 'Пятница',
            'Saturday': 'Суббота',
            'Sunday': 'Воскресенье'
        }
        
        # Создаем копию с русскими названиями дней недели
        df_copy = df.copy()
        df_copy['weekday_ru'] = df_copy['weekday'].map(weekday_mapping)
        
        # Создаем сводную таблицу
        heatmap_data = pd.pivot_table(
            df_copy, 
            values='administrator', 
            index='weekday_ru',
            columns='hour', 
            aggfunc='count',
            fill_value=0
        )
        
        # Сортируем дни недели
        heatmap_data = heatmap_data.reindex([weekday_mapping[day] for day in weekday_order])
        
        # Создаем тепловую карту с аннотациями
        ax = sns.heatmap(
            heatmap_data,
            annot=True,
            fmt="d",
            cmap="viridis",
            linewidths=.5,
            cbar_kws={'label': 'Количество звонков'}
        )
        
        plt.title('Распределение звонков по дням недели и часам', fontsize=16, color='white')
        plt.xlabel('Час дня', fontsize=12)
        plt.ylabel('День недели', fontsize=12)
        plt.tight_layout()
        
        # Настраиваем цвета для тепловой карты в темной теме
        ax.collections[0].colorbar.outline.set_edgecolor('white')
        ax.collections[0].colorbar.ax.yaxis.set_tick_params(color='white')
        ax.collections[0].colorbar.ax.tick_params(axis='y', colors='white')
        
        plt.savefig(buffer, format=format, dpi=100, bbox_inches='tight')
        plt.close()
        buffer.seek(0)


# Создаем экземпляр для использования в API
call_report_service = CallReportService()
