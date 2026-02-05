import os
from io import BytesIO
from datetime import datetime, timedelta, date
from motor.motor_asyncio import AsyncIOMotorClient
from typing import Dict, Any, List, Optional

from app.services.recommendation_analysis_service import RecommendationAnalysisService

import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import textwrap
import matplotlib.patheffects as pe

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak,
    Table, TableStyle, Flowable
)
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import inch

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
from openpyxl.styles import Alignment

from app.settings.paths import MONGO_URI, DB_NAME, REPORTS_DIR
import logging

logger = logging.getLogger(__name__)


class CallReportService:
    def __init__(self, mongodb_uri=None, mongodb_name=None):
        self.mongodb_uri = mongodb_uri or MONGO_URI
        self.mongodb_name = mongodb_name or DB_NAME
        self.client = None
        self.db = None
        self.calls_collection = None

        # light seaborn theme
        plt.style.use('default')
        sns.set_theme(style="whitegrid")

        # Попытка регистрации шрифтов Roboto
        font_dir = os.getenv("FONTS_DIR", os.path.join(os.path.dirname(__file__), "..", "fonts"))
        try:
            roboto_regular_path = os.path.join(font_dir, 'Roboto-Regular.ttf')
            roboto_bold_path = os.path.join(font_dir, 'Roboto-Bold.ttf')
            
            if os.path.exists(roboto_regular_path) and os.path.exists(roboto_bold_path):
                pdfmetrics.registerFont(TTFont('Roboto', roboto_regular_path))
                pdfmetrics.registerFont(TTFont('Roboto-Bold', roboto_bold_path))
                self.font_name = "Roboto"
                self.font_bold = "Roboto-Bold"
                logger.info("Шрифты Roboto успешно зарегистрированы")
            else:
                logger.warning(f"Файлы шрифтов не найдены: {roboto_regular_path}, {roboto_bold_path}")
                self.font_name = "Helvetica"
                self.font_bold = "Helvetica-Bold"
                logger.info("Будут использованы стандартные шрифты")
        except Exception as e:
            logger.error(f"Ошибка при регистрации шрифтов: {e}")
            self.font_name = "Helvetica"
            self.font_bold = "Helvetica-Bold"
            logger.info("Будут использованы стандартные шрифты")
        
        # Флаг, указывающий, доступны ли шрифты Roboto
        self.roboto_available = 'Roboto' in pdfmetrics.getRegisteredFontNames()

    async def connect_db(self):
        self.client = AsyncIOMotorClient(self.mongodb_uri)
        self.db = self.client[self.mongodb_name]
        self.calls_collection = self.db["calls"]
        logger.info("MongoDB connected")

    async def close_db(self):
        if self.client:
            self.client.close()

    async def get_calls_data(self, start_date, end_date, clinic_id=None):
        """
        Получение данных о звонках из базы данных MongoDB
        """
        if not self.client:
            await self.connect_db()

        # Подготовка запроса к базе
        # Форматируем даты для фильтрации в строковом формате YYYY-MM-DD
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')
        
        query = {
            "created_date_for_filtering": {"$gte": start_date_str, "$lte": end_date_str},
            "metrics.overall_score": {
                "$exists": True
            }
        }
        
        # Добавляем фильтр по клинике, если указан
        if clinic_id:
            query["client_id"] = clinic_id

        # Получаем данные из базы
        data = await self.calls_collection.find(query).to_list(length=None)
        logger.info(f"Fetched {len(data)} records with overall_score from Mongo")
        
        return data

    def create_dataframe(self, calls_data):
        """
        Преобразование данных MongoDB в pandas DataFrame с разворачиванием всех метрик
        """
        import pandas as pd
        df = pd.DataFrame(calls_data)
        if df.empty:
            return df

        # Преобразуем поля с датами в datetime
        if 'created_date_for_filtering' in df.columns:
            # Преобразуем строковые даты в формате YYYY-MM-DD в datetime для общих целей
            df['created_date'] = pd.to_datetime(df['created_date_for_filtering'])
            
            # Добавляем поле с номером недели и годом для анализа по неделям
            df['week'] = df['created_date'].dt.isocalendar().week
            df['year'] = df['created_date'].dt.year
        
        # Используем created_date_iso для точного времени (для тепловой карты)
        if 'created_date_iso' in df.columns:
            # Преобразуем ISO datetime в pandas datetime с учетом часового пояса
            df['created_date_with_time'] = pd.to_datetime(df['created_date_iso'])
            
            # Добавляем поле с днём недели и часом для тепловых карт на основе точного времени
            df['weekday'] = df['created_date_with_time'].dt.day_name()
            df['hour'] = df['created_date_with_time'].dt.hour
        elif 'created_date_for_filtering' in df.columns:
            # Fallback: если created_date_iso нет, используем created_date_for_filtering
            df['weekday'] = df['created_date'].dt.day_name()
            df['hour'] = df['created_date'].dt.hour
        
        # Мы используем только created_date_for_filtering, поэтому created_at не обрабатываем
        
        # Собираем все уникальные ключи из metrics по всей коллекции
        all_metrics = set()
        for row in calls_data:
            metrics = row.get('metrics', {})
            if isinstance(metrics, dict):
                all_metrics.update(metrics.keys())
        all_metrics = list(all_metrics)

        # Разворачиваем все метрики на верхний уровень
        for metric in all_metrics:
            df[metric] = df['metrics'].apply(
                lambda x: x.get(metric) if isinstance(x, dict) and metric in x else None
            )

        # Гарантируем наличие ключевых столбцов и правильные типы
        if 'overall_score' in df.columns:
            df['overall_score'] = pd.to_numeric(df['overall_score'], errors='coerce')
        else:
            df['overall_score'] = None
        # Всегда создаём conversion_int на основе conversion (0/1)
        if 'conversion' in df.columns:
            df['conversion'] = df['conversion'].apply(lambda x: bool(x) if pd.notnull(x) else False)
            df['conversion_int'] = df['conversion'].astype(int)
        else:
            df['conversion'] = False
            df['conversion_int'] = 0
        if 'call_type_classification' not in df.columns:
            df['call_type_classification'] = 'Неопределенный'
        else:
            df['call_type_classification'] = df['call_type_classification'].fillna('Неопределенный').astype(str)

        # Для удобства: если нужны другие метрики — добавлять аналогично
        return df

    def create_calls_by_admin_chart(self, df):
        plt.figure(figsize=(10, 6))
        palette = sns.color_palette('Set2', n_colors=df['administrator'].nunique())
        ax = sns.countplot(data=df, y='administrator', hue='administrator', legend=False, palette=palette)
        ax.set_title('Количество звонков по администраторам', fontsize=16)
        ax.set_xlabel('Количество звонков', fontsize=13)
        ax.set_ylabel('Администратор', fontsize=13)
        ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        for p in ax.patches:
            width = p.get_width()
            ax.text(width + 0.1, p.get_y() + p.get_height()/2, f'{int(width)}', ha='left', va='center', fontsize=12, fontweight='bold')
        plt.tight_layout()
        img_data = BytesIO()
        plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
        img_data.seek(0)
        plt.close()
        return img_data

    def create_duration_by_admin_chart(self, df):
        duration_by_admin = df.groupby('administrator')['duration'].sum() / 60
        duration_df = pd.DataFrame({'administrator': duration_by_admin.index, 'duration_minutes': duration_by_admin.values})
        plt.figure(figsize=(10, 6))
        palette = sns.color_palette('Set2', n_colors=duration_df['administrator'].nunique())
        ax = sns.barplot(data=duration_df, x='duration_minutes', y='administrator', hue='administrator', legend=False, palette=palette)
        ax.set_title('Общая длительность звонков по администраторам (минуты)', fontsize=16)
        ax.set_xlabel('Длительность (минуты)', fontsize=13)
        ax.set_ylabel('Администратор', fontsize=13)
        ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        for p in ax.patches:
            width = p.get_width()
            ax.text(width + 0.1, p.get_y() + p.get_height()/2, f'{int(width)} мин', ha='left', va='center', fontsize=12, fontweight='bold')
        plt.tight_layout()
        img_data = BytesIO()
        plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
        img_data.seek(0)
        plt.close()
        return img_data

    def create_traffic_pie_chart(self, df: pd.DataFrame) -> BytesIO:
        """
        Создание круговой диаграммы по источникам трафика с правильными пропорциями
        """
        buf = BytesIO()
        
        # Проверка наличия данных
        if df.empty or 'source' not in df.columns:
            plt.figure(figsize=(6, 6))
            plt.text(0.5, 0.5, "Нет данных о трафике", ha="center", va="center")
            plt.axis("off")
            plt.savefig(buf, format="png", dpi=150)
            buf.seek(0)
            plt.close()
            return buf
            
        try:
            # Подсчет количества звонков по источникам
            traffic_counts = df['source'].value_counts()
            
            # Создаем КВАДРАТНУЮ фигуру для правильного отображения круга
            plt.figure(figsize=(8, 8))
            
            # Создаем круговую диаграмму
            plt.pie(
                traffic_counts, 
                labels=traffic_counts.index, 
                autopct='%1.1f%%', 
                startangle=90,
                colors=sns.color_palette('viridis', len(traffic_counts))
            )
            
            # Критически важно для корректного отображения круга!
            plt.axis('equal')
            
            plt.title('Распределение звонков по источникам трафика', fontsize=16)
            plt.tight_layout()
            
            # Сохраняем график
            plt.savefig(buf, format="png", dpi=150)
            buf.seek(0)
            plt.close()
            return buf
            
        except Exception as e:
            logger.error(f"Ошибка при создании круговой диаграммы: {e}")
            plt.close()
            # Возвращаем пустой график при ошибке
            plt.figure(figsize=(6, 6))
            plt.text(0.5, 0.5, "Ошибка построения", ha="center", va="center")
            plt.axis("off")
            plt.savefig(buf, format="png", dpi=150)
            buf.seek(0)
            plt.close()
            return buf

    def create_calls_heatmap(self, df):
        """
        Создание тепловой карты звонков по дням недели и времени суток
        """
        # Создаем копию DataFrame, чтобы не изменять оригинал
        df_copy = df.copy()
        
        # Проверяем, есть ли столбцы с датами
        date_column = None
        if 'created_date_with_time' in df_copy.columns:
            # Приоритет у created_date_with_time (создается из created_date_iso)
            date_column = 'created_date_with_time'
        elif 'created_date' in df_copy.columns:
            # Fallback на created_date (создается из created_date_for_filtering)
            date_column = 'created_date'
        
        if date_column:
            # Используем поля weekday и hour, которые мы добавили в методе create_dataframe
            if 'weekday' not in df_copy.columns or 'hour' not in df_copy.columns:
                # Если полей нет, создаем их на основе доступного столбца с датой
                df_copy['weekday'] = df_copy[date_column].dt.day_name()
                df_copy['hour'] = df_copy[date_column].dt.hour
            
            # Словарь для перевода названий дней недели на русский
            days_translation = {
                'Monday': 'Понедельник',
                'Tuesday': 'Вторник',
                'Wednesday': 'Среда',
                'Thursday': 'Четверг',
                'Friday': 'Пятница',
                'Saturday': 'Суббота',
                'Sunday': 'Воскресенье'
            }
            
            # Переводим названия дней недели, если они на английском
            if df_copy['weekday'].iloc[0] in days_translation:
                df_copy['weekday'] = df_copy['weekday'].map(days_translation)
            
            # Создаем сводную таблицу для подсчета звонков по дням недели и часам
            heatmap_data = df_copy.pivot_table(
                index='weekday', 
                columns='hour', 
                values='_id', 
                aggfunc='count',
                fill_value=0
            )

            # Сортируем дни недели в правильном порядке
            day_order = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
            heatmap_data = heatmap_data.reindex(day_order)
            
            # Заполняем NaN нулями
            heatmap_data = heatmap_data.fillna(0)
            
            # Создаем тепловую карту
            plt.figure(figsize=(14, 8))
            ax = sns.heatmap(
                heatmap_data, 
                cmap='viridis', 
                linewidths=0.5, 
                annot=True, 
                fmt='g',  # Общий формат для int и float
                cbar_kws={'label': 'Количество звонков'}
            )
            # Сделать деления colorbar только по целым значениям
            colorbar = ax.collections[0].colorbar
            vmin, vmax = int(heatmap_data.values.min()), int(heatmap_data.values.max())
            if vmax > vmin:  # Избегаем ошибки, если все значения одинаковые
                colorbar.set_ticks(np.arange(vmin, vmax + 1, max(1, (vmax - vmin) // 10)))
            
            plt.title('Распределение звонков по дням недели и времени суток', fontsize=16)
            plt.xlabel('Час дня', fontsize=14)
            plt.ylabel('День недели', fontsize=14)
            
            # Сохраняем график в BytesIO
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
            img_data.seek(0)
            plt.close()
            
            return img_data
        else:
            logger.warning("Столбцы с датами не найдены в данных")
            return None

    def create_traffic_line_chart(self, df):
        """
        Создание линейного графика звонков по источникам трафика
        """
        # Проверяем наличие необходимых полей
        if 'created_date' not in df.columns or 'source' not in df.columns:
            logger.warning("Отсутствуют обязательные поля для графика по трафику")
            return None
        
        try:
            # Копируем DataFrame для обработки
            df_copy = df.copy()
            
            # Используем поле created_date, которое уже преобразовано в datetime в методе create_dataframe
            # Добавляем поле с датой (без времени) для группировки, если оно не было создано ранее
            if 'date' not in df_copy.columns:
                df_copy['date'] = df_copy['created_date'].dt.date
            
            # Выбираем топ-5 источников по количеству звонков
            top_sources = df_copy['source'].value_counts().nlargest(5).index.tolist()
            logger.debug(f"Топ-5 источников: {top_sources}")
            
            # Фильтруем данные только по топ источникам
            df_filtered = df_copy[df_copy['source'].isin(top_sources)]
            
            # Группируем по источнику и дате, считаем количество звонков
            traffic_by_date = df_filtered.groupby(['source', 'date']).size().reset_index(name='count')
            
            # Создаем широкий формат данных для построения графика
            pivot_traffic = traffic_by_date.pivot(index='date', columns='source', values='count')
            
            # Заполняем NaN нулями
            pivot_traffic = pivot_traffic.fillna(0)
            
            # Если дат слишком много, выбираем только 5 равномерно распределенных дат
            if len(pivot_traffic) > 5:
                # Выбираем индексы для 5 дат
                selected_indices = np.linspace(0, len(pivot_traffic)-1, 5, dtype=int)
                pivot_traffic = pivot_traffic.iloc[selected_indices]
            
            # Создаем график
            plt.figure(figsize=(12, 7))
            
            # Подготавливаем форматированные даты для отображения на оси X (в формате ДД.ММ.ГГГГ)
            formatted_dates = [date.strftime('%d.%m.%Y') for date in pivot_traffic.index]
            
            # Строим линии для каждого источника, используя позиции на оси X
            x_positions = range(len(pivot_traffic.index))
            for source in pivot_traffic.columns:
                plt.plot(x_positions, pivot_traffic[source].astype(int), marker='o', linewidth=2, label=source)
            
            # Настраиваем график
            plt.title('Звонки по источнику трафика', fontsize=16)
            plt.xlabel('Дата', fontsize=12)
            plt.ylabel('Количество звонков', fontsize=12)
            plt.grid(True, alpha=0.3)
            plt.legend(fontsize=10)
            
            # Устанавливаем отформатированные метки дат на оси X
            plt.xticks(x_positions, formatted_dates, rotation=45)
            
            # Настраиваем ось Y для отображения только целых чисел
            ax = plt.gca()
            ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
            ax.yaxis.set_major_formatter(ticker.FormatStrFormatter('%d'))
            
            # Принудительно выставить лимиты оси Y, если все значения одинаковые
            ymin, ymax = ax.get_ylim()
            if ymax - ymin < 2:
                ax.set_ylim(0, max(2, int(ymax) + 1))
            
            # Добавляем отступы
            plt.tight_layout()
            
            # Сохраняем график в BytesIO
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
            img_data.seek(0)
            plt.close()
            
            return img_data
        
        except Exception as e:
            logger.error(f"Ошибка при создании графика звонков по трафику: {e}")
            return None

    def create_call_types_line_chart(self, df):
        """
        Создание линейного графика звонков по типам
        """
        try:
            # Копируем DataFrame для обработки
            df_copy = df.copy()
            
            # Проверяем и добавляем поле типа звонка, если его нет
            if 'call_type_classification' not in df_copy.columns:
                logger.debug("Поле call_type_classification не найдено напрямую, пробуем извлечь из metrics...")
                # Пробуем извлечь из metrics, если оно есть
                if 'metrics' in df_copy.columns:
                    df_copy['call_type_classification'] = df_copy['metrics'].apply(
                        lambda x: x.get('call_type_classification', 'Неопределенный') if isinstance(x, dict) else 'Неопределенный'
                    )
            
            # Проверяем, есть ли теперь поле call_type_classification
            if 'call_type_classification' not in df_copy.columns or df_copy['call_type_classification'].count() == 0:
                logger.warning("Не удалось получить данные о типах звонков")
                return None
            
            # Заполняем пропущенные значения
            df_copy['call_type_classification'] = df_copy['call_type_classification'].fillna('Неопределенный')
            
            # Проверяем наличие поля created_date_for_filtering и используем его вместо created_at
            if 'created_date_for_filtering' in df_copy.columns:
                logger.debug("Используем поле created_date_for_filtering для отображения дат на графике")
                # Преобразуем строки в даты
                df_copy['date'] = pd.to_datetime(df_copy['created_date_for_filtering']).dt.date
            else:
                # Если нет created_date_for_filtering, пробуем использовать created_at
                logger.debug("Поле created_date_for_filtering не найдено, используем created_at")
                # Убедимся, что 'created_at' в формате datetime
                if not pd.api.types.is_datetime64_any_dtype(df_copy['created_at']):
                    df_copy['created_at'] = pd.to_datetime(df_copy['created_at'], unit='s')
                
                # Добавляем поле с датой (без времени) для группировки
                df_copy['date'] = df_copy['created_at'].dt.date
            
            # Известные типы звонков
            known_types = ['Неопределенный', 'первичка', 'вторичка', 'перезвон', 'подтверждение']
            existing_types = df_copy['call_type_classification'].unique()
            
            # Выбираем типы для отображения на графике
            types_count = df_copy['call_type_classification'].value_counts()
            if len(types_count) <= 5:
                top_types = types_count.index.tolist()
            else:
                # Если типов более 5, берем самые частые, но обязательно включаем "подтверждение" и "первичка"
                top_types = types_count.nlargest(5).index.tolist()
                # Убеждаемся, что важные типы включены
                for important_type in ['подтверждение', 'первичка']:
                    if important_type in existing_types and important_type not in top_types:
                        # Заменяем наименее часто встречающийся тип
                        top_types[-1] = important_type
            
            # Фильтруем данные только по выбранным типам
            df_filtered = df_copy[df_copy['call_type_classification'].isin(top_types)]
            
            # Если после фильтрации не осталось данных, возвращаем None
            if len(df_filtered) == 0:
                logger.warning("После фильтрации не осталось данных для построения графика")
                return None
            
            # Группируем по типу и дате, считаем количество звонков
            types_by_date = df_filtered.groupby(['call_type_classification', 'date']).size().reset_index(name='count')
            
            # Создаем широкий формат данных для построения графика
            pivot_types = types_by_date.pivot(index='date', columns='call_type_classification', values='count')
            
            # Получаем минимальную и максимальную дату
            min_date = df_filtered['date'].min()
            max_date = df_filtered['date'].max()
            
            # Создаем полный диапазон дат
            full_date_range = pd.date_range(start=min_date, end=max_date, freq='D')
            
            # Создаем новый DataFrame с полным диапазоном дат
            complete_pivot = pd.DataFrame(index=full_date_range.date)
            
            # Копируем данные из исходной pivot-таблицы
            for column in pivot_types.columns:
                complete_pivot[column] = pivot_types.get(column, 0)
            
            # Заполняем отсутствующие значения нулями
            complete_pivot = complete_pivot.fillna(0)
            
            # Используем полную таблицу вместо исходной pivot-таблицы
            pivot_types = complete_pivot
            
            # Проверяем, достаточно ли данных для построения графика
            if len(pivot_types) < 2:
                logger.warning("Недостаточно дат для построения графика")
                return None
            
            # Сортируем даты
            pivot_types = pivot_types.sort_index()
            
            # Создаем график
            plt.figure(figsize=(14, 8))  # Увеличиваем размер для большей читаемости
            
            # Подготавливаем форматированные даты для отображения на оси X (в формате ДД.ММ.ГГГГ)
            formatted_dates = [date.strftime('%d.%m.%Y') for date in pivot_types.index]
            
            # Строим линии для каждого типа звонка, используя позиции на оси X
            x_positions = range(len(pivot_types.index))
            for call_type in pivot_types.columns:
                plt.plot(x_positions, pivot_types[call_type].astype(int), marker='o', linewidth=2, label=call_type)
            
            # Настраиваем график
            plt.title('Звонки по типу', fontsize=16)
            plt.xlabel('Дата', fontsize=12)
            plt.ylabel('Количество звонков', fontsize=12)
            plt.grid(True, alpha=0.3)
            plt.legend(fontsize=10)
            
            # Устанавливаем отформатированные метки дат на оси X
            # Если дат много, показываем каждую вторую метку для лучшей читаемости
            if len(formatted_dates) > 10:
                every_nth = 2
                plt.xticks(x_positions[::every_nth], formatted_dates[::every_nth], rotation=45)
            else:
                plt.xticks(x_positions, formatted_dates, rotation=45)
            
            # Настраиваем ось Y для отображения только целых чисел
            ax = plt.gca()
            ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
            ax.yaxis.set_major_formatter(ticker.FormatStrFormatter('%d'))
            
            # Принудительно выставить лимиты оси Y, если все значения одинаковые
            ymin, ymax = ax.get_ylim()
            if ymax - ymin < 2:
                ax.set_ylim(0, max(2, int(ymax) + 1))
            
            # Добавляем отступы
            plt.tight_layout()
            
            # Сохраняем график в BytesIO
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
            img_data.seek(0)
            plt.close()
            
            return img_data
        
        except Exception as e:
            logger.error(f"Ошибка при создании графика звонков по типам: {e}")
            import traceback
            traceback.print_exc()
            return None

    def create_summary_statistics(self, df, start_date=None, end_date=None):
        """
        Создание сводной статистики
        """
        # Используем явно переданные даты периода, если они заданы
        if start_date and end_date:
            date_range = f"{start_date} - {end_date}"
        elif 'created_date_for_filtering' in df.columns:
            try:
                # Используем created_date_for_filtering для определения диапазона дат
                df_dates = pd.to_datetime(df['created_date_for_filtering'])
                min_date = df_dates.min()
                max_date = df_dates.max()
                start_date_str = min_date.strftime('%d.%m.%Y')
                end_date_str = max_date.strftime('%d.%m.%Y')
                date_range = f"{start_date_str} - {end_date_str}"
            except Exception as e:
                logger.error(f"Ошибка при форматировании дат created_date_for_filtering: {e}")
                date_range = "Неизвестный период"
        else:
            date_range = "Неизвестный период"
        
        # Общее количество звонков
        total_calls = len(df)
        
        # Логируем длительности звонков для отладки
        logger.debug(f"Длительности звонков (сек): {df['duration'].tolist()}")
        
        # Средняя длительность звонка в минутах и секундах
        avg_duration_sec = df['duration'].mean()
        avg_min = int(avg_duration_sec // 60)
        avg_sec = int(avg_duration_sec % 60)
        avg_duration_str = f"{avg_min}:{avg_sec:02d}"
        
        # Общая длительность всех звонков в минутах и секундах
        total_duration_sec = df['duration'].sum()
        total_min = int(total_duration_sec // 60)
        total_sec = int(total_duration_sec % 60)
        total_duration_str = f"{total_min}:{total_sec:02d}"
        
        logger.debug(f"Средняя длительность: {avg_duration_sec} сек = {avg_duration_str}")
        logger.debug(f"Общая длительность: {total_duration_sec} сек = {total_duration_str}")
        
        # Количество входящих и исходящих звонков
        incoming_calls = len(df[df['call_direction'] == 'Входящий'])
        outgoing_calls = len(df[df['call_direction'] == 'Исходящий'])
        
        # Расчет процента конверсии в запись (оставляем как есть, если данные доступны)
        conversion_percentage = None
        try:
            if 'conversion' in df.columns:
                # Подсчитываем количество True значений
                converted_calls = df['conversion'].sum() if df['conversion'].dtype == 'bool' else len(df[df['conversion'] == True])
                if total_calls > 0:
                    conversion_percentage = (converted_calls / total_calls) * 100
        except Exception as e:
            logger.error(f"Ошибка при расчете конверсии: {e}")
        
        # Расчет FG% звонков на основе overall_score
        fg_percentage = None
        try:
            if 'overall_score' in df.columns:
                # По формуле: FG% = (Сумма всех оценок overall_score) / (Количество оцененных звонков) * 10
                total_overall_score = df['overall_score'].sum()
                scored_calls_count = df[df['overall_score'] > 0].shape[0]
                
                if scored_calls_count > 0:
                    fg_percentage = (total_overall_score / scored_calls_count) * 10
        except Exception as e:
            logger.error(f"Ошибка при расчете FG%: {e}")
        
        # Расчет средней скорости обработки в минутах
        avg_processing_speed = None
        try:
            if 'processing_speed' in df.columns:
                avg_processing_speed = df['processing_speed'].mean()
        except Exception as e:
            logger.error(f"Ошибка при расчете средней скорости обработки: {e}")
        
        # Создаем сводную таблицу статистики
        summary_data = [
            ["Период анализа", date_range],
            ["Общее количество звонков", str(total_calls)],
            ["Входящие звонки", str(incoming_calls)],
            ["Исходящие звонки", str(outgoing_calls)],
            ["Средняя длительность звонка", avg_duration_str],
            ["Общая длительность всех звонков", total_duration_str]
        ]
        
        # Добавляем процент конверсии если данные доступны
        if conversion_percentage is not None:
            summary_data.append(["Конверсия в запись", f"{round(conversion_percentage)}%"])
        
        # Добавляем FG% если данные доступны
        if fg_percentage is not None:
            summary_data.append(["FG% звонков", f"{round(fg_percentage)}%"])
        
        # Добавляем среднюю скорость обработки если данные доступны
        if avg_processing_speed is not None:
            summary_data.append(["Средняя скорость обработки", f"{avg_processing_speed:.2f} мин"])
        
        return summary_data

    def get_criterion_display_name(self, criterion_key):
        """
        Вспомогательная функция для получения отображаемых имен критериев
        """
        display_names = {
            'greeting': 'Приветствие',
            'patient_name': 'Имя пациента',
            'needs_identification': 'Выявление потребностей',
            'service_presentation': 'Презентация услуги',
            'clinic_presentation': 'Презентация клиники',
            'doctor_presentation': 'Презентация врача',
            'patient_booking': 'Запись',
            'clinic_address': 'Адрес клиники',
            'passport': 'Паспорт',
            'price': 'Цена',
            'expertise': 'Экспертность',
            'next_step': 'Следующий шаг',
            'appointment': 'Запись на прием',
            'emotional_tone': 'Эмоциональный окрас',
            'speech': 'Речь',
            'initiative': 'Инициатива',
            # Новые/английские ключи:
            'appeal': 'Апелляция',
            'objection_handling': 'Работа с возражениями',
            'question_clarification': 'Уточнение вопроса',
            'communication': 'Коммуникация',
        }
        return display_names.get(criterion_key, criterion_key)

    def create_traffic_conversion_chart(self, df):
        """
        Создание графика конверсии по источникам трафика
        """
        try:
            # Проверяем наличие необходимых колонок
            if 'source' not in df.columns or 'conversion_int' not in df.columns:
                logger.warning("Отсутствуют необходимые поля для графика конверсии по трафику")
                return None
                
            # Группируем данные по источникам трафика
            grouped = df.groupby('source').agg(
                total_calls=('conversion_int', 'count'),
                converted_calls=('conversion_int', 'sum')
            )
            
            # Исключаем источники с нулевым количеством звонков
            grouped = grouped[grouped['total_calls'] > 0]
            
            # Вычисляем проценты конверсии и округляем до целых чисел
            grouped['conversion_pct'] = (grouped['converted_calls'] / grouped['total_calls'] * 100).round(0).astype(int)
            
            # Сортируем по проценту конверсии в убывающем порядке и берем топ-6
            grouped = grouped.sort_values('conversion_pct', ascending=False)
            grouped = grouped.head(6)
            
            # Если у нас нет данных для построения графика
            if len(grouped) == 0:
                logger.warning("Нет данных для построения графика конверсии по трафику")
                return None
                
            # Создаем фигуру подходящего размера
            plt.figure(figsize=(8, 5))
            ax = plt.gca()
            
            # Определяем цветовую схему
            colors = sns.color_palette("viridis", len(grouped))
            
            # Выбираем тип графика в зависимости от количества источников
            if len(grouped) < 3:
                # Если мало категорий — barplot с более ярким оформлением
                bars = ax.bar(grouped.index, grouped['conversion_pct'], color=colors)
                
                # Добавляем значения на столбцы
                for i, (bar, y) in enumerate(zip(bars, grouped['conversion_pct'])):
                    ax.text(
                        bar.get_x() + bar.get_width()/2, 
                        y + 1, 
                        f"{int(y)}%", 
                        ha='center', 
                        va='bottom', 
                        fontsize=12, 
                        fontweight='bold',
                        color='black'
                    )
            else:
                # Если больше 2 источников — line plot
                x_positions = range(len(grouped.index))
                
                # Создаем линейный график с точками и подписями
                ax.plot(x_positions, grouped['conversion_pct'], marker='o', linewidth=2, color=colors[0])
                
                # Добавляем точки и подписи
                for i, (x, y) in enumerate(zip(x_positions, grouped['conversion_pct'])):
                    # Точки делаем более заметными
                    ax.scatter(x, y, s=80, color=colors[i % len(colors)], zorder=5)
                    
                    # Смещение подписи: если 100% — внутрь, иначе чуть выше
                    offset = -7 if y > 95 else 2
                    ax.text(
                        x, 
                        y + offset, 
                        f"{int(y)}%", 
                        ha='center', 
                        va='bottom', 
                        fontsize=12, 
                        fontweight='bold',
                        color='black'
                    )
                
                # Устанавливаем метки на оси X
                ax.set_xticks(x_positions)
                ax.set_xticklabels(grouped.index, rotation=30, ha='right')
            
            # Настройка осей Y
            ax.set_ylim(0, 110)
                
            # Добавляем сетку для лучшей читаемости
            ax.grid(True, alpha=0.3, linestyle='--')
            
            # Добавляем подписи и заголовок
            ax.set_ylabel("Конверсия в запись, %")
            ax.set_title("Конверсия трафик", fontsize=14, fontweight='bold')
            
            # Добавляем значения конверсии на графике
            for i, (source, row) in enumerate(grouped.iterrows()):
                if i < len(grouped):
                    ax.text(
                        i,
                        row['conversion_pct'] / 2,
                        f"{int(row['converted_calls'])}/{int(row['total_calls'])}",
                        ha='center',
                        va='center',
                        color='white' if row['conversion_pct'] > 30 else 'black',
                        fontsize=9,
                        fontweight='bold'
                    )
            
            plt.tight_layout()
            
            # Сохраняем график в BytesIO
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
            img_data.seek(0)
            plt.close()
            
            return img_data
            
        except Exception as e:
            logger.error(f"Ошибка при создании графика конверсии по трафику: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def create_call_type_conversion_chart(self, df):
        """
        Создание графика конверсии по типам звонков
        """
        try:
            # Проверяем наличие необходимых колонок
            if 'conversion_int' not in df.columns:
                logger.warning("Отсутствует поле conversion_int для графика конверсии по типам звонков")
                return None
            
            # Получаем данные о типах звонков
            if 'call_type_classification' not in df.columns:
                # Пробуем извлечь из metrics, если это возможно
                if 'metrics' in df.columns:
                    df = df.copy()  # Создаем копию для безопасного изменения
                    df['call_type_classification'] = df['metrics'].apply(
                        lambda x: x.get('call_type_classification', 'Неопределенный') if isinstance(x, dict) else 'Неопределенный'
                    )
                else:
                    logger.warning("Не найдены данные о типах звонков")
                    return None
            
            # Проверяем, что данные о типах звонков не пустые
            if df['call_type_classification'].isna().all():
                logger.warning("Все значения типов звонков пустые")
                return None
                
            # Группируем данные по типам звонков
            grouped = df.groupby('call_type_classification').agg(
                total_calls=('conversion_int', 'count'),
                converted_calls=('conversion_int', 'sum')
            )
            
            # Исключаем типы с нулевым количеством звонков
            grouped = grouped[grouped['total_calls'] > 0]
            
            # Вычисляем проценты конверсии и округляем до целых чисел
            grouped['conversion_pct'] = (grouped['converted_calls'] / grouped['total_calls'] * 100).round(0).astype(int)
            
            # Сортируем по проценту конверсии в убывающем порядке и берем топ-5
            grouped = grouped.sort_values('conversion_pct', ascending=False)
            grouped = grouped.head(5)
            
            # Если у нас нет данных для построения графика
            if len(grouped) == 0:
                logger.warning("Нет данных для построения графика конверсии по типам звонков")
                return None
                
            # Создаем фигуру подходящего размера
            plt.figure(figsize=(8, 5))
            ax = plt.gca()
            
            # Определяем цветовую схему
            colors = sns.color_palette("viridis", len(grouped))
            
            # Выбираем тип графика в зависимости от количества типов
            if len(grouped) < 3:
                # Если мало категорий — barplot
                bars = ax.bar(grouped.index, grouped['conversion_pct'], color=colors)
                
                # Добавляем значения на столбцы
                for i, (bar, y) in enumerate(zip(bars, grouped['conversion_pct'])):
                    ax.text(
                        bar.get_x() + bar.get_width()/2, 
                        y + 1, 
                        f"{int(y)}%", 
                        ha='center', 
                        va='bottom', 
                        fontsize=12, 
                        fontweight='bold',
                        color='black'
                    )
            else:
                # Если больше 2 типов — line plot
                x_positions = range(len(grouped.index))
                
                # Создаем линейный график с точками и подписями
                ax.plot(x_positions, grouped['conversion_pct'], marker='o', linewidth=2, color=colors[0])
                
                # Добавляем точки и подписи
                for i, (x, y) in enumerate(zip(x_positions, grouped['conversion_pct'])):
                    # Точки делаем более заметными
                    ax.scatter(x, y, s=80, color=colors[i % len(colors)], zorder=5)
                    
                    # Смещение подписи: если 100% — внутрь, иначе чуть выше
                    offset = -7 if y > 95 else 2
                    ax.text(
                        x, 
                        y + offset, 
                        f"{int(y)}%", 
                        ha='center', 
                        va='bottom', 
                        fontsize=12, 
                        fontweight='bold',
                        color='black'
                    )
                
                # Устанавливаем метки на оси X
                ax.set_xticks(x_positions)
                ax.set_xticklabels(grouped.index, rotation=30, ha='right')
            
            # Настройка осей Y
            ax.set_ylim(0, 110)
                
            # Добавляем сетку для лучшей читаемости
            ax.grid(True, alpha=0.3, linestyle='--')
            
            # Добавляем подписи и заголовок
            ax.set_ylabel("Конверсия в запись, %")
            ax.set_title("Конверсия тип звонка", fontsize=14, fontweight='bold')
            
            # Добавляем значения количества звонков на графике
            for i, (call_type, row) in enumerate(grouped.iterrows()):
                if i < len(grouped):
                    ax.text(
                        i,
                        row['conversion_pct'] / 2,
                        f"{int(row['converted_calls'])}/{int(row['total_calls'])}",
                        ha='center',
                        va='center',
                        color='white' if row['conversion_pct'] > 30 else 'black',
                        fontsize=9,
                        fontweight='bold'
                    )
            
            plt.tight_layout()
            
            # Сохраняем график в BytesIO
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
            img_data.seek(0)
            plt.close()
            
            return img_data
            
        except Exception as e:
            logger.error(f"Ошибка при создании графика конверсии по типам звонков: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def create_admin_table(self, df):
        try:
            grp = df.groupby("administrator")
            total = grp.size().sort_values(ascending=False)
            fg_pct = []
            for admin in total.index:
                group = grp.get_group(admin)
                total_score = group['overall_score'].sum()
                scored_calls = group[group['overall_score'] > 0].shape[0]
                if scored_calls > 0:
                    admin_fg = (total_score / scored_calls) * 10
                else:
                    admin_fg = 0
                fg_pct.append(admin_fg)
            fg_pct = pd.Series(fg_pct, index=total.index).round().astype(int)
            table = [["Администратор", "Звонков", "FG %"]]
            for adm in total.index:
                table.append([adm, int(total[adm]), f"{int(fg_pct[adm])}%"])
            return table
        except Exception as e:
            logger.error(f"Ошибка при создании таблицы администраторов: {e}")
            return [["Администратор", "Звонков", "FG %"], ["Нет данных", "0", "0%"]]

    def get_week_date_range_label(self, year, week):
        """
        Возвращает строку с номером недели в формате DataLens: 'Неделя 34'
        """
        return f"Неделя {week}"

    def create_weekly_scores_table(self, df):
        call_types_metrics = self.get_call_types_and_metrics()
        tables = []
        # Определяем границы периода по данным
        if 'created_date' in df.columns:
            min_date = df['created_date'].min()
            max_date = df['created_date'].max()
        else:
            min_date = max_date = None
        for call_type, metrics in call_types_metrics.items():
            df_type = df[df['call_type_classification'] == call_type]
            if df_type.empty:
                continue
            # Используем created_date вместо created_at
            if 'created_date' not in df_type.columns:
                return tables
            # Проверяем, существуют ли поля week и year в данных
            if 'week' not in df_type.columns or 'year' not in df_type.columns:
                df_type['week'] = df_type['created_date'].dt.isocalendar().week
                df_type['year'] = df_type['created_date'].dt.year

            # Фильтруем метрики - оставляем только те, которые существуют в DataFrame
            available_metrics = [m for m in metrics if m in df_type.columns]
            if not available_metrics:
                logger.warning(f"Нет доступных метрик для типа звонка '{call_type}', пропускаем")
                continue

            weekly = df_type.groupby(['year', 'week'])[available_metrics].mean().reset_index()
            if weekly.empty:
                continue
            metrics_groups = [available_metrics[i:i + 8] for i in range(0, len(available_metrics), 8)]
            for group_idx, metrics_group in enumerate(metrics_groups):
                group_suffix = f" (часть {group_idx + 1})" if len(metrics_groups) > 1 else ""
                header = ["" + group_suffix] + [self.get_criterion_display_name(m) for m in metrics_group]
                table_data = [header]
                for _, row in weekly.iterrows():
                    week_label = self.get_week_date_range_label(int(row['year']), int(row['week']))
                    row_data = [week_label]
                    for m in metrics_group:
                        value = row[m] if m in row and pd.notnull(row[m]) else 0
                        row_data.append(f"{round(value)}")
                    table_data.append(row_data)
                table_title = f"{call_type}{group_suffix}"
                tables.append((table_title, table_data))
        return tables

    def get_call_types_and_metrics(self):
        """
        Возвращает словарь: {тип_звонка: [метрики этого типа]}
        """
        return {
            'первичка': [
                'greeting', 'patient_name', 'needs_identification', 'service_presentation', 'clinic_presentation',
                'doctor_presentation', 'appointment', 'price', 'expertise', 'next_step', 'patient_booking',
                'emotional_tone', 'speech', 'initiative', 'clinic_address', 'passport', 'objection_handling'
            ],
            'вторичка': [
                'greeting', 'patient_name', 'question_clarification', 'expertise', 'next_step', 'patient_booking',
                'emotional_tone', 'speech', 'initiative', 'objection_handling'
            ],
            'перезвон': [
                'greeting', 'patient_name', 'appeal', 'next_step', 'initiative', 'speech',
                'clinic_address', 'passport', 'objection_handling'
            ],
            'подтверждение': [
                'greeting', 'patient_name', 'appeal', 'next_step', 'initiative', 'speech',
                'clinic_address', 'passport', 'objection_handling'
            ],
            'прочее': [
                'greeting', 'communication', 'emotional_tone', 'speech'
            ]
        }

    def create_recommendations_table(self, df):
        """Создает таблицу рекомендаций: каждая строка — отдельный звонок администратора, ссылки кликабельны и ведут на /api/transcriptions/<filename>/download"""
        if df.empty or 'administrator' not in df.columns:
            return [["Нет данных для отображения"]]

        def clean_text(text):
            if not text or text == "—":
                return "—"
            import re
            text = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)
            text = re.sub(r'\.{3,}', '...', text)
            return text

        def get_transcript_url(filename):
            import os
            API_BASE_URL = os.getenv("API_BASE_URL")
            if not filename or filename == "—":
                return None, None
            # Если это уже путь, извлекаем только имя файла
            import os
            # fname = os.path.basename(filename)
            url = API_BASE_URL + f"/transcriptions/{filename}/download"
            return url

        headers = ["Администратор", "Транскрибация", "Запись", "Рекомендации"]
        data = [headers]

        for admin, group in df.groupby('administrator'):
            admin = clean_text(admin)
            for _, row in group.iterrows():
                transcript = clean_text(row.get('filename_transcription', '') or "—")
                call_link = clean_text(row.get('call_link', '') or "—")
                # Формируем ссылку на транскрибацию
                transcript_url = get_transcript_url(transcript)
                if transcript_url:
                    transcript_cell = f'<a href="{transcript_url}">{transcript_url}</a>'
                else:
                    transcript_cell = "—"
                if call_link != "—":
                    call_link_cell = f'<a href="{call_link}">{call_link}</a>'
                else:
                    call_link_cell = "—"
                # Рекомендации по звонку
                recs = row.get('recommendations', [])
                if isinstance(recs, str):
                    recs = [recs]
                recs = [clean_text(r) for r in recs if r and r != "—"]
                # Форматируем рекомендации списком
                if recs:
                    rec_text = "\n".join(f"• {r}" for r in recs)
                else:
                    rec_text = "—"
                data.append([admin, transcript_cell, call_link_cell, rec_text])
        return data

    async def get_clinic_name(self, clinic_id):
        """
        Получение названия клиники из базы данных MongoDB
        """
        if not self.client:
            await self.connect_db()
        
        try:
            # Получаем коллекцию clinics
            clinics_collection = self.db["clinics"]
            
            # Ищем клинику по ID
            clinic = await clinics_collection.find_one({"client_id": clinic_id})
            
            if clinic and "name" in clinic:
                return clinic["name"]
            return None
        except Exception as e:
            logger.error(f"Ошибка при получении названия клиники: {e}")
            return None

    async def generate_report(self, start_date_str, end_date_str, clinic_id=None):
        """
        Генерация отчета по звонкам
        """
        try:
            # parse dates
            start = datetime.strptime(start_date_str, "%d.%m.%Y")
            end = datetime.strptime(end_date_str, "%d.%m.%Y")
            end = end.replace(hour=23, minute=59, second=59)
            
            # Получаем название клиники, если указан clinic_id
            clinic_name = None
            if clinic_id:
                clinic_name = await self.get_clinic_name(clinic_id)

            try:
                data = await self.get_calls_data(start, end, clinic_id)
                df = self.create_dataframe(data)
            except Exception as e:
                logger.error(f"Ошибка при получении или обработке данных: {e}")
                return None
            finally:
                await self.close_db()

            if df.empty:
                logger.warning("Нет данных для создания отчета")
                return None

            # prepare output file
            try:
                os.makedirs(REPORTS_DIR, exist_ok=True)
                suffix = f"_clinic_{clinic_id}" if clinic_id else ""
                period = f"_{start_date_str.replace('.','_')}-{end_date_str.replace('.','_')}"
                fn = os.path.join(REPORTS_DIR, f"call_report{suffix}{period}.pdf")

                await self._create_pdf_report(df, fn, clinic_name, start_date_str, end_date_str)
                return fn
            except Exception as e:
                logger.error(f"Ошибка при создании PDF отчета: {e}")
                import traceback
                logger.error(traceback.format_exc())
                return None
        except Exception as e:
            logger.error(f"Общая ошибка в generate_report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    async def _create_pdf_report(self, df: pd.DataFrame, out_path: str, clinic_name=None, start_date=None, end_date=None):
        """
        Создание PDF отчета с графиками и статистикой
        """
        try:
            doc = SimpleDocTemplate(out_path, pagesize=A4,
                                    rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
            elements = []
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                    fontName=self.font_bold, alignment=1, fontSize=18)
            subtitle_style = ParagraphStyle("Subtitle", parent=styles["Heading2"],
                    fontName=self.font_bold, alignment=1, fontSize=16)
            heading = ParagraphStyle("H2", parent=styles["Heading2"],
                    fontName=self.font_bold, fontSize=14)
            normal = ParagraphStyle("N", parent=styles["Normal"],
                                fontName=self.font_name, fontSize=8)

            # Переместили справку по метрикам в конец титульной страницы
            metrics_help = (
                "<b>FG%</b> — средний балл по всем оценкам звонков (0-10, умноженный на 10).<br/>"
                "<b>Конверсия</b> — доля звонков, завершившихся записью.<br/>"
                "<b>Оценки по критериям</b> — баллы по ключевым этапам разговора (см. таблицы ниже).<br/>"
                "<b>Тип звонка</b> — классификация: первичка, вторичка, подтверждение и др.<br/>"
                "<b>Администратор</b> — сотрудник, принявший звонок.<br/>"
            )

            # СТРАНИЦА 1 - ТИТУЛЬНАЯ
            # ======================
            
            # Название клиники, если есть (временно отключено)
            # if clinic_name:
            #     elements.append(Paragraph(f"{clinic_name}", subtitle_style))
            #     elements.append(Spacer(1, 10))

            # Обработка логотипа
            try:
                logo_path = "app/resources/logo.png"
                if os.path.exists(logo_path):
                    logo_img = Image(logo_path)
                    logo_width = 300
                    logo_height = logo_width * 517 / 837
                    logo_img.drawHeight = logo_height
                    logo_img.drawWidth = logo_width
                    elements.append(logo_img)
                else:
                    logger.warning(f"Файл логотипа {logo_path} не найден")
                elements.append(Spacer(1, 30))  # Увеличиваем отступ после логотипа
            except Exception as e:
                logger.error(f"Ошибка при добавлении логотипа: {e}")
                elements.append(Spacer(1, 30))

            # Title
            elements.append(Paragraph(f"Отчет по звонкам: {datetime.now():%d.%m.%Y}", title_style))
            elements.append(Spacer(1, 25))  # Увеличиваем отступ
                
            # Переместили справку по метрикам под таблицу "Сводная статистика"
            

            # Сводная статистика
            elements.append(Paragraph("Сводная статистика", heading))
            elements.append(Spacer(1, 10))
            
            summary_data = self.create_summary_statistics(df, start_date, end_date)
            summary_table = Table(summary_data, colWidths=[250, 250])
            
            summary_table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), self.font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ])
            
            summary_table.setStyle(summary_table_style)
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
            
            # Добавляем справку по метрикам после таблицы "Сводная статистика"
            elements.append(Paragraph("Справка по метрикам отчёта:", heading))
            elements.append(Paragraph(metrics_help, normal))
            elements.append(Spacer(1, 20))

            # СТРАНИЦА 2 - ТАБЛИЦА АДМИНИСТРАТОРОВ И ТАБЛИЦА ОЦЕНОК ПО НЕДЕЛЯМ
            # ================================================================
            elements.append(PageBreak())

            # Таблица администраторов - теперь первая на странице
            elements.append(Paragraph("Таблица администраторов", heading))
            elements.append(Spacer(1, 10))

            admin_table_data = self.create_admin_table(df)
            admin_table = Table(admin_table_data, colWidths=[250, 100, 100])
            admin_table.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, -1), self.font_name),
            ]))
            elements.append(admin_table)
            elements.append(Spacer(1, 30))  # Увеличенный отступ

            # Сводная таблица оценок по неделям - теперь вторая
            elements.append(Paragraph("Сводная таблица оценок по неделям", heading))
            elements.append(Spacer(1, 10))

            scores_table_data = self.create_weekly_scores_table(df)
            for call_type, table_data in scores_table_data:
                # Добавляем заголовок для каждой таблицы
                elements.append(Paragraph(f"Тип звонка: {call_type}", heading))
                elements.append(Spacer(1, 5))
                
                # Определяем ширины колонок с учетом специфики содержимого
                col_count = len(table_data[0])
                first_col_width = 45  # Ширина для номера недели
                
                # Создаем словарь с коэффициентами для каждой колонки
                column_factors = {}
                
                # Заполняем словарь для всех заголовков таблицы
                for i, header in enumerate(table_data[0][1:], 1):
                    # Устанавливаем базовый коэффициент 1.0 для всех колонок
                    column_factors[i] = 1.0
                    
                    # Изменяем коэффициенты для определенных колонок
                    if 'Цена ОТ' in header:
                        column_factors[i] = 0.7  # Уменьшаем ширину колонки Цена ОТ
                    elif 'Речь' in header:
                        column_factors[i] = 0.7  # Уменьшаем ширину колонки Речь
                    elif 'Экспертность' in header:
                        column_factors[i] = 1.25  # Увеличиваем ширину колонки Экспертность
                    elif 'Эмоциональный' in header:
                        column_factors[i] = 1.4  # Увеличиваем ширину колонки Эмоциональный окрас
                
                # Рассчитываем сумму коэффициентов
                total_factor = sum(column_factors.values())
                
                # Вычисляем базовую ширину колонки
                base_width = (500 - first_col_width) / total_factor
                
                # Создаем массив ширин колонок
                col_widths = [first_col_width]
                for i in range(1, col_count):
                    width = base_width * column_factors.get(i, 1.0)
                    col_widths.append(width)
                
                # Оборачиваем все строки в Paragraph для корректного отображения переносов
                table_wrapped = self.wrap_table(table_data, normal)
                table = Table(table_wrapped, colWidths=col_widths)
                
                # Стиль таблицы
                table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Центрирование всего содержимого
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Вертикальное центрирование
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Границы всех ячеек
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Жирный шрифт для заголовков
                    ('FONTSIZE', (0, 0), (-1, -1), 4),  # Размер шрифта для всей таблицы
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Фон заголовков
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Цвет текста
                    ('TOPPADDING', (0, 0), (-1, -1), 3),  # Отступ сверху
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),  # Отступ снизу
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),  # Отступ слева
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),  # Отступ справа
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 10))  # Пространство между таблицами

            # СТРАНИЦА 3 - ГРАФИКИ АДМИНИСТРАТОРОВ
            # ===================================
            elements.append(PageBreak())
            
            # График количества звонков по администраторам
            elements.append(Paragraph("Количество звонков по администраторам", heading))
            elements.append(Spacer(1, 10))
            
            calls_by_admin_img = self.create_calls_by_admin_chart(df)
            if calls_by_admin_img:
                elements.append(Image(calls_by_admin_img, width=450, height=270))
                elements.append(Spacer(1, 20))
            
            # График длительности звонков по администраторам
            elements.append(Paragraph("Длительность звонков по администраторам", heading))
            elements.append(Spacer(1, 10))
            
            duration_by_admin_img = self.create_duration_by_admin_chart(df)
            if duration_by_admin_img:
                elements.append(Image(duration_by_admin_img, width=450, height=270))
                elements.append(Spacer(1, 20))
            
            # СТРАНИЦА 4 - ГРАФИКИ КОНВЕРСИИ
            # ==============================
            elements.append(PageBreak())
            elements.append(Paragraph("Конверсия трафик", heading))
            elements.append(Spacer(1, 10))
            
            traffic_conv_img = self.create_traffic_conversion_chart(df)
            if traffic_conv_img:
                elements.append(Image(traffic_conv_img, width=350, height=250))
                elements.append(Spacer(1, 20))

            elements.append(Paragraph("Конверсия тип звонка", heading))
            elements.append(Spacer(1, 10))
            
            calltype_conv_img = self.create_call_type_conversion_chart(df)
            if calltype_conv_img:
                elements.append(Image(calltype_conv_img, width=350, height=250))
                elements.append(Spacer(1, 20))
            
            # СТРАНИЦА 5 - ЛИНЕЙНЫЕ ГРАФИКИ
            # =============================
            elements.append(PageBreak())
            elements.append(Paragraph("Звонки по типу", heading))
            elements.append(Spacer(1, 10))
            
            call_types_img = self.create_call_types_line_chart(df)
            if call_types_img:
                elements.append(Image(call_types_img, width=450, height=270))
                elements.append(Spacer(1, 20))
            
            elements.append(Paragraph("Динамика звонков по источникам трафика", heading))
            elements.append(Spacer(1, 10))
            
            traffic_line_img = self.create_traffic_line_chart(df)
            if traffic_line_img:
                elements.append(Image(traffic_line_img, width=450, height=270))
                elements.append(Spacer(1, 20))
            
            # СТРАНИЦА 6 - КРУГОВАЯ ДИАГРАММА
            # ===============================
            elements.append(PageBreak())
            elements.append(Paragraph("Распределение звонков по источникам трафика", heading))
            elements.append(Spacer(1, 10))
            
            traffic_pie_img = self.create_traffic_pie_chart(df)
            if traffic_pie_img:
                elements.append(Image(traffic_pie_img, width=450, height=450))
                elements.append(Spacer(1, 20))
            
            # СТРАНИЦА 7 - ТЕПЛОВАЯ КАРТА И ОЦЕНКИ ПО НЕДЕЛЯМ
            # ==============================================
            elements.append(PageBreak())
            elements.append(Paragraph("Тепловая карта звонков по дням недели и времени суток", heading))
            elements.append(Spacer(1, 10))
            
            calls_heatmap_img = self.create_calls_heatmap(df)
            if calls_heatmap_img:
                elements.append(Image(calls_heatmap_img, width=450, height=270))
                elements.append(Spacer(1, 30))
            
            elements.append(Paragraph("Оценки по неделям", heading))
            elements.append(Spacer(1, 10))
            
            overall_scores_img = self.create_overall_scores_by_week_chart(df)
            if overall_scores_img:
                elements.append(Image(overall_scores_img, width=450, height=270))
                elements.append(Spacer(1, 20))
            
            
            # Вставка графиков по типам звонков, размещаем Первичку и Перезвон на одной странице
            charts = self.create_scores_by_week_chart(df)
            first_chart = True
            prev_call_type = None
            
            for idx, (call_type, img_data) in enumerate(charts):
                # Добавляем PageBreak только перед первым графиком или если это не Первичка/Перезвон
                if first_chart:
                    elements.append(PageBreak())
                    first_chart = False
                elif prev_call_type == 'первичка' and call_type != 'перезвон':
                    elements.append(PageBreak())
                elif prev_call_type == 'перезвон':
                    elements.append(PageBreak())
                elif prev_call_type and prev_call_type != 'первичка' and call_type != 'первичка' and call_type != 'перезвон':
                    elements.append(PageBreak())
                
                # Добавляем увеличенный отступ между графиками на одной странице
                if not first_chart and prev_call_type == 'первичка' and call_type == 'перезвон':
                    elements.append(Spacer(1, 30))
                
                elements.append(Paragraph(f"Оценки по неделям — {call_type.capitalize()}", heading))
                elements.append(Image(img_data, width=7*inch, height=3.5*inch))
                
                prev_call_type = call_type
            
            # --- ВСТАВКА ТАБЛИЦЫ РЕКОМЕНДАЦИЙ В САМЫЙ КОНЕЦ ---
            try:
                elements.append(PageBreak())
                elements.append(Paragraph("Рекомендации для администраторов", heading))
                elements.append(Spacer(1, 10))
                recommendations_data = self.create_recommendations_table(df)
                recommendations_data = self.wrap_table(recommendations_data, normal)
                rec_table = Table(
                    recommendations_data,
                    colWidths=[70, 70, 150, 210],
                    repeatRows=1,
                    splitByRow=True
                )
                rec_table.setStyle(TableStyle([
                    ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
                    ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                    ("FONTNAME", (0,0), (-1,-1), self.font_name),
                    ("ALIGN", (0,0), (-1,-1), "LEFT"),
                    ("VALIGN", (0,0), (-1,-1), "TOP"),
                    ("WORDWRAP", (0,0), (-1,-1), True),
                    ("SPLITLONGWORDS", (0,0), (-1,-1), True),
                    ("FONTSIZE", (0,0), (-1,0), 8),
                    ("FONTSIZE", (0,1), (-1,-1), 6),
                    ("LEADING", (0,0), (-1,-1), 7),
                    ("TOPPADDING", (0,0), (-1,-1), 2),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 2),
                    ("LEFTPADDING", (0,0), (-1,-1), 2), 
                    ("RIGHTPADDING", (0,0), (-1,-1), 2),
                ]))
                elements.append(rec_table)
                elements.append(Spacer(1, 20))
            except Exception as e:
                logger.error(f"Ошибка при создании таблицы рекомендаций: {e}")
                elements.append(Paragraph("Не удалось создать таблицу рекомендаций", normal))

            # --- ДОБАВЛЯЕМ ОБЩИЕ РЕКОМЕНДАЦИИ КЛИНИКЕ ---
            try:
                client_id = df['client_id'].iloc[0] if not df.empty and 'client_id' in df.columns else None
                if client_id and start_date and end_date:
                    logger.info(f"Запрос общих рекомендаций для client_id={client_id} за период {start_date} - {end_date}")
                    service = RecommendationAnalysisService(client_id=client_id)
                    
                    start_date_obj = datetime.strptime(start_date, "%d.%m.%Y").date()
                    end_date_obj = datetime.strptime(end_date, "%d.%m.%Y").date()

                    analysis_result = await service.analyze_recommendations_for_period(
                        start_date=start_date_obj,
                        end_date=end_date_obj
                    )

                    # Извлекаем только итоговый вывод
                    overall_summary = analysis_result.get("overall_summary")

                    if overall_summary:
                        # Убираем префикс "Итог: "
                        summary_text = overall_summary.replace("Итог:", "").strip()

                        # Добавляем заголовок
                        elements.append(Paragraph("Общие Рекомендации по клинике", heading))
                        elements.append(Spacer(1, 10))

                        # Оборачиваем текст в параграф с автоматическим переносом
                        wrapped_summary = Paragraph(summary_text.replace('\n', '<br/>'), normal)
                        elements.append(wrapped_summary)
                        elements.append(Spacer(1, 20))
                    else:
                        logger.warning("Анализ рекомендаций не вернул текст для отчета (отсутствуют 'summary_points' и 'overall_summary').")
                else:
                    logger.warning("Недостаточно данных для запроса общих рекомендаций (client_id, start_date, end_date).")

            except Exception as e:
                logger.error(f"Ошибка при создании общих рекомендаций: {e}")
                import traceback
                logger.error(traceback.format_exc())
                elements.append(Paragraph("Не удалось создать блок общих рекомендаций.", normal))

            # Строим документ
            doc.build(elements)
            
            logger.info(f"Отчет сохранен в файл {out_path}")
        except Exception as e:
            logger.error(f"Общая ошибка в _create_pdf_report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            raise

    def _create_summary_excel(self, report_data: Dict[str, Any], output_path: str, df) -> None:
        """
        Создает Excel-отчет с сводной статистикой
        :param report_data: Данные для отчета
        :param output_path: Путь для сохранения Excel-файла
        :param df: DataFrame с данными звонков
        """
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        import os

        API_BASE_URL = os.getenv("API_BASE_URL")
        # Создаем рабочую книгу и листы
        workbook = Workbook()
        
        # Основной лист со сводной информацией
        ws_summary = workbook.active
        ws_summary.title = "Сводная информация"
        
        # Лист с данными по администраторам
        ws_admins = workbook.create_sheet("Администраторы")
        
        # Лист с данными по источникам
        ws_sources = workbook.create_sheet("Источники")
        
        # Лист с данными по типам звонков
        ws_call_types = workbook.create_sheet("Типы звонков")
        
        # Лист с оценками по неделям
        ws_weekly_scores = workbook.create_sheet("Оценки по неделям")
        
        # Лист с рекомендациями
        ws_recommendations = workbook.create_sheet("Рекомендации")
        
        # Импортируем необходимые стили для форматирования
        from openpyxl.styles import PatternFill, Alignment, Border, Side
        
        # Создаем стили для ячеек
        header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зеленый цвет для заголовков
        data_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")   # Голубой цвет для данных
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Граница для ячеек
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 1. Заполняем лист со сводной информацией
        ws_summary.append(["Сводный отчет по звонкам"])
        ws_summary.append(["Клиника:", report_data.get("clinic_name", "")])
        ws_summary.append(["Период:", f"{report_data.get('start_date', '')} - {report_data.get('end_date', '')}" ])
        ws_summary.append(["Дата создания:", report_data.get("generated_at", "")])
        ws_summary.append([])
        
        # Основная статистика в формате двухколоночной таблицы с заголовками
        # Добавляем заголовки
        ws_summary.append(["Показатель", "Значение"])
        row = 6  # Текущая строка
        # Применяем форматирование к заголовкам
        for col in range(1, 3):
            cell = ws_summary.cell(row=row, column=col)
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        # Данные
        stat_rows = [
            ["Общее количество звонков", report_data.get("total_calls", 0)],
            ["Входящие звонки", report_data.get("incoming_calls", 0)],
            ["Исходящие звонки", report_data.get("outgoing_calls", 0)]
        ]
        
        # Добавляем конверсию, если данные доступны
        if "conversion_percentage" in report_data:
            stat_rows.append(["Общая конверсия (%)", f"{report_data.get('conversion_percentage', 0)}%"])
            
        # Добавляем FG%, если данные доступны
        if "fg_percentage" in report_data:
            stat_rows.append(["Средний FG% (0-100)", f"{report_data.get('fg_percentage', 0)}%"])
        
        # Добавляем среднюю скорость обработки, если данные доступны
        if "avg_processing_speed" in report_data:
            stat_rows.append(["Средняя скорость обработки (мин)", report_data.get("avg_processing_speed", 0)])
        
        # Добавляем дополнительные данные с правильным форматированием
        stat_rows.extend([
            ["Средняя длительность звонка", report_data.get("avg_duration", "0:00")],
            ["Общая длительность всех звонков", report_data.get("total_duration", "0:00")]
        ])
        
        # Добавляем данные и форматируем
        for stat_row in stat_rows:
            ws_summary.append(stat_row)
            row += 1
            # Заголовок в первой колонке (зеленый фон)
            cell = ws_summary.cell(row=row, column=1)
            cell.fill = header_fill
            cell.border = thin_border
            # Значение во второй колонке (синий фон)
            cell = ws_summary.cell(row=row, column=2)
            cell.fill = data_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # 2. Заполняем лист с данными по администраторам
        ws_admins.append(["Статистика по администраторам"])
        ws_admins.append([
            "№", "Администратор", "Всего звонков", "Входящие", "Исходящие", 
            "Конверсия входящих (%)", "Конверсия исходящих (%)", "Средний FG%"
        ])
        
        # Форматируем заголовки
        for col in range(1, 9):
            cell = ws_admins.cell(row=2, column=col)
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        # Добавляем данные администраторов
        for i, admin in enumerate(report_data.get("administrators", [])):
            conversion_incoming = admin.get("fg_percentage", 0) if admin.get("incoming_calls", 0) > 0 else 0
            conversion_outgoing = admin.get("fg_percentage", 0) if admin.get("outgoing_calls", 0) > 0 else 0
            
            ws_admins.append([
                i+1,  # Номер строки
                admin.get("name", ""),
                admin.get("total_calls", 0),
                admin.get("incoming_calls", 0),
                admin.get("outgoing_calls", 0),
                conversion_incoming,
                conversion_outgoing,
                admin.get("fg_percentage", 0)
            ])
            
            # Форматируем строку
            row = i + 3  # 3 - начальный индекс строк с данными
            for col in range(1, 9):
                cell = ws_admins.cell(row=row, column=col)
                if col == 2:  # Имя администратора без особого форматирования
                    cell.border = thin_border
                else:
                    cell.fill = data_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
        
        # 3. Заполняем лист с данными по источникам
        ws_sources.append(["Статистика по источникам трафика"])
        ws_sources.append([
            "№", "Источник", "Количество звонков", "Процент от общего", 
            "Входящие", "Исходящие", "Конверсия (%)"
        ])
        
        # Форматируем заголовки
        for col in range(1, 8):
            cell = ws_sources.cell(row=2, column=col)
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        # Добавляем данные по источникам
        for i, source in enumerate(report_data.get("sources", [])):
            row_data = [
                i+1,  # Номер строки
                source.get("source", ""),
                source.get("count", 0),
                source.get("percentage", 0) / 100,  # Явно делим на 100 для формата процентов
                source.get("incoming_calls", 0),
                source.get("outgoing_calls", 0)
            ]
            
            # Добавляем конверсию, если она есть
            if "conversion_percentage" in source:
                row_data.append(source.get('conversion_percentage', 0))
            else:
                row_data.append(0)
                
            # Добавляем строку
            row = ws_sources.append(row_data)
            row_idx = len(ws_sources[ws_sources.dimensions])
            
            # Форматируем строку
            row = i + 3  # 3 - начальный индекс строк с данными
            for col in range(1, 8):
                cell = ws_sources.cell(row=row, column=col)
                if col == 2:  # Название источника без особого форматирования
                    cell.border = thin_border
                else:
                    cell.fill = data_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
            
            # Форматируем проценты (используем строки, а не значения)
            cell = ws_sources.cell(row=row, column=4)
            cell.number_format = '0.00%'
            
        # 4. Заполняем лист с данными по типам звонков
        ws_call_types.append(["Статистика по типам звонков"])
        ws_call_types.append([
            "№", "Тип звонка", "Количество звонков", "Процент от общего", 
            "Входящие", "Исходящие", "Конверсия (%)"
        ])
        
        # Форматируем заголовки
        for col in range(1, 8):
            cell = ws_call_types.cell(row=2, column=col)
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        # Добавляем данные по типам звонков
        for i, call_type in enumerate(report_data.get("call_types", [])):
            # Считаем реальный процент (доля от общего количества звонков)
            percentage = call_type.get("count", 0) / report_data.get("total_calls", 1) * 100 if report_data.get("total_calls", 0) > 0 else 0
            
            row_data = [
                i+1,  # Номер строки
                call_type.get("type", ""),
                call_type.get("count", 0),
                percentage / 100,  # Переводим в формат для Excel (0.75 для 75%)
                call_type.get("incoming_calls", 0),
                call_type.get("outgoing_calls", 0)
            ]
            
            # Добавляем конверсию, если она есть
            if "conversion_percentage" in call_type:
                row_data.append(call_type.get('conversion_percentage', 0))
            else:
                row_data.append(0)
                
            ws_call_types.append(row_data)
            
            # Форматируем строку
            row = i + 3  # 3 - начальный индекс строк с данными
            for col in range(1, 8):
                cell = ws_call_types.cell(row=row, column=col)
                if col == 2:  # Название типа без особого форматирования
                    cell.border = thin_border
                else:
                    cell.fill = data_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
            
            # Форматируем проценты
            cell = ws_call_types.cell(row=row, column=4)
            cell.number_format = '0.00%'
            
        # 5. Заполняем лист с оценками по неделям
        ws_weekly_scores.append(["Оценки по неделям"])
        ws_weekly_scores.append([])  # Пустая строка для разделения

        # Получаем данные таблиц
        weekly_tables = self.create_weekly_scores_table(df)
        current_row = 3  # Начинаем с третьей строки (после двух служебных)

        for call_type, table_data in weekly_tables:
            # Добавляем заголовок типа звонка
            ws_weekly_scores.cell(row=current_row, column=1, value=f"Тип звонка: {call_type}")
            ws_weekly_scores.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1

            # Добавляем заголовки и данные таблицы
            for row_idx, row in enumerate(table_data):
                for col_idx, value in enumerate(row, start=1):
                    # Если value — список (например, из-за wrap_table), превращаем в строку
                    if isinstance(value, list):
                        value = ' '.join(str(v) for v in value)
                    cell = ws_weekly_scores.cell(row=current_row + row_idx, column=col_idx)
                    cell.value = value
                    # Форматирование заголовков
                    if row_idx == 0:
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.font = Font(bold=True, size=10)
                    else:
                        cell.fill = data_fill
                        cell.alignment = center_alignment
                    cell.border = thin_border
            current_row += len(table_data)
            current_row += 1  # Пустая строка между таблицами

        # Автоматическая ширина колонок
        for column in ws_weekly_scores.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_weekly_scores.column_dimensions[column_letter].width = adjusted_width
        
        # 6. Заполняем лист с рекомендациями
        ws_recommendations.append(["Рекомендации для администраторов"])
        ws_recommendations.append([
            "№", "Администратор", "Транскрибация", "Запись", "Рекомендации"
        ])
        # Форматируем заголовки
        for col in range(1, 6):
            cell = ws_recommendations.cell(row=2, column=col)
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.font = Font(bold=True)
            cell.border = thin_border
        idx = 1
        for admin, group in df.groupby('administrator'):
            admin_name = str(admin) if admin else "—"
            for _, row in group.iterrows():
                transcript = row.get('filename_transcription', '—') or "—"
                call_link = row.get('call_link', '—') or "—"
                recs = row.get('recommendations', [])
                if isinstance(recs, str):
                    recs = [recs]
                recs = [str(r) for r in recs if r and r != "—"]
                # Формируем рекомендации с переносом строк и маркером
                if recs:
                    rec_lines = []
                    for r in recs:
                        wrapped = textwrap.fill(r, width=80)
                        rec_lines.append(f"• {wrapped}")
                    rec_text = "\n".join(rec_lines)
                else:
                    rec_text = "—"
                transcript_url = None
                transcript_cell = "—"
                if transcript != "—" and API_BASE_URL:
                    if API_BASE_URL.startswith("http://") or API_BASE_URL.startswith("https://"):
                        transcript_url = f"{API_BASE_URL}/transcriptions/{transcript}/download"
                        transcript_cell = textwrap.fill(transcript_url, width=50)
                call_link_cell = call_link
                if call_link != "—":
                    call_link_cell = textwrap.fill(call_link, width=50)
                ws_recommendations.append([
                    idx, admin_name, transcript_cell, call_link_cell, rec_text
                ])
                current_row = ws_recommendations.max_row
                # Назначаем гиперссылку на транскрибацию
                cell = ws_recommendations.cell(row=current_row, column=3)
                if transcript_url:
                    cell.value = transcript_cell
                    cell.hyperlink = transcript_url
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                else:
                    cell.value = "—"
                # Назначаем гиперссылку на запись
                cell = ws_recommendations.cell(row=current_row, column=4)
                if call_link != "—":
                    cell.value = call_link_cell
                    cell.hyperlink = call_link
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                else:
                    cell.value = "—"
                # Форматирование остальных колонок
                for col in range(1, 6):
                    cell = ws_recommendations.cell(row=current_row, column=col)
                    if col in [2, 5]:
                        cell.border = thin_border
                    elif col not in [3, 4]:
                        cell.fill = data_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border
                idx += 1
        # Ограничиваем ширину колонки "Рекомендации"
        ws_recommendations.column_dimensions['E'].width = 50
        # Ограничиваем ширину колонок "Транскрибация" и "Запись"
        ws_recommendations.column_dimensions['C'].width = 35
        ws_recommendations.column_dimensions['D'].width = 35

        # 7. Добавляем лист с общими рекомендациями
        ws_common_recs = workbook.create_sheet("Общие рекомендации")
        # Сбор рекомендаций по администраторам (аналогично PDF)
        rec_admins = dict()  # {рекомендация: set(админов)}
        admins = set()
        if 'administrator' in df.columns and 'recommendations' in df.columns:
            for _, row in df.iterrows():
                admin = row.get('administrator')
                admins.add(admin)
                recs = row.get('recommendations')
                if isinstance(recs, list):
                    for rec in recs:
                        if isinstance(rec, str) and rec.strip():
                            rec_admins.setdefault(rec.strip(), set()).add(admin)
                elif isinstance(recs, str) and recs.strip():
                    rec_admins.setdefault(recs.strip(), set()).add(admin)
        # Оставляем только рекомендации, которые встречаются у 2+ админов, либо все если админ один
        if len(admins) > 1:
            unique_recs = [rec for rec, adms in rec_admins.items() if len(adms) >= 2]
        else:
            unique_recs = list(rec_admins.keys())
        unique_recs.sort()
        # Заполняем таблицу
        ws_common_recs.append(["Рекомендация"])
        for rec in unique_recs:
            ws_common_recs.append([rec])
        # Оформление: ширина, шрифт, границы
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws_common_recs.iter_rows(min_row=1, max_row=ws_common_recs.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
        ws_common_recs.column_dimensions['A'].width = 80

        # Вставка таблиц по типам звонков:
        tables = self.create_weekly_scores_table(df)
        for call_type, table_data in tables:
            # ... (создать новый лист, вставить table_data, оформить) ...
            pass
        
        # Сохраняем файл
        workbook.save(output_path)

    async def prepare_report_data(self, df: pd.DataFrame, clinic_name: Optional[str] = None, 
                                start_date: Optional[str] = None, end_date: Optional[str] = None) -> Dict[str, Any]:
        """
        Подготавливает данные для Excel-отчета
        
        :param df: DataFrame с данными звонков
        :param clinic_name: Название клиники
        :param start_date: Начальная дата периода
        :param end_date: Конечная дата периода
        :return: Словарь с данными для отчета
        """
        # Логируем длительности звонков для отладки
        logger.debug(f"Длительности звонков (сек): {df['duration'].tolist()}")
        
        # Средняя длительность звонка в минутах и секундах
        avg_duration_sec = df['duration'].mean()
        avg_min = int(avg_duration_sec // 60)
        avg_sec = int(avg_duration_sec % 60)
        avg_duration_str = f"{avg_min}:{avg_sec:02d}"
        
        # Общая длительность всех звонков в минутах и секундах
        total_duration_sec = df['duration'].sum()
        total_min = int(total_duration_sec // 60)
        total_sec = int(total_duration_sec % 60)
        total_duration_str = f"{total_min}:{total_sec:02d}"
        
        logger.debug(f"Средняя длительность: {avg_duration_sec} сек = {avg_duration_str}")
        logger.debug(f"Общая длительность: {total_duration_sec} сек = {total_duration_str}")
        
        report_data = {
            "clinic_name": clinic_name or "",
            "start_date": start_date or "",
            "end_date": end_date or "",
            "generated_at": datetime.now().strftime('%d.%m.%Y %H:%M'),
            "total_calls": len(df),
            "incoming_calls": len(df[df['call_direction'] == 'Входящий']),
            "outgoing_calls": len(df[df['call_direction'] == 'Исходящий']),
            "avg_duration": avg_duration_str,
            "total_duration": total_duration_str,
            "avg_duration_sec": round(avg_duration_sec, 2),
            "total_duration_sec": round(total_duration_sec, 2),
        }
        
        # Расчет средней скорости обработки, если данные доступны
        try:
            if 'processing_speed' in df.columns:
                report_data["avg_processing_speed"] = round(df['processing_speed'].mean(), 2)
        except Exception as e:
            logger.error(f"Ошибка при расчете средней скорости обработки: {e}")
        
        # Расчет процента конверсии в запись
        try:
            if 'conversion' in df.columns:
                # Подсчитываем количество True значений
                if df['conversion'].dtype == 'bool':
                    converted_calls = df['conversion'].sum()
                else:
                    converted_calls = len(df[df['conversion'] == True])
                
                if report_data["total_calls"] > 0:
                    report_data["conversion_percentage"] = round((converted_calls / report_data["total_calls"]) * 100)
        except Exception as e:
            logger.error(f"Ошибка при расчете конверсии: {e}")
        
        # Расчет FG% по формуле: FG% = (Сумма всех оценок overall_score) / (Количество оцененных звонков) * 10
        if 'overall_score' in df.columns:
            total_overall_score = df['overall_score'].sum()
            scored_calls_count = df[df['overall_score'] > 0].shape[0]
            
            if scored_calls_count > 0:
                report_data["fg_percentage"] = round((total_overall_score / scored_calls_count) * 10)
        
        # Данные по администраторам
        admin_data = []
        for admin, group in df.groupby("administrator"):
            admin_info = {
                "name": admin,
                "total_calls": len(group),
                "incoming_calls": len(group[group['call_direction'] == 'Входящий']),
                "outgoing_calls": len(group[group['call_direction'] == 'Исходящий']),
                "avg_duration": round(group['duration'].mean() / 60, 2),
                "total_duration": round(group['duration'].sum() / 60, 2),
            }
            
            # Расчет конверсии для администратора
            try:
                if 'conversion' in group.columns:
                    if group['conversion'].dtype == 'bool':
                        converted_calls = group['conversion'].sum()
                    else:
                        converted_calls = len(group[group['conversion'] == True])
                    
                    if len(group) > 0:
                        admin_info["conversion_percentage"] = round((converted_calls / len(group)) * 100)
            except Exception as e:
                logger.error(f"Ошибка при расчете конверсии для администратора {admin}: {e}")
            
            # Расчет FG% для администратора по формуле: FG% = (Сумма всех оценок overall_score) / (Количество оцененных звонков) * 10
            if 'overall_score' in group.columns:
                total_score = group['overall_score'].sum()
                scored_calls = group[group['overall_score'] > 0].shape[0]
                
                if scored_calls > 0:
                    admin_info["fg_percentage"] = round((total_score / scored_calls) * 10)
            
            # Собираем рекомендации для администратора
            admin_recommendations = []
            if 'recommendations' in group.columns:
                unique_recommendations = set()
                for _, row in group.iterrows():
                    if isinstance(row.get('recommendations'), list):
                        for rec in row['recommendations']:
                            if isinstance(rec, str) and rec.strip() and rec not in unique_recommendations:
                                unique_recommendations.add(rec)
                    elif isinstance(row.get('recommendations'), str) and row['recommendations'].strip():
                        rec = row['recommendations']
                        if rec not in unique_recommendations:
                            unique_recommendations.add(rec)
                admin_recommendations = list(unique_recommendations)
            
            # Добавляем информацию о транскрипции и ссылке на запись
            # Используем первую запись для этого администратора
            first_call = group.iloc[0] if not group.empty else None
            if first_call is not None:
                admin_info["transcription"] = first_call.get('filename_transcription', '—')
                admin_info["call_link"] = first_call.get('call_link', '—')
            else:
                admin_info["transcription"] = "—"
                admin_info["call_link"] = "—"
                
            admin_info["recommendations"] = admin_recommendations
            admin_data.append(admin_info)
        
        report_data["administrators"] = admin_data
        
        # Данные по источникам
        source_data = []
        if 'source' in df.columns:
            source_counts = df['source'].value_counts()
            for source, count in source_counts.items():
                source_group = df[df['source'] == source]
                source_info = {
                    "source": source,
                    "count": count,
                    "percentage": round((count / len(df)) * 100),
                    "incoming_calls": len(source_group[source_group['call_direction'] == 'Входящий']),
                    "outgoing_calls": len(source_group[source_group['call_direction'] == 'Исходящий']),
                }
                
                # Расчет конверсии для источника
                try:
                    if 'conversion' in source_group.columns:
                        if source_group['conversion'].dtype == 'bool':
                            converted_calls = source_group['conversion'].sum()
                        else:
                            converted_calls = len(source_group[source_group['conversion'] == True])
                        
                        if len(source_group) > 0:
                            source_info["conversion_percentage"] = round((converted_calls / len(source_group)) * 100)
                except Exception as e:
                    logger.error(f"Ошибка при расчете конверсии для источника {source}: {e}")
                
                # Расчет FG% для источника по формуле: FG% = (Сумма всех оценок overall_score) / (Количество оцененных звонков) * 10
                if 'overall_score' in source_group.columns:
                    total_score = source_group['overall_score'].sum()
                    scored_calls = source_group[source_group['overall_score'] > 0].shape[0]
                    
                    if scored_calls > 0:
                        source_info["fg_percentage"] = round((total_score / scored_calls) * 10)
                
                source_data.append(source_info)
        
        report_data["sources"] = source_data
        
        # Данные по типам звонков
        call_type_data = []
        if 'call_type_classification' in df.columns:
            type_counts = df['call_type_classification'].value_counts()
            for call_type, count in type_counts.items():
                type_group = df[df['call_type_classification'] == call_type]
                type_info = {
                    "type": call_type,
                    "count": count,
                    "percentage": round((count / len(df)) * 100),
                    "incoming_calls": len(type_group[type_group['call_direction'] == 'Входящий']),
                    "outgoing_calls": len(type_group[type_group['call_direction'] == 'Исходящий']),
                }
                
                # Расчет конверсии для типа звонка
                try:
                    if 'conversion' in type_group.columns:
                        if type_group['conversion'].dtype == 'bool':
                            converted_calls = type_group['conversion'].sum()
                        else:
                            converted_calls = len(type_group[type_group['conversion'] == True])
                        
                        if len(type_group) > 0:
                            type_info["conversion_percentage"] = round((converted_calls / len(type_group)) * 100)
                except Exception as e:
                    logger.error(f"Ошибка при расчете конверсии для типа звонка {call_type}: {e}")
                
                # Расчет FG% для типа звонка по формуле: FG% = (Сумма всех оценок overall_score) / (Количество оцененных звонков) * 10
                if 'overall_score' in type_group.columns:
                    total_score = type_group['overall_score'].sum()
                    scored_calls = type_group[type_group['overall_score'] > 0].shape[0]
                    
                    if scored_calls > 0:
                        type_info["fg_percentage"] = round((total_score / scored_calls) * 10)
                
                call_type_data.append(type_info)
        elif 'metrics' in df.columns:
            # Если типы звонков находятся в metrics.call_type_classification
            df_copy = df.copy()
            df_copy['call_type_classification'] = df_copy['metrics'].apply(
                lambda x: x.get('call_type_classification', 'Неопределенный') if isinstance(x, dict) else 'Неопределенный'
            )
            
            type_counts = df_copy['call_type_classification'].value_counts()
            for call_type, count in type_counts.items():
                type_group = df_copy[df_copy['call_type_classification'] == call_type]
                type_info = {
                    "type": call_type,
                    "count": count,
                    "percentage": round((count / len(df_copy)) * 100),
                    "incoming_calls": len(type_group[type_group['call_direction'] == 'Входящий']),
                    "outgoing_calls": len(type_group[type_group['call_direction'] == 'Исходящий']),
                }
                
                # Расчет конверсии для типа звонка
                try:
                    if 'conversion' in type_group.columns:
                        if type_group['conversion'].dtype == 'bool':
                            converted_calls = type_group['conversion'].sum()
                        else:
                            converted_calls = len(type_group[type_group['conversion'] == True])
                        
                        if len(type_group) > 0:
                            type_info["conversion_percentage"] = round((converted_calls / len(type_group)) * 100)
                except Exception as e:
                    logger.error(f"Ошибка при расчете конверсии для типа звонка {call_type}: {e}")
                
                # Расчет FG% для типа звонка по формуле: FG% = (Сумма всех оценок overall_score) / (Количество оцененных звонков) * 10
                if 'overall_score' in type_group.columns:
                    total_score = type_group['overall_score'].sum()
                    scored_calls = type_group[type_group['overall_score'] > 0].shape[0]
                    
                    if scored_calls > 0:
                        type_info["fg_percentage"] = round((total_score / scored_calls) * 10)
                
                call_type_data.append(type_info)
        
        report_data["call_types"] = call_type_data
        
        # --- Формируем weekly_scores через create_weekly_scores_table(df) ---
        weekly_table = self.create_weekly_scores_table(df)
        report_data["weekly_scores"] = []
        if weekly_table and len(weekly_table) > 1:
            headers = weekly_table[0]
            for i, row in enumerate(weekly_table[1:], 1):
                # Преобразуем ключи в строки, если это список
                week_score = {(headers[j] if not isinstance(headers[j], list) else ' '.join(str(x) for x in headers[j])): row[j] for j in range(min(len(headers), len(row)))}
                # Для совместимости с Excel-отчетом добавляем стандартные ключи
                report_data["weekly_scores"].append({
                    "week_label": week_score.get("Неделя", ""),
                    "greeting": week_score.get("Приветствие,\nимя пациента", week_score.get("Приветствие", 0)),
                    "patient_name": week_score.get("Имя\nпациента", week_score.get("Имя пациента", 0)),
                    "needs_identification": week_score.get("Потребность", 0),
                    "service_presentation": week_score.get("Презентация\nуслуг", week_score.get("Презентация услуг", 0)),
                    "clinic_presentation": week_score.get("Презентация\nклиники", week_score.get("Презентация клиники", 0)),
                    "doctor_presentation": week_score.get("Презентация врача", 0),
                    "patient_booking": week_score.get("Запись пациента", 0),
                    "emotional_tone": week_score.get("Эмоциональный фон", 0),
                    "speech": week_score.get("Речь", 0),
                    "initiative": week_score.get("Инициатива", 0),
                    "overall_score": week_score.get("Средний балл", 0)
                })
        
        return report_data

    async def generate_excel_report(self, start_date_str: str, end_date_str: str, clinic_id: Optional[str] = None) -> Optional[str]:
        """
        Генерация Excel-отчета по звонкам
        :param start_date_str: Начальная дата в формате DD.MM.YYYY
        :param end_date_str: Конечная дата в формате DD.MM.YYYY
        :param clinic_id: ID клиники (опционально)
        :return: Путь к сгенерированному отчету или None в случае ошибки
        """
        try:
            # parse dates
            start = datetime.strptime(start_date_str, "%d.%m.%Y")
            end = datetime.strptime(end_date_str, "%d.%m.%Y")
            end = end.replace(hour=23, minute=59, second=59)
            
            # Получаем название клиники, если указан clinic_id
            clinic_name = None
            if clinic_id:
                clinic_name = await self.get_clinic_name(clinic_id)

            try:
                data = await self.get_calls_data(start, end, clinic_id)
                df = self.create_dataframe(data)
            except Exception as e:
                logger.error(f"Ошибка при получении или обработке данных: {e}")
                return None
            finally:
                await self.close_db()

            if df.empty:
                logger.warning("Нет данных для создания отчета")
                return None

            # Подготавливаем данные для отчета
            report_data = await self.prepare_report_data(df, clinic_name, start_date_str, end_date_str)
            
            # Prepare output file
            try:
                os.makedirs(REPORTS_DIR, exist_ok=True)
                suffix = f"_clinic_{clinic_id}" if clinic_id else ""
                period = f"_{start_date_str.replace('.','_')}-{end_date_str.replace('.','_')}"
                output_path = os.path.join(REPORTS_DIR, f"call_report{suffix}{period}.xlsx")
                
                self._create_summary_excel(report_data, output_path, df)
                return output_path
            except Exception as e:
                logger.error(f"Ошибка при создании Excel отчета: {e}")
                import traceback
                logger.error(traceback.format_exc())
                return None
        except Exception as e:
            logger.error(f"Общая ошибка в generate_excel_report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def create_scores_by_week_chart(self, df):
        import matplotlib.pyplot as plt
        from io import BytesIO
        import matplotlib.ticker as ticker
        results = []
        call_types_metrics = self.get_call_types_and_metrics()
        # Используем created_date вместо created_at
        if 'created_date' not in df.columns:
            return results
        # Проверяем, существуют ли поля week и year в данных
        if 'week' not in df.columns or 'year' not in df.columns:
            df['week'] = df['created_date'].dt.isocalendar().week
            df['year'] = df['created_date'].dt.year
        for call_type, metrics in call_types_metrics.items():
            df_type = df[df['call_type_classification'] == call_type]
            if df_type.empty:
                continue

            # Фильтруем метрики - оставляем только те, которые существуют в DataFrame
            available_metrics = [m for m in metrics if m in df_type.columns]
            if not available_metrics:
                logger.warning(f"Нет доступных метрик для типа звонка '{call_type}', пропускаем график")
                continue

            weekly = df_type.groupby(['year', 'week'])[available_metrics].mean().reset_index()
            if weekly.empty:
                continue
            # Унифицируем размер графика для всех типов звонков
            plt.figure(figsize=(11, 5))
            # Используем единую цветовую схему для всех графиков
            color_palette = sns.color_palette('pastel', n_colors=max(len(weekly), 2))  # Минимум 2 цвета для единообразия
            metric_labels = [self.get_criterion_display_name(m) for m in available_metrics]
            for idx, (_, week_row) in enumerate(weekly.iterrows()):
                y = [week_row[m] if m in week_row and pd.notnull(week_row[m]) else 0 for m in available_metrics]
                plt.plot(
                    metric_labels, y,
                    marker='o',
                    linewidth=1.2,
                    markersize=4,
                    label=f'Неделя {idx+1}',
                    color=color_palette[idx]
                )
            plt.grid(True, which='major', axis='y', color='#e0e0e0', linewidth=1, alpha=0.7)
            plt.grid(True, which='minor', axis='y', color='#f5f5f5', linewidth=0.7, alpha=0.5)
            plt.gca().set_axisbelow(True)
            plt.title(f'по неделям', fontsize=15, fontweight='normal', pad=15)
            plt.xlabel('Критерии', fontsize=11)
            plt.ylabel('Средний балл', fontsize=11)
            plt.xticks(rotation=35, ha='right', fontsize=8)
            plt.yticks(fontsize=8)
            # --- Ровная сетка по Y: всегда 0-10, шаг 1 ---
            ax = plt.gca()
            ax.set_ylim(0, 10)
            ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
            # --- Легенда ---
            legend_fontsize = 9
            # Унифицируем стиль легенды для всех графиков
            plt.legend(fontsize=legend_fontsize, loc='center right', bbox_to_anchor=(1.15, 0.5), frameon=False, handletextpad=0.5)
            plt.tight_layout(rect=[0, 0, 0.85, 1])
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
            img_data.seek(0)
            plt.close()
            results.append((call_type, img_data))
        return results

    def create_overall_scores_by_week_chart(self, df):
        import matplotlib.pyplot as plt
        from io import BytesIO
        import matplotlib.ticker as ticker
        if 'created_date' not in df.columns or 'overall_score' not in df.columns:
            return None
        # Проверяем, существуют ли поля week и year в данных
        if 'week' not in df.columns or 'year' not in df.columns:
            df['week'] = df['created_date'].dt.isocalendar().week
            df['year'] = df['created_date'].dt.year
        weekly = df.groupby(['year', 'week'])['overall_score'].mean().reset_index()
        if weekly.empty:
            plt.figure(figsize=(8, 3))
            plt.text(0.5, 0.5, 'Нет данных за период', ha='center', va='center', fontsize=16)
            plt.axis('off')
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
            img_data.seek(0)
            plt.close()
            return img_data
        min_date = df['created_date'].min()
        max_date = df['created_date'].max()
        plt.figure(figsize=(11, 5))  # Унифицируем размер с графиком Первички
        # Используем ту же цветовую схему, что и в графике Первички
        color_palette = sns.color_palette('pastel', n_colors=2)
        color = color_palette[0]  # Используем первый цвет из pastel палитры
        week_labels = [self.get_week_date_range_label(int(row['year']), int(row['week'])) for _, row in weekly.iterrows()]
        x = range(len(week_labels))
        plt.plot(x, weekly['overall_score'], marker='o', linewidth=1.2, markersize=4, color=color, label='Средний балл')
        for xi, y in zip(x, weekly['overall_score']):
            plt.text(xi, y + 0.1, f'{y:.2f}', ha='center', va='bottom', fontsize=10, fontweight='bold')
        plt.title('по неделям', fontsize=15, fontweight='normal', pad=15)  # Унифицируем заголовок
        plt.xlabel('Недели', fontsize=11)  # Унифицируем подписи осей
        plt.ylabel('Средний балл', fontsize=11)
        # Унифицируем стиль сетки с графиком Первички
        plt.grid(True, which='major', axis='y', color='#e0e0e0', linewidth=1, alpha=0.7)
        plt.grid(True, which='minor', axis='y', color='#f5f5f5', linewidth=0.7, alpha=0.5)
        plt.gca().set_axisbelow(True)
        # Унифицируем размер шрифта для подписей осей
        plt.xticks(x, week_labels, rotation=0, fontsize=8)
        plt.yticks(fontsize=8)
        # --- Ровная сетка по Y: всегда 0-10, шаг 1 ---
        ax = plt.gca()
        ax.set_ylim(0, 10)
        ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        # Унифицируем стиль легенды с графиком Первички
        legend_fontsize = 9
        plt.legend(fontsize=legend_fontsize, loc='center right', bbox_to_anchor=(1.15, 0.5), frameon=False, handletextpad=0.5)
        plt.tight_layout(rect=[0, 0, 0.85, 1])
        img_data = BytesIO()
        plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
        img_data.seek(0)
        plt.close()
        return img_data

    # --- ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ДЛЯ ОБЁРТКИ ЯЧЕЕК В PARAGRAPH ---
    def wrap_table(self, table, style):
        """
        Оборачивает все строки (кроме чисел и None) в Paragraph для корректного отображения в PDF.
        """
        from reportlab.platypus import Paragraph
        wrapped = []
        for row in table:
            new_row = []
            for cell in row:
                if isinstance(cell, (int, float)) or cell is None:
                    new_row.append(cell)
                else:
                    new_row.append(Paragraph(str(cell), style))
            wrapped.append(new_row)
        return wrapped
